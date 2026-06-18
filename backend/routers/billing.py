"""
WMS Kiwkiw - Router de Faturamento
Configurações de cobrança por seller e relatórios de faturamento.
"""

from typing import Optional, List
from datetime import date, datetime as _dt

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from ..database import get_db
from ..auth import get_current_user, require_manager_or_above, require_admin
from .. import models, schemas
from ..timezone_utils import end_of_day

router = APIRouter(prefix="/billing", tags=["Faturamento"])


@router.get("/config/{seller_id}", response_model=List[schemas.BillingConfigResponse])
def get_billing_config(
    seller_id: int,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Retorna configurações de cobrança de um seller."""
    return db.query(models.BillingConfig).filter(
        models.BillingConfig.seller_id == seller_id
    ).all()


@router.post("/config", response_model=schemas.BillingConfigResponse)
def upsert_billing_config(
    config: schemas.BillingConfigCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Cria ou atualiza configuração de cobrança (upsert por seller+chave)."""
    existing = db.query(models.BillingConfig).filter(
        models.BillingConfig.seller_id == config.seller_id,
        models.BillingConfig.config_key == config.config_key,
    ).first()

    if existing:
        existing.config_value = config.config_value
        existing.valid_from = config.valid_from
        existing.valid_to = config.valid_to
        db.commit()
        db.refresh(existing)
        return existing

    c = models.BillingConfig(**config.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@router.get("/report/{seller_id}")
def billing_report(
    seller_id: int,
    date_from: date = Query(...),
    date_to: date = Query(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Gera relatório de faturamento para um seller em um período.
    Calcula manuseio, caixas, seguro e armazenagem.
    """
    # Busca configurações de cobrança
    configs = db.query(models.BillingConfig).filter(
        models.BillingConfig.seller_id == seller_id
    ).all()
    config_dict = {c.config_key: c.config_value for c in configs}

    # Busca pedidos do período
    orders = db.query(models.Order).filter(
        models.Order.seller_id == seller_id,
        models.Order.imported_at >= _dt.combine(date_from, _dt.min.time()),
        models.Order.imported_at <= end_of_day(date_to),
        models.Order.for_billing == True,
        models.Order.status == "completed",
    ).all()

    total_orders = len(orders)
    unit_price = float(config_dict.get("Preço Unitário", 0))
    franchise = int(config_dict.get("Franquia", 1))
    franchise_qty = int(config_dict.get("Número Mínimo de Pedidos", 0))
    handling = float(config_dict.get("Manuseio", 0))
    storage = float(config_dict.get("Armazenagem", 0) if config_dict.get("Armazenagem Incluso", "0") != "0" else 0)

    # Calcula faturamento
    base = total_orders * unit_price
    franchise_value = 0
    if franchise and total_orders > franchise_qty:
        excess = total_orders - franchise_qty
        additional_price = float(config_dict.get("Preço Adicional", 0))
        franchise_value = excess * additional_price

    total = base + franchise_value + storage

    return {
        "seller_id": seller_id,
        "period_from": date_from.strftime("%d/%m/%Y"),
        "period_to": date_to.strftime("%d/%m/%Y"),
        "total_orders": total_orders,
        "unit_price": unit_price,
        "base_value": round(base, 2),
        "franchise_value": round(franchise_value, 2),
        "storage": round(storage, 2),
        "total": round(total, 2),
        "config": config_dict,
        "orders": [
            {
                "nf": o.nf_number,
                "customer": o.customer_name,
                "date": o.order_date.strftime("%d/%m/%Y"),
                "items": len(o.items),
            }
            for o in orders[:100]  # limita para não sobrecarregar
        ],
    }


@router.get("/export")
def export_billing_excel(
    date_from: date = Query(..., description="Data inicial (data de upload/importação)"),
    date_to: date   = Query(..., description="Data final (data de upload/importação)"),
    seller_id: Optional[int] = Query(None, description="Filtrar por seller (opcional)"),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Exporta relatório de faturamento em Excel (.xlsx).

    Colunas: Seller | Número NF | Data | Cód SKU | Quantidade | Entrada/Saída | Caixa
    Filtro: data de importação (imported_at) entre date_from e date_to.
    Caixa: usa sempre box_used (ajustado pelo operador) se preenchido,
           caso contrário tenta calcular via algoritmo de caixa.
    """
    import io
    from datetime import datetime as _dt
    from fastapi.responses import StreamingResponse

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl não instalado. Execute: pip install openpyxl")

    # ── Busca pedidos no período ──────────────────────────────────────────────
    from sqlalchemy.orm import joinedload
    from sqlalchemy import and_

    q = db.query(models.Order).options(
        joinedload(models.Order.seller),
        joinedload(models.Order.items),
    ).filter(
        and_(
            models.Order.for_billing == True,
            models.Order.imported_at >= _dt.combine(date_from, _dt.min.time()),
            models.Order.imported_at <= _dt.combine(date_to,   _dt.max.time()),
        )
    )
    if seller_id:
        q = q.filter(models.Order.seller_id == seller_id)

    orders = q.order_by(models.Order.imported_at, models.Order.seller_id, models.Order.nf_number).all()

    # ── Status label helper ──────────────────────────────────────────────────
    STATUS_LABELS = {
        "pending":     "A iniciar",
        "validated":   "A iniciar",
        "separating":  "A iniciar",
        "scanning":    "Em Bipagem",
        "completed":   "Finalizado",
        "interrupted": "Interrompido",
        "cancelled":   "Cancelado",
    }
    def _status_label(order) -> str:
        raw = order.status.value if hasattr(order.status, "value") else str(order.status)
        return STATUS_LABELS.get(raw, raw)

    # ── Box algorithm helper (inline) ────────────────────────────────────────
    # score = soma de (quantidade × box_type_do_produto) — o box_type do produto é seu "peso"
    def _get_box(order) -> str:
        if order.box_used:
            return order.box_used
        total_qty   = sum(i.quantity for i in order.items)
        total_score = 0
        for item in order.items:
            prod = db.query(models.Product).filter(
                models.Product.seller_id == order.seller_id,
                models.Product.sku       == item.sku,
            ).first()
            if prod and prod.box_type:
                try:
                    total_score += int(prod.box_type) * item.quantity
                except (ValueError, TypeError):
                    pass
        rule = db.query(models.BoxAlgorithm).filter(
            models.BoxAlgorithm.seller_id    == order.seller_id,
            models.BoxAlgorithm.num_products == total_qty,
            models.BoxAlgorithm.score        == total_score,
        ).first()
        return rule.box_type if rule else "N.A"

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faturamento"

    PURPLE  = "7B63E8"
    TEAL    = "3DD9A4"
    DARK    = "14122A"
    LIGHT   = "F0EEFF"
    WHITE   = "FFFFFF"
    GRAY    = "F5F5F5"

    header_font    = Font(name="Calibri", bold=True, color=WHITE, size=10)
    header_fill    = PatternFill("solid", fgColor=PURPLE)
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font      = Font(name="Calibri", size=9)
    alt_fill       = PatternFill("solid", fgColor=LIGHT)
    thin_border    = Border(
        left=Side(style="thin", color="D0CCEE"),
        right=Side(style="thin", color="D0CCEE"),
        top=Side(style="thin", color="D0CCEE"),
        bottom=Side(style="thin", color="D0CCEE"),
    )

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Relatório de Faturamento  |  {date_from.strftime('%d/%m/%Y')} a {date_to.strftime('%d/%m/%Y')}"
    title_cell.font  = Font(name="Calibri", bold=True, color=WHITE, size=12)
    title_cell.fill  = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header row
    headers = ["Seller", "Número NF", "Data", "Cód SKU", "Quantidade", "Entrada/Saída", "Caixa", "Status NF"]
    col_widths = [22, 14, 12, 18, 12, 14, 10, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    # Data rows — one row per OrderItem
    row_num = 3
    for order in orders:
        seller_name = order.seller.trade_name if order.seller else "—"
        nf          = order.nf_number or "—"
        date_str    = order.order_date.strftime("%d/%m/%Y") if order.order_date else "—"
        file_type   = "Saída" if (order.file_type and order.file_type.value in ("export", "saida", "Saída")) else "Entrada"
        box         = _get_box(order)
        fill        = PatternFill("solid", fgColor=LIGHT) if row_num % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        status_label = _status_label(order)
        status_fill_color = {
            "Finalizado":    "D1FAE5",   # verde claro
            "Interrompido":  "FEF3C7",   # âmbar claro
            "Em Bipagem":    "EDE9FE",   # lilás claro
            "Cancelado":     "FEE2E2",   # vermelho claro
        }.get(status_label, None)

        for item in order.items:
            row = [seller_name, nf, date_str, item.sku or "—", item.quantity, file_type, box, status_label]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font   = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col in (2, 5, 6, 7, 8) else "left",
                                           vertical="center")
                # Status column (8) gets color-coded; others get normal zebra
                if col == 8 and status_fill_color:
                    cell.fill = PatternFill("solid", fgColor=status_fill_color)
                    cell.font = Font(name="Calibri", size=9, bold=True,
                                     color={"D1FAE5":"065F46","FEF3C7":"92400E",
                                            "EDE9FE":"4C1D95","FEE2E2":"991B1B"}.get(status_fill_color,"000000"))
                else:
                    cell.fill = fill
            row_num += 1

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Summary row
    ws.cell(row=row_num, column=1, value=f"Total: {len(orders)} pedidos  |  {row_num - 3} itens").font = \
        Font(name="Calibri", bold=True, size=9, color=PURPLE)
    ws.row_dimensions[row_num].height = 16

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"faturamento_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
