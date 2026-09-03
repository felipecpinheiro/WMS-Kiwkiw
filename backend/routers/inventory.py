"""
WMS Kiwkiw - Router de Estoque
Consulta, exportação e movimentações de estoque.
"""

import csv
import io
import json
import tempfile
import os
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text

from ..database import get_db
from ..auth import get_current_user, require_manager_or_above, require_admin
from ..timezone_utils import now_brasilia, today_brasilia
from .. import models
from ..services.stock_manager import (
    get_stock_report, update_stock_position, get_sku_history, calculate_stock_level,
)

import os as _os

router = APIRouter(prefix="/inventory", tags=["Estoque"])

# Passphrase de edição de movimentações — SEMPRE de variável de ambiente
# (repositório é PÚBLICO). Em produção (Postgres) o app recusa subir sem ela;
# em dev local (SQLite) cai num valor fixo só de desenvolvimento.
_DATABASE_URL = _os.environ.get("DATABASE_URL", "")
_IS_PRODUCTION_DB = _DATABASE_URL.startswith("postgres")

EDIT_PASSPHRASE = _os.environ.get("WMS_EDIT_PASSPHRASE")
if not EDIT_PASSPHRASE:
    if _IS_PRODUCTION_DB:
        raise RuntimeError(
            "WMS_EDIT_PASSPHRASE não definida. Configure a variável de ambiente "
            "no serviço antes de subir em produção (Postgres)."
        )
    EDIT_PASSPHRASE = "k&ksmt$"  # só dev local


# ─────────────────────────────────────────────────────────
# VERIFICAÇÃO DE SENHA (resposta imediata, sem abrir janela de edição)
# ─────────────────────────────────────────────────────────

@router.post("/verify-passphrase")
def verify_passphrase(
    body: dict,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Verifica a senha de edição de movimentações imediatamente.
    Retorna 200 se correta, 403 se incorreta.
    Permite que o frontend barre o acesso SEM abrir a janela de edição.
    """
    passphrase = body.get("passphrase", "")
    if passphrase != EDIT_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Senha de edição incorreta")
    return {"valid": True}


# ─────────────────────────────────────────────────────────
# POSIÇÃO DE ESTOQUE
# ─────────────────────────────────────────────────────────

@router.get("/stock/{seller_id}")
def get_stock(
    seller_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Retorna posição de estoque de um seller.
    Sellers só acessam o próprio estoque.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")

    stock_positions = get_stock_report(seller_id, db)

    # ── Pré-carrega catálogo de produtos cadastrados para este seller ─────────
    # Fonte de verdade para product_name: tabela products (não o campo armazenado
    # na posição, que pode estar desatualizado ou preenchido só com o código SKU).
    products_catalog: dict[str, str] = {
        p.sku: p.name
        for p in db.query(models.Product).filter(
            models.Product.seller_id == seller_id,
            models.Product.active == True,
        ).all()
    }

    # ── Enriquece cada posição com previsão de dias (média 60d) ──────────────
    from datetime import datetime as _dt, timedelta as _td
    cutoff_60d = (_dt.now() - _td(days=60)).date()

    # Busca saídas dos últimos 60 dias — SQL raw para suportar valores legados.
    # CAST obrigatório: movement_type é ENUM nativo no PostgreSQL e comparar com
    # 'Saída' dispara InvalidTextRepresentation (o UPPER também não aceita enum).
    sales_60d = dict(
        db.execute(
            text("""
                SELECT sku, SUM(quantity) FROM stock_movements
                WHERE seller_id = :sid
                  AND UPPER(CAST(movement_type AS VARCHAR)) IN ('OUT','S','SAIDA','SAÍDA')
                  AND movement_date >= :cutoff
                GROUP BY sku
            """),
            {"sid": seller_id, "cutoff": str(cutoff_60d)},
        ).fetchall()
    )

    today_d = _dt.now().date()
    result = []
    for pos in stock_positions:
        sku = pos.get("sku") if isinstance(pos, dict) else getattr(pos, "sku", None)
        if not sku:
            result.append(pos)
            continue

        pos_dict = dict(pos) if isinstance(pos, dict) else pos.__dict__.copy()
        pos_dict.pop("_sa_instance_state", None)

        # ── Nome oficial vem do cadastro de Produtos (source of truth) ───────
        if sku in products_catalog:
            pos_dict["product_name"]       = products_catalog[sku]
            pos_dict["product_registered"] = True
        else:
            # SKU existe em movimentações mas não está cadastrado em Produtos
            pos_dict["product_registered"] = False
            # Mantém o que estiver salvo na posição (pode ser o próprio código)

        # ── Previsão de dias restantes ────────────────────────────────────────
        current       = pos_dict.get("current_stock", 0) or 0
        total_out_60d = sales_60d.get(sku) or 0

        if current <= 0:
            pos_dict["days_remaining"]  = 0
            pos_dict["forecast_status"] = "Sem Produto"
        elif total_out_60d == 0:
            pos_dict["days_remaining"]  = None
            pos_dict["forecast_status"] = "Sem Saídas 60d"
        else:
            avg_daily = total_out_60d / 60.0
            days_rem  = round(current / avg_daily)
            pos_dict["days_remaining"] = days_rem
            if days_rem < 30:
                pos_dict["forecast_status"] = "Baixo"
            elif days_rem <= 60:
                pos_dict["forecast_status"] = "Médio"
            else:
                pos_dict["forecast_status"] = "Alto"

        result.append(pos_dict)

    return result


# ─────────────────────────────────────────────────────────
# HISTÓRICO / GRÁFICO DE SKU
# ─────────────────────────────────────────────────────────

@router.get("/sku-history/{seller_id}/{sku}")
def sku_history(
    seller_id: int,
    sku: str,
    days: int = Query(90, ge=7, le=365),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Retorna histórico de movimentações de um SKU para o popup de análise.
    Inclui gráfico diário e estatísticas de média de consumo.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")

    return get_sku_history(seller_id=seller_id, sku=sku, db=db, days=days)


# ─────────────────────────────────────────────────────────
# MOVIMENTAÇÕES
# ─────────────────────────────────────────────────────────

# Mapa de normalização para valores legados ('IN'/'OUT') que podem estar no banco
_MT_NORMALIZE = {
    "IN":     "Entrada",
    "OUT":    "Saída",
    "E":      "Entrada",
    "S":      "Saída",
    "ENTRADA":"Entrada",
    "SAÍDA":  "Saída",
    "SAIDA":  "Saída",
}


@router.get("/movements/{seller_id}")
def get_movements(
    seller_id: int,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    sku: Optional[str] = None,
    movement_type: Optional[str] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Retorna histórico de movimentações de estoque com filtros opcionais.
    Usa SQL raw para evitar erros de conversão ORM quando o banco contém
    valores legados ('IN'/'OUT') no campo movement_type.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")

    # Monta query SQL diretamente — evita falha do ORM ao converter
    # valores legados ('IN','OUT') para o enum MovementType
    conditions = ["seller_id = :seller_id"]
    params: dict = {"seller_id": seller_id}

    if date_from:
        conditions.append("movement_date >= :date_from")
        params["date_from"] = str(date_from)
    if date_to:
        conditions.append("movement_date <= :date_to")
        params["date_to"] = str(date_to)
    if sku:
        conditions.append("sku LIKE :sku")
        params["sku"] = f"%{sku}%"
    if movement_type:
        # Normaliza o filtro e compara sobre o texto: movement_type é ENUM nativo
        # no PostgreSQL, então "movement_type = 'Entrada'" quebraria. As listas
        # cobrem os rótulos atuais (IN/OUT) e os legados gravados antes deles.
        mt_normalized = _MT_NORMALIZE.get(movement_type.upper(), movement_type)
        if mt_normalized == models.MovementType.IN.value:
            conditions.append("UPPER(CAST(movement_type AS VARCHAR)) IN ('IN','ENTRADA','E')")
        else:
            conditions.append("UPPER(CAST(movement_type AS VARCHAR)) IN ('OUT','SAIDA','SAÍDA','S')")

    where_clause = " AND ".join(conditions)
    sql = text(f"""
        SELECT id, seller_id, sku, product_name, movement_date, movement_type,
               quantity, adjusted_quantity, nf_number, nf_date, nature,
               order_id, session_id, observation, created_at
        FROM stock_movements
        WHERE {where_clause}
        ORDER BY movement_date DESC, id DESC
    """)

    rows = db.execute(sql, params).fetchall()

    result = []
    for r in rows:
        raw_mt = r.movement_type or ""
        # Normaliza valores legados para o formato canônico
        mt_display = _MT_NORMALIZE.get(raw_mt.upper(), raw_mt)
        result.append({
            "id":                r.id,
            "seller_id":         r.seller_id,
            "sku":               r.sku,
            "product_name":      r.product_name,
            "movement_date":     str(r.movement_date) if r.movement_date else None,
            "movement_type":     mt_display,
            "quantity":          r.quantity,
            "adjusted_quantity": r.adjusted_quantity,
            "nf_number":         r.nf_number,
            "nf_date":           str(r.nf_date) if r.nf_date else None,
            "nature":            r.nature,
            "order_id":          r.order_id,
            "session_id":        r.session_id,
            "observation":       r.observation,
            "created_at":        r.created_at if isinstance(r.created_at, str) else (r.created_at.isoformat() if r.created_at else None),
        })

    return result


@router.post("/movements/manual", status_code=201)
def create_manual_movement(
    body: dict,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Cria uma movimentação manual de estoque (entrada ou saída).
    Requer papel master ou superior.
    """
    seller_id = body.get("seller_id")
    sku = body.get("sku", "").strip()
    movement_type_str = body.get("movement_type", "")
    quantity = int(body.get("quantity", 0))
    movement_date_str = body.get("movement_date") or str(today_brasilia())
    observation = body.get("observation", "")
    nf_number = body.get("nf_number", "")
    product_name = body.get("product_name", sku)

    if not seller_id or not sku or not movement_type_str or quantity <= 0:
        raise HTTPException(
            status_code=422,
            detail="seller_id, sku, movement_type e quantity (>0) são obrigatórios",
        )

    # Sem esta checagem o INSERT viola a FK e o usuário recebe "Internal Server
    # Error" sem explicação. Não filtra `active`: só verifica que existe.
    if not db.query(models.Seller.id).filter(models.Seller.id == seller_id).first():
        raise HTTPException(status_code=404, detail=f"Seller {seller_id} não encontrado")

    # Normaliza tipo: aceita "Saida" / "saida" / "Saída" / "Entrada" etc.
    _type_norm = {
        'entrada': 'Entrada',
        'saida': 'Saída',
        'saída': 'Saída',
    }
    movement_type_str = _type_norm.get(movement_type_str.lower().strip(), movement_type_str)

    try:
        mt_enum = models.MovementType(movement_type_str)
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail=f"movement_type deve ser um de: {[e.value for e in models.MovementType]}",
        )

    try:
        mov_date = date.fromisoformat(movement_date_str)
    except ValueError:
        mov_date = today_brasilia()

    # ⚠️ SKU não diferencia caixa (28/08/2026): resolve contra o cadastro e passa
    # a usar a GRAFIA DO CADASTRO daqui pra frente. Sem isso, lançar 'MOSQ2' à
    # mão cria uma posição separada de 'mosq2' e parte o estoque do produto.
    prod = db.query(models.Product).filter(
        models.Product.seller_id == seller_id,
        func.lower(models.Product.sku) == sku.lower(),
    ).order_by(models.Product.active.desc(), models.Product.id).first()
    if prod:
        sku = prod.sku
        if not product_name or product_name == sku:
            product_name = prod.name

    movement = models.StockMovement(
        seller_id=seller_id,
        sku=sku,
        product_name=product_name,
        movement_date=mov_date,
        movement_type=mt_enum,
        quantity=quantity,
        adjusted_quantity=quantity,
        nf_number=nf_number or None,
        observation=observation or None,
        operator_id=current_user.id,
        created_at=now_brasilia(),
    )
    db.add(movement)

    update_stock_position(
        seller_id=seller_id,
        sku=sku,
        product_name=product_name,
        movement_type=mt_enum,
        quantity=quantity,
        db=db,
    )

    # ── Trilha de auditoria: registra lançamento manual ──────────────────────
    seller_obj = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    seller_label = seller_obj.trade_name if seller_obj else str(seller_id)
    db.add(models.AuditLog(
        entity_type="StockMovement",
        entity_id=None,  # será preenchido após commit se necessário
        action="MANUAL_MOVEMENT",
        detail=(
            f"Lançamento manual | Tipo: {movement_type_str} | SKU: {sku} | "
            f"Qtd: {quantity} | Seller: {seller_label} | "
            f"Data: {mov_date} | NF: {nf_number or '-'} | "
            f"Obs: {observation or '-'}"
        ),
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(movement)

    mt_val = (
        movement.movement_type.value
        if hasattr(movement.movement_type, "value")
        else movement.movement_type
    )
    return {
        "id": movement.id,
        "sku": movement.sku,
        "product_name": movement.product_name,
        "movement_type": mt_val,
        "quantity": movement.quantity,
        "movement_date": str(movement.movement_date),
        "observation": movement.observation,
    }


@router.post("/movements/bulk", status_code=201)
def create_bulk_movements(
    body: dict,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Importação em lote de movimentações de estoque.
    Muito mais eficiente que chamar /movements/manual N vezes:
    - Pré-carrega produtos e posições de estoque com 2 queries únicas
    - Atualiza posições em memória (sem query por linha)
    - Insere todos os movimentos com um único commit
    """
    seller_id = body.get("seller_id")
    rows = body.get("rows", [])

    if not seller_id:
        raise HTTPException(status_code=422, detail="seller_id é obrigatório")
    if not rows:
        raise HTTPException(status_code=422, detail="Nenhuma linha fornecida")

    _type_norm = {
        'entrada': 'Entrada',
        'saida': 'Saída',
        'saída': 'Saída',
    }

    # ── 1. Pré-carrega todos os produtos do seller em memória ─────────────────
    skus_in_batch = {r.get("sku", "").strip() for r in rows if r.get("sku", "").strip()}
    products_by_sku: dict[str, str] = {
        p.sku: p.name
        for p in db.query(models.Product).filter(
            models.Product.seller_id == seller_id,
            models.Product.sku.in_(skus_in_batch),
        ).all()
    }

    # ── 2. Pré-carrega todas as posições de estoque em memória ───────────────
    positions_by_sku: dict[str, models.StockPosition] = {
        p.sku: p
        for p in db.query(models.StockPosition).filter(
            models.StockPosition.seller_id == seller_id,
            models.StockPosition.sku.in_(skus_in_batch),
        ).all()
    }

    movements_to_add: list[models.StockMovement] = []
    ok = 0
    errors: list[str] = []
    now = now_brasilia()

    # ── 3. Processa cada linha sem tocar no banco ─────────────────────────────
    for i, row in enumerate(rows):
        try:
            sku = row.get("sku", "").strip()
            if not sku:
                errors.append(f"Linha {i + 1}: SKU vazio")
                continue

            raw_type = row.get("movement_type", "")
            mt_str = _type_norm.get(raw_type.lower().strip(), raw_type)
            try:
                mt_enum = models.MovementType(mt_str)
            except ValueError:
                errors.append(f"Linha {i + 1}: tipo inválido '{raw_type}'")
                continue

            quantity = int(row.get("quantity", 0))
            if quantity <= 0:
                errors.append(f"Linha {i + 1}: quantidade inválida")
                continue

            movement_date_str = row.get("movement_date") or str(today_brasilia())
            try:
                mov_date = date.fromisoformat(movement_date_str)
            except ValueError:
                mov_date = today_brasilia()

            product_name = products_by_sku.get(sku, sku)
            nf_number = row.get("nf_number") or None
            observation = row.get("observation") or None

            # Cria objeto de movimento (ainda não persiste)
            movements_to_add.append(models.StockMovement(
                seller_id=seller_id,
                sku=sku,
                product_name=product_name,
                movement_date=mov_date,
                movement_type=mt_enum,
                quantity=quantity,
                adjusted_quantity=quantity,
                nf_number=nf_number,
                observation=observation,
                operator_id=current_user.id,
                created_at=now,
            ))

            # Atualiza posição em memória (zero queries extras)
            if sku not in positions_by_sku:
                pos = models.StockPosition(
                    seller_id=seller_id,
                    sku=sku,
                    product_name=product_name,
                    initial_stock=0,
                    total_in=0,
                    total_out=0,
                    current_stock=0,
                )
                db.add(pos)
                positions_by_sku[sku] = pos

            pos = positions_by_sku[sku]
            if mt_enum == models.MovementType.IN:
                pos.total_in += quantity
            else:
                pos.total_out += quantity
            pos.current_stock = pos.initial_stock + pos.total_in - pos.total_out
            pos.level = calculate_stock_level(pos.current_stock)
            pos.updated_at = now
            if not pos.product_name:
                pos.product_name = product_name

            ok += 1

        except Exception as e:
            errors.append(f"Linha {i + 1}: {str(e)}")

    # ── 4. Um único insert em lote + commit ──────────────────────────────────
    if movements_to_add:
        db.bulk_save_objects(movements_to_add)

    # Trilha de auditoria resumida (1 registro para o lote inteiro)
    seller_obj = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    seller_label = seller_obj.trade_name if seller_obj else str(seller_id)
    db.add(models.AuditLog(
        entity_type="StockMovement",
        entity_id=None,
        action="BULK_IMPORT",
        detail=(
            f"Importação em lote | Seller: {seller_label} | "
            f"Linhas OK: {ok} | Erros: {len(errors)}"
        ),
        user_id=current_user.id,
    ))

    db.commit()

    return {"imported": ok, "errors": errors}


@router.put("/movements/{movement_id}")
def update_movement(
    movement_id: int,
    body: dict,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Edita uma movimentação existente.
    Requer papel admin + senha especial.
    Mudanças de TIPO e de QUANTIDADE são refletidas na posição de estoque:
    o efeito antigo do registro é desfeito e o novo é aplicado.
    """
    passphrase = body.get("passphrase", "")
    if passphrase != EDIT_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Senha de edição incorreta")

    movement = db.query(models.StockMovement).filter(
        models.StockMovement.id == movement_id
    ).first()
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")

    # Campos editáveis
    old_qty = movement.quantity
    old_type = movement.movement_type

    def _norm_type(value) -> str:
        """Normaliza enum/str para 'Entrada'/'Saída' (aceita legados 'IN'/'OUT')."""
        raw = value.value if hasattr(value, "value") else str(value or "")
        return _MT_NORMALIZE.get(raw.strip().upper(), raw)

    old_type_str = _norm_type(old_type)

    # ── Tipo (Entrada/Saída) ────────────────────────────────────────────────
    # ⚠️ Até 28/08/2026 este campo era LIDO (para escolher o balde da posição) e
    # NUNCA gravado: a tela mandava "Saída", o backend descartava em silêncio e
    # devolvia o tipo antigo na resposta — com o toast de sucesso mentindo.
    new_type_str = old_type_str
    if "movement_type" in body and body["movement_type"] is not None:
        candidate = _norm_type(body["movement_type"])
        try:
            new_type_str = models.MovementType(candidate).value
        except ValueError:
            raise HTTPException(
                status_code=422,
                detail=(
                    "movement_type deve ser um de: "
                    f"{[e.value for e in models.MovementType]}"
                ),
            )

    # ── Quantidade ──────────────────────────────────────────────────────────
    qty_informada = "quantity" in body and body["quantity"] is not None
    new_qty = old_qty
    if qty_informada:
        # int() cru virava 500 com qualquer texto ("abc"): o ValueError subia sem
        # tratamento. Pela tela e dificil chegar aqui (campo e type=number), mas
        # uma chamada forjada estourava o endpoint.
        try:
            new_qty = int(body["quantity"])
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=422, detail="Quantidade deve ser um numero inteiro"
            )
        if new_qty <= 0:
            raise HTTPException(status_code=422, detail="Quantidade deve ser > 0")

    tipo_mudou = new_type_str != old_type_str
    qtd_mudou = new_qty != old_qty

    # ── Posição de estoque: desfaz o efeito ANTIGO e aplica o NOVO ───────────
    # ⚠️ Precisa rodar quando o TIPO muda, não só a quantidade. Inverter
    # Entrada→Saída de N peças tira N de total_in E põe N em total_out: o saldo
    # cai 2N. A conta anterior assumia tipo constante e, num swap com a mesma
    # quantidade, não mexia em nada — a linha passaria a dizer "Saída" e a
    # posição continuaria contando como entrada.
    if tipo_mudou or qtd_mudou:
        position = db.query(models.StockPosition).filter(
            models.StockPosition.seller_id == movement.seller_id,
            models.StockPosition.sku == movement.sku,
        ).first()

        if position:
            total_in = position.total_in or 0
            total_out = position.total_out or 0

            # Desfaz o registro como ele estava
            if old_type_str == models.MovementType.IN.value:
                total_in -= old_qty
            else:
                total_out -= old_qty

            # Aplica o registro como ele fica
            if new_type_str == models.MovementType.IN.value:
                total_in += new_qty
            else:
                total_out += new_qty

            position.total_in = max(0, total_in)
            position.total_out = max(0, total_out)
            position.current_stock = position.initial_stock + position.total_in - position.total_out
            from ..services.stock_manager import calculate_stock_level
            position.level = calculate_stock_level(position.current_stock)
            position.updated_at = now_brasilia()

    if tipo_mudou:
        movement.movement_type = models.MovementType(new_type_str)
    if qty_informada:
        movement.quantity = new_qty
        movement.adjusted_quantity = new_qty

    if "observation" in body:
        movement.observation = body["observation"]
    if "nf_number" in body:
        movement.nf_number = body["nf_number"]
    if "movement_date" in body and body["movement_date"]:
        try:
            movement.movement_date = date.fromisoformat(body["movement_date"])
        except ValueError:
            pass

    # ── Trilha de auditoria: registra edição manual ──────────────────────────
    detalhe = [
        f"Edição de movimentação #{movement_id}",
        f"SKU: {movement.sku}",
        f"Seller: {movement.seller_id}",
    ]
    if tipo_mudou:
        detalhe.append(f"Tipo: {old_type_str} -> {new_type_str}")
    detalhe.append(
        f"Qtd: {old_qty} -> {new_qty}" if qtd_mudou else f"Qtd: {movement.quantity}"
    )
    detalhe.append(f"Data: {movement.movement_date}")
    if movement.order_id:
        # Editar movimento de NF é permitido (decisão do dono do sistema), mas
        # fica marcado: o sinal dele passa a divergir do da NF de origem.
        detalhe.append(f"MOVIMENTO VINCULADO A NF (order_id={movement.order_id})")

    db.add(models.AuditLog(
        entity_type="StockMovement",
        entity_id=movement_id,
        action="EDIT_MOVEMENT",
        detail=" | ".join(detalhe),
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(movement)

    mt_val = (
        movement.movement_type.value
        if hasattr(movement.movement_type, "value")
        else movement.movement_type
    )
    return {
        "id": movement.id,
        "sku": movement.sku,
        "quantity": movement.quantity,
        "movement_type": mt_val,
        "movement_date": str(movement.movement_date),
        "observation": movement.observation,
    }


@router.delete("/movements/{movement_id}")
def delete_movement(
    movement_id: int,
    body: dict,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Apaga DEFINITIVAMENTE uma movimentação (a linha some de stock_movements,
    como se nunca tivesse sido lançada). Requer papel admin + senha especial.

    O efeito da movimentação na posição de estoque é DESFEITO: se era Entrada
    de N, tira N de total_in; se era Saída, tira N de total_out. O saldo é
    recalculado na hora.

    ⚠️ Apaga a linha inclusive de movimento vinculado a NF (order_id preenchido)
    — o pedido em si NÃO é tocado. Decisão do dono do sistema, ciente de que uma
    baixa automática posterior (destravamento) pode reintroduzir o movimento.
    """
    passphrase = body.get("passphrase", "")
    if passphrase != EDIT_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Senha de edição incorreta")

    movement = db.query(models.StockMovement).filter(
        models.StockMovement.id == movement_id
    ).first()
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")

    raw_type = (
        movement.movement_type.value
        if hasattr(movement.movement_type, "value")
        else str(movement.movement_type or "")
    )
    type_str = _MT_NORMALIZE.get(raw_type.strip().upper(), raw_type)
    qty = movement.quantity or 0

    # ── Desfaz o efeito na posição de estoque ───────────────────────────────
    position = db.query(models.StockPosition).filter(
        models.StockPosition.seller_id == movement.seller_id,
        models.StockPosition.sku == movement.sku,
    ).first()
    if position:
        total_in = position.total_in or 0
        total_out = position.total_out or 0
        if type_str == models.MovementType.IN.value:
            total_in -= qty
        else:
            total_out -= qty
        position.total_in = max(0, total_in)
        position.total_out = max(0, total_out)
        position.current_stock = (
            position.initial_stock + position.total_in - position.total_out
        )
        position.level = calculate_stock_level(position.current_stock)
        position.updated_at = now_brasilia()

    # ── Trilha de auditoria: snapshot do que foi apagado ────────────────────
    detalhe = [
        f"Exclusão definitiva de movimentação #{movement_id}",
        f"SKU: {movement.sku}",
        f"Seller: {movement.seller_id}",
        f"Tipo: {type_str}",
        f"Qtd: {qty}",
        f"Data: {movement.movement_date}",
        f"NF: {movement.nf_number or '—'}",
        f"Obs: {movement.observation or '—'}",
    ]
    if movement.order_id:
        detalhe.append(f"MOVIMENTO VINCULADO A NF (order_id={movement.order_id})")

    db.add(models.AuditLog(
        entity_type="StockMovement",
        entity_id=movement_id,
        action="DELETE_MOVEMENT",
        detail=" | ".join(detalhe),
        user_id=current_user.id,
    ))

    db.delete(movement)
    db.commit()

    return {"deleted": True, "movement_id": movement_id}


# ─────────────────────────────────────────────────────────
# SKU LOOKUP
# ─────────────────────────────────────────────────────────

@router.get("/sku-lookup")
def sku_lookup(
    seller_id: int,
    sku: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Verifica se um SKU está cadastrado para o seller informado.
    Retorna nome do produto e barcode caso encontrado.
    """
    product = db.query(models.Product).filter(
        models.Product.seller_id == seller_id,
        models.Product.sku == sku,
    ).first()

    if product:
        return {
            "found": True,
            "sku": product.sku,
            "name": product.name,
            "barcode_seller": product.barcode_seller,
        }

    # Tenta pelo barcode
    product = db.query(models.Product).filter(
        models.Product.seller_id == seller_id,
        models.Product.barcode_seller == sku,
    ).first()

    if product:
        return {
            "found": True,
            "sku": product.sku,
            "name": product.name,
            "barcode_seller": product.barcode_seller,
        }

    return {"found": False, "sku": sku}


# ─────────────────────────────────────────────────────────
# EXPORTAÇÃO CSV
# ─────────────────────────────────────────────────────────

@router.get("/stock/{seller_id}/export/csv")
def export_stock_csv(
    seller_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Exporta posição de estoque em CSV para download.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")
    if user_role == "operator":
        raise HTTPException(status_code=403, detail="Operador não tem permissão para exportar estoque")

    positions = db.query(models.StockPosition).filter(
        models.StockPosition.seller_id == seller_id
    ).order_by(models.StockPosition.sku).all()

    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    seller_name = seller.name if seller else str(seller_id)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "SKU", "Produto", "Estoque Inicial", "Total Entradas",
        "Total Saidas", "Estoque Atual", "Nivel", "Tipo Insumo",
        "Valor Unit.", "Atualizado em",
    ])

    for p in positions:
        writer.writerow([
            p.sku,
            p.product_name or "",
            p.initial_stock,
            p.total_in,
            p.total_out,
            p.current_stock,
            p.level or "",
            p.supply_type or "",
            p.unit_value or 0,
            p.updated_at.strftime("%d/%m/%Y %H:%M") if p.updated_at else "",
        ])

    output.seek(0)
    filename = f"estoque_{seller_name}_{today_brasilia().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/stock/{seller_id}/export/xlsx")
def export_stock_xlsx(
    seller_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Exporta posição de estoque em Excel (.xlsx) para download.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")
    if user_role == "operator":
        raise HTTPException(status_code=403, detail="Operador não tem permissão para exportar estoque")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl não instalado. Execute: pip install openpyxl")

    positions = db.query(models.StockPosition).filter(
        models.StockPosition.seller_id == seller_id
    ).order_by(models.StockPosition.sku).all()

    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    seller_name = seller.name if seller else str(seller_id)

    PURPLE = "7B63E8"
    DARK   = "14122A"
    LIGHT  = "F0EEFF"
    WHITE  = "FFFFFF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"

    header_font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    header_fill  = PatternFill("solid", fgColor=PURPLE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font    = Font(name="Calibri", size=9)
    thin_border  = Border(
        left=Side(style="thin", color="D0CCEE"),
        right=Side(style="thin", color="D0CCEE"),
        top=Side(style="thin", color="D0CCEE"),
        bottom=Side(style="thin", color="D0CCEE"),
    )

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Posição de Estoque  |  {seller_name}  |  {today_brasilia().strftime('%d/%m/%Y')}"
    title_cell.font  = Font(name="Calibri", bold=True, color=WHITE, size=12)
    title_cell.fill  = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers    = ["SKU", "Nome do Produto", "Estoque Inicial", "Entrada", "Saída", "Estoque Final"]
    col_widths = [16, 40, 14, 12, 12, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    row_num = 3
    for p in positions:
        fill = PatternFill("solid", fgColor=LIGHT) if row_num % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        row = [p.sku, p.product_name or "", p.initial_stock, p.total_in, p.total_out, p.current_stock]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font      = data_font
            cell.border    = thin_border
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
        row_num += 1

    ws.freeze_panes = "A3"

    ws.cell(row=row_num, column=1, value=f"Total: {len(positions)} SKUs").font = \
        Font(name="Calibri", bold=True, size=9, color=PURPLE)
    ws.row_dimensions[row_num].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"estoque_{seller_name}_{today_brasilia().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/movements/{seller_id}/export/csv")
def export_movements_csv(
    seller_id: int,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    sku: Optional[str] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Exporta movimentações de estoque em CSV para download.
    """
    user_role = (
        current_user.role.value
        if hasattr(current_user.role, "value")
        else current_user.role
    )
    if user_role == "client" and current_user.seller_id != seller_id:
        raise HTTPException(status_code=403, detail="Acesso negado")

    # Usa SQL raw para evitar falha do ORM com valores legados ('IN'/'OUT')
    conds = ["seller_id = :seller_id"]
    params: dict = {"seller_id": seller_id}
    if date_from:
        conds.append("movement_date >= :date_from")
        params["date_from"] = str(date_from)
    if date_to:
        conds.append("movement_date <= :date_to")
        params["date_to"] = str(date_to)
    if sku:
        conds.append("sku LIKE :sku")
        params["sku"] = f"%{sku}%"

    movements = db.execute(
        text(f"""
            SELECT id, seller_id, sku, product_name, movement_date, movement_type,
                   quantity, adjusted_quantity, nf_number, nature,
                   observation, created_at
            FROM stock_movements
            WHERE {' AND '.join(conds)}
            ORDER BY movement_date DESC, id DESC
        """),
        params,
    ).fetchall()

    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    seller_name = seller.name if seller else str(seller_id)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "ID", "Data", "SKU", "Produto", "Tipo", "Quantidade",
        "Qtd Ajustada", "NF", "Natureza", "Observacao", "Criado em",
    ])

    for m in movements:
        raw_mt = m.movement_type or ""
        mt = _MT_NORMALIZE.get(raw_mt.upper(), raw_mt)
        ca = m.created_at
        if ca and not isinstance(ca, str):
            ca_str = ca.strftime("%d/%m/%Y %H:%M")
        elif ca:
            ca_str = ca[:16].replace("T", " ")
        else:
            ca_str = ""
        writer.writerow([
            m.id,
            str(m.movement_date) if m.movement_date else "",
            m.sku,
            m.product_name or "",
            mt,
            m.quantity,
            m.adjusted_quantity or m.quantity,
            m.nf_number or "",
            m.nature or "",
            m.observation or "",
            ca_str,
        ])

    output.seek(0)
    filename = f"movimentacoes_{seller_name}_{today_brasilia().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────
# IMPORTAÇÃO DE HISTÓRICO (Excel DETALHADO)
# ─────────────────────────────────────────────────────────

def _parse_history_excel(file_bytes: bytes):
    """
    Faz o parse do Excel no formato DETALHADO.
    Linha 1: cabeçalho mesclado (ignorado)
    Linha 2: nomes das colunas
    Linha 3+: dados
    Colunas: B=Log, C=Tipo, D=Data, E=SKU, F=Quantidade, H=Nome, I=Observação, AB=#NF

    Duas datas distintas, ambas preservadas:
      - B (Log)  → data em que a Kiwkiw processou a movimentação → movement_date.
                   A planilha de origem só preenche o Log na PRIMEIRA linha de cada
                   bloco lançado; as seguintes vêm vazias e herdam o valor de cima
                   (forward-fill).
      - D (Data) → data em que o seller emitiu a NF → nf_date (controle do seller).
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl não instalado no servidor. Execute: pip install openpyxl",
        )

    from io import BytesIO
    # read_only=True usa streaming — muito mais rápido para arquivos grandes
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

    # Encontra a aba DETALHADO (case-insensitive)
    sheet = None
    for name in wb.sheetnames:
        if "detalha" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        raise HTTPException(
            status_code=422,
            detail="Aba 'DETALHADO' não encontrada no arquivo Excel.",
        )

    def _to_date(value):
        """Converte célula do Excel em date; devolve None se não for data."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    rows = []
    last_log = None   # forward-fill: último Log preenchido
    # Pula linhas 1 (merge) e 2 (cabeçalho), lê a partir da linha 3
    for i, row in enumerate(sheet.iter_rows(values_only=True, min_row=3)):
        # Garante acesso seguro a qualquer índice (linhas curtas retornam None)
        def _cell(idx: int):
            return row[idx] if len(row) > idx else None

        log    = _cell(1)   # col B (Log — quando a Kiwkiw processou)
        tipo   = _cell(2)   # col C
        nf_dt  = _cell(3)   # col D (data da NF emitida pelo seller)
        sku    = _cell(4)   # col E
        qty    = _cell(5)   # col F
        nome   = _cell(7)   # col H
        obs    = _cell(8)   # col I
        nf     = _cell(27)  # col AB

        # Forward-fill: linha sem Log herda o da linha de cima
        log_date = _to_date(log)
        if log_date is not None:
            last_log = log_date
        data = last_log

        # Ignora linhas sem SKU ou quantidade
        if not sku or qty is None:
            continue
        sku = str(sku).strip()
        if not sku:
            continue

        try:
            qty_int = int(float(qty))
        except (ValueError, TypeError):
            continue
        if qty_int <= 0:
            continue

        # Normaliza tipo
        tipo_str = str(tipo).strip() if tipo else "Entrada"
        if tipo_str.lower() in ("saida", "saída", "out", "s"):
            tipo_norm = "Saida"
        else:
            tipo_norm = "Entrada"

        # Data da movimentação: Log (com forward-fill). Sem Log em nenhuma linha
        # anterior, cai na data da NF; em último caso, hoje.
        nf_date = _to_date(nf_dt)
        mov_date = data or nf_date or today_brasilia()

        rows.append({
            "sku": sku,
            "product_name_from_sheet": str(nome).strip() if nome else sku,
            "movement_type": tipo_norm,
            "movement_date": str(mov_date),
            "nf_date": str(nf_date) if nf_date else None,
            "quantity": qty_int,
            "observation": str(obs).strip() if obs else None,
            "nf_number": str(nf).strip() if nf else None,
        })

    wb.close()  # libera memória do workbook read_only
    return rows


# Síncrono de propósito: leitura de Excel é bloqueante e travaria o event loop.
@router.post("/import-history/{seller_id}/analyze")
def analyze_history(
    seller_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Fase 1: analisa o Excel e retorna SKUs não cadastrados.
    Não importa nada — apenas lê e informa o que precisa ser cadastrado.
    """
    file_bytes = file.file.read()
    try:
        rows = _parse_history_excel(file_bytes)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Falha ao processar o Excel: {exc}")

    if not rows:
        raise HTTPException(status_code=422, detail="Nenhuma linha válida encontrada no arquivo. Verifique se a aba DETALHADO existe e tem dados a partir da linha 3.")

    # Descobre SKUs únicos presentes no arquivo
    sku_names: dict = {}  # sku -> nome sugerido pela planilha
    for r in rows:
        if r["sku"] not in sku_names:
            sku_names[r["sku"]] = r["product_name_from_sheet"]

    # Quais SKUs já estão cadastrados?
    # ⚠️ Comparação sem diferenciar caixa (28/08/2026): 'MOSQ2' na planilha e
    # 'mosq2' no cadastro são o mesmo produto. Tratar como diferente faria a tela
    # pedir para cadastrar um SKU que já existe, criando a duplicata.
    existing_skus = {
        p.sku.lower() for p in db.query(models.Product.sku).filter(
            models.Product.seller_id == seller_id,
            func.lower(models.Product.sku).in_([s.lower() for s in sku_names]),
        ).all()
    }

    unknown = [
        {"sku": sku, "suggested_name": name, "count": sum(1 for r in rows if r["sku"] == sku)}
        for sku, name in sku_names.items()
        if sku.lower() not in existing_skus
    ]

    return {
        "total_rows": len(rows),
        "total_skus": len(sku_names),
        "unknown_skus": sorted(unknown, key=lambda x: x["sku"]),
        "already_registered": len(sku_names) - len(unknown),
    }


# Síncrono de propósito: 150k linhas de Excel + gravação em blocos travariam
# o event loop e, com ele, a API inteira.
@router.post("/import-history/{seller_id}/execute")
def execute_history_import(
    seller_id: int,
    file: UploadFile = File(...),
    product_names: str = Form("{}"),
    force: bool = Form(False),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Fase 2: importa o histórico de movimentações.
    - product_names: JSON {sku: name} para produtos a cadastrar
    - force: quando False (padrão), a importação é BLOQUEADA se sobrar algum SKU
      sem cadastro no seller — evita criar estoque de produto inexistente.
      O frontend oferece "Cadastrar mesmo assim", que envia force=True.
    - Cria produtos novos, depois importa todas as movimentações
    - Recalcula posições de estoque
    """
    file_bytes = file.file.read()
    rows = _parse_history_excel(file_bytes)

    # Cadastra produtos novos
    try:
        names_map: dict = json.loads(product_names)
    except Exception:
        names_map = {}

    # ── Trava: nenhum SKU pode ficar sem produto cadastrado ──────────────────
    if not force:
        skus_arquivo = {r["sku"] for r in rows}
        # comparação sem diferenciar caixa — ver comentário no /analyze
        cadastrados = {
            p.sku.lower() for p in db.query(models.Product.sku).filter(
                models.Product.seller_id == seller_id,
                func.lower(models.Product.sku).in_([s.lower() for s in skus_arquivo]),
            ).all()
        }
        # SKUs que o usuário preencheu nome serão criados abaixo — não bloqueiam
        a_criar = {sku.lower() for sku, nome in names_map.items() if nome and nome.strip()}
        faltantes = sorted(s for s in skus_arquivo
                           if s.lower() not in cadastrados and s.lower() not in a_criar)
        if faltantes:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": (
                        f"{len(faltantes)} SKU(s) do arquivo não estão cadastrados neste seller. "
                        f"Cadastre-os antes de importar ou use 'Cadastrar mesmo assim'."
                    ),
                    "missing_skus": faltantes[:200],
                    "missing_count": len(faltantes),
                },
            )

    products_created = 0
    for sku, name in names_map.items():
        if not name or not name.strip():
            continue
        # sem diferenciar caixa — não recria um produto que já existe com outra grafia
        existing = db.query(models.Product).filter(
            models.Product.seller_id == seller_id,
            func.lower(models.Product.sku) == sku.lower(),
        ).first()
        if not existing:
            prod = models.Product(
                seller_id=seller_id,
                sku=sku,
                name=name.strip(),
            )
            db.add(prod)
            products_created += 1

    if products_created:
        db.commit()

    # ── Importa movimentações ────────────────────────────────────────────────
    # Planilhas de histórico chegam a 150k linhas. Uma consulta por linha
    # (produto + posição) estouraria o timeout, então pré-carrega tudo em
    # memória e grava em blocos, como já faz o /bulk-stock-upload.
    imported = 0
    skipped_dup = 0
    errors = []

    # ⚠️ Chaveado em MINÚSCULA e guardando a grafia do cadastro (28/08/2026):
    # a movimentação é gravada com o SKU do CADASTRO, não com o da planilha.
    # Sem isso, uma planilha escrevendo 'MOSQ2' cria uma posição paralela à do
    # 'mosq2' cadastrado e o estoque do produto se parte em duas.
    cadastro_por_sku: dict = {}
    for p in db.query(models.Product.sku, models.Product.name, models.Product.active).filter(
        models.Product.seller_id == seller_id,
    ).order_by(models.Product.active.desc(), models.Product.sku).all():
        cadastro_por_sku.setdefault(p.sku.lower(), (p.sku, p.name))

    now_ts = now_brasilia()
    pending: list[dict] = []
    delta_in: dict = defaultdict(int)
    delta_out: dict = defaultdict(int)
    nome_final: dict = {}

    for r in rows:
        try:
            mov_date = date.fromisoformat(r["movement_date"])
        except Exception:
            mov_date = today_brasilia()

        try:
            nf_date = date.fromisoformat(r["nf_date"]) if r.get("nf_date") else None
        except Exception:
            nf_date = None

        is_in = r["movement_type"] == "Entrada"
        mt = models.MovementType.IN if is_in else models.MovementType.OUT

        sku = r["sku"]
        cad = cadastro_por_sku.get(sku.lower())
        if cad:
            sku = cad[0]                     # grafia do cadastro, não a da planilha
        # Nome cadastrado tem prioridade; senão o informado no modal; senão o da planilha
        product_name = (
            (cad[1] if cad else None)
            or names_map.get(r["sku"])
            or r["product_name_from_sheet"]
        )

        pending.append({
            "seller_id":     seller_id,
            "sku":           sku,
            "product_name":  product_name,
            "movement_date": str(mov_date),
            # .name (IN/OUT), nao .value: em producao a coluna e um ENUM nativo
            # do PostgreSQL cujos rotulos sao os nomes do enum Python. Gravar
            # "Entrada" quebra com InvalidTextRepresentation.
            "movement_type": mt.name,
            "quantity":      r["quantity"],
            "adjusted_quantity": r["quantity"],
            "nf_number":     r["nf_number"],
            "nf_date":       str(nf_date) if nf_date else None,
            "observation":   r["observation"],
            "operator_id":   current_user.id,
            "created_at":    str(now_ts),
        })

        if is_in:
            delta_in[sku] += r["quantity"]
        else:
            delta_out[sku] += r["quantity"]
        nome_final.setdefault(sku, product_name)
        imported += 1

    try:
        CHUNK = 5_000
        raw_conn = db.connection()
        for i in range(0, len(pending), CHUNK):
            raw_conn.execute(
                text("""
                    INSERT INTO stock_movements
                        (seller_id, sku, product_name, movement_date, movement_type,
                         quantity, adjusted_quantity, nf_number, nf_date, observation,
                         operator_id, created_at)
                    VALUES
                        (:seller_id, :sku, :product_name, :movement_date, :movement_type,
                         :quantity, :adjusted_quantity, :nf_number, :nf_date, :observation,
                         :operator_id, :created_at)
                """),
                pending[i: i + CHUNK],
            )

        # Posições: carrega as existentes de uma vez e aplica os deltas acumulados
        afetados = set(delta_in) | set(delta_out)
        positions_map = {
            p.sku: p
            for p in db.query(models.StockPosition).filter(
                models.StockPosition.seller_id == seller_id,
                models.StockPosition.sku.in_(list(afetados)),
            ).all()
        } if afetados else {}

        for sku in afetados:
            pos = positions_map.get(sku)
            if pos is None:
                pos = models.StockPosition(
                    seller_id=seller_id, sku=sku, product_name=nome_final.get(sku, sku),
                    initial_stock=0, total_in=0, total_out=0, current_stock=0,
                )
                db.add(pos)
                positions_map[sku] = pos
            pos.total_in  = (pos.total_in  or 0) + delta_in[sku]
            pos.total_out = (pos.total_out or 0) + delta_out[sku]
            pos.current_stock = (pos.initial_stock or 0) + pos.total_in - pos.total_out
            pos.level = calculate_stock_level(pos.current_stock)
            pos.updated_at = now_ts
            if not pos.product_name:
                pos.product_name = nome_final.get(sku, sku)

        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(500, f"Erro ao gravar no banco — nenhuma movimentação foi salva: {exc}")

    return {
        "imported": imported,
        "products_created": products_created,
        "skipped": skipped_dup,
        "errors": errors[:10],  # primeiros 10 erros
    }



# ============================================================
# BULK CSV UPLOAD DE MOVIMENTAÇÕES DE ESTOQUE — ADMIN ONLY
# ============================================================


# Síncrono de propósito: validação linha a linha + gravação em chunks.
@router.post("/bulk-stock-upload")
def bulk_stock_upload(
    file: UploadFile = File(..., description="CSV: seller;data;tipo;sku;quantity;nf;observ"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Importação em massa multi-seller via CSV.

    Comportamento:
      - FASE 1 (validação): percorre TODAS as linhas sem tocar no banco.
        Se qualquer linha tiver erro → retorna lista completa de erros e NADA é salvo.
      - FASE 2 (gravação): só executa se fase 1 passou sem erros.
        Insere movimentos em chunks + atualiza posições em memória,
        tudo em uma única transação com rollback automático em caso de falha.

    Colunas obrigatórias (delimitador vírgula ou ponto-e-vírgula):
        seller   — trade_name do seller (ou apelido cadastrado)
        data     — dd/mm/aaaa
        tipo     — Entrada | Saída | E | S
        sku      — código do produto
        quantity — inteiro > 0
        nf       — número NF (opcional)
        observ   — observação (opcional)
    """
    import csv as _csv, io as _io, time as _time, re as _re
    from datetime import datetime as _dt

    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Apenas arquivos .csv são aceitos")

    start = _time.perf_counter()
    raw = file.file.read()

    try:
        text_data = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text_data = raw.decode("latin-1")

    # Auto-detecta delimitador
    first_line = text_data.split("\n")[0] if text_data else ""
    delim = ";" if first_line.count(";") >= first_line.count(",") else ","

    reader = _csv.DictReader(_io.StringIO(text_data), delimiter=delim)
    if not reader.fieldnames:
        raise HTTPException(422, "CSV vazio ou sem cabeçalho")

    # Normaliza nomes das colunas
    reader.fieldnames = [f.strip().lower().lstrip("#") for f in reader.fieldnames]

    required = {"seller", "data", "tipo", "sku", "quantity"}
    present  = set(reader.fieldnames)
    if not required.issubset(present):
        missing = sorted(required - present)
        raise HTTPException(
            422,
            f"Colunas obrigatórias ausentes: {missing}. "
            f"Encontradas: {sorted(present)}. "
            f"Delimitador detectado: \'{delim}\'",
        )

    # Cache de sellers (trade_name, name, apelidos)
    seller_cache: dict[str, int] = {}
    for s in db.query(models.Seller).filter(models.Seller.active == True).all():
        seller_cache[s.trade_name.strip().lower()] = s.id
        if s.name:
            seller_cache[s.name.strip().lower()] = s.id
        if s.other_aliases:
            for alias in _re.split(r"[;,]", s.other_aliases):
                a = alias.strip().lower()
                if a:
                    seller_cache[a] = s.id

    rows_raw = list(reader)
    total_rows = len(rows_raw)
    if total_rows == 0:
        return {"ok": True, "total_rows": 0, "created": 0, "errors": [], "duration_sec": 0.0}

    tipo_map = {
        "e": "E", "s": "S",
        "entrada": "E", "saída": "S", "saida": "S",
        "in": "E", "out": "S",
    }

    # ── FASE 1: Validação completa — zero gravações ───────────────────────────
    valid_rows: list[dict] = []
    errors: list[str] = []
    now_ts = now_brasilia()

    for i, row in enumerate(rows_raw):
        line = i + 2  # 1-based + cabeçalho
        try:
            seller_name = (row.get("seller") or "").strip()
            date_str    = (row.get("data")   or "").strip()
            tipo        = (row.get("tipo")   or "").strip()
            sku         = (row.get("sku")    or "").strip()
            qty_raw     = (row.get("quantity") or "0").strip().replace(",", ".")
            nf          = (row.get("nf") or row.get("#nf") or "").strip() or None
            observ      = (row.get("observ") or "").strip() or None

            if not seller_name:
                errors.append(f"Linha {line}: coluna \'seller\' vazia"); continue
            if not sku:
                errors.append(f"Linha {line}: coluna \'sku\' vazia"); continue

            tipo_norm = tipo_map.get(tipo.lower(), "")
            if tipo_norm not in ("E", "S"):
                errors.append(
                    f"Linha {line}: tipo inválido \'{tipo}\' (use Entrada, Saída, E ou S)"
                ); continue

            try:
                qty = int(float(qty_raw))
            except (ValueError, TypeError):
                errors.append(f"Linha {line}: quantidade inválida \'{qty_raw}\'"); continue
            if qty <= 0:
                errors.append(
                    f"Linha {line}: quantidade deve ser > 0 — SKU {sku}, valor recebido: {qty}"
                ); continue

            try:
                mov_date = _dt.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                errors.append(
                    f"Linha {line}: data inválida \'{date_str}\' — use dd/mm/aaaa"
                ); continue

            seller_id = seller_cache.get(seller_name.lower())
            if seller_id is None:
                errors.append(
                    f"Linha {line}: seller \'{seller_name}\' não encontrado "
                    f"(verifique trade_name ou apelidos cadastrados)"
                ); continue

            mt = models.MovementType.IN if tipo_norm == "E" else models.MovementType.OUT
            valid_rows.append({
                "seller_id":     seller_id,
                "sku":           sku,
                "product_name":  None,
                "quantity":      qty,
                # .name (IN/OUT): a coluna e um ENUM nativo no PostgreSQL
                "movement_type": mt.name,
                "movement_date": str(mov_date),
                "nf_number":     nf,
                "observation":   observ,
                "operator_id":   current_user.id,
                "created_at":    str(now_ts),
            })

        except Exception as exc:
            errors.append(f"Linha {line}: erro inesperado — {exc}")

    # Se há QUALQUER erro → devolve tudo sem salvar nada
    if errors:
        duration = round(_time.perf_counter() - start, 2)
        return {
            "ok":          False,
            "total_rows":  total_rows,
            "valid_rows":  len(valid_rows),
            "created":     0,
            "errors":      errors,
            "duration_sec": duration,
        }

    # ── FASE 2: Gravação em transação única ───────────────────────────────────
    try:
        CHUNK = 5_000
        raw_conn = db.connection()
        for start_i in range(0, len(valid_rows), CHUNK):
            chunk = valid_rows[start_i: start_i + CHUNK]
            raw_conn.execute(
                text("""
                    INSERT INTO stock_movements
                        (seller_id, sku, product_name, quantity, movement_type,
                         movement_date, nf_number, observation, operator_id, created_at)
                    VALUES
                        (:seller_id, :sku, :product_name, :quantity, :movement_type,
                         :movement_date, :nf_number, :observation, :operator_id, :created_at)
                """),
                chunk,
            )

        # Atualiza posições em memória (pré-carrega → atualiza → commit único)
        from collections import defaultdict as _dd

        affected = {(r["seller_id"], r["sku"]) for r in valid_rows}
        all_seller_ids = list({s for s, _ in affected})
        all_skus       = list({k for _, k in affected})

        positions_map: dict = {
            (p.seller_id, p.sku): p
            for p in db.query(models.StockPosition).filter(
                models.StockPosition.seller_id.in_(all_seller_ids),
                models.StockPosition.sku.in_(all_skus),
            ).all()
        }

        delta_in:  dict = _dd(int)
        delta_out: dict = _dd(int)
        for r in valid_rows:
            k = (r["seller_id"], r["sku"])
            if r["movement_type"] == models.MovementType.IN.name:
                delta_in[k]  += r["quantity"]
            else:
                delta_out[k] += r["quantity"]

        now_dt = now_brasilia()
        for (sid, sku) in affected:
            k = (sid, sku)
            if k not in positions_map:
                pos = models.StockPosition(
                    seller_id=sid, sku=sku, product_name=sku,
                    initial_stock=0, total_in=0, total_out=0, current_stock=0,
                )
                db.add(pos)
                positions_map[k] = pos
            pos = positions_map[k]
            pos.total_in    = (pos.total_in  or 0) + delta_in[k]
            pos.total_out   = (pos.total_out or 0) + delta_out[k]
            pos.current_stock = (pos.initial_stock or 0) + pos.total_in - pos.total_out
            pos.level       = calculate_stock_level(pos.current_stock)
            pos.updated_at  = now_dt

        db.commit()

    except Exception as exc:
        db.rollback()
        raise HTTPException(500, f"Erro ao gravar no banco — nenhum dado foi salvo: {exc}")

    duration = round(_time.perf_counter() - start, 2)
    return {
        "ok":          True,
        "total_rows":  total_rows,
        "valid_rows":  len(valid_rows),
        "created":     len(valid_rows),
        "errors":      [],
        "duration_sec": duration,
    }


# ── Template CSV para download ─────────────────────────────────────────────────
@router.get("/bulk-stock-upload/template")
def bulk_stock_upload_template():
    """Retorna um CSV de exemplo para upload em massa de estoque."""
    from fastapi.responses import Response
    import csv, io

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["seller", "data", "tipo", "sku", "quantity", "nf", "observ"])
    writer.writerow(["NOME_SELLER", "01/01/2025", "Entrada", "SKU001", "10", "123456", "Obs opcional"])
    writer.writerow(["NOME_SELLER", "02/01/2025", "Saída",   "SKU001",  "3", "",       ""])
    csv_content = output.getvalue()

    return Response(
        content=csv_content.encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="template_estoque.csv"'},
    )
