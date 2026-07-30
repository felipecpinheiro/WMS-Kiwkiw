"""
WMS Kiwkiw - Router de Cadastros
Produtos, Kits, Algoritmo de Caixa, Sellers, Unidades e Usuários.
"""

import os, shutil, json
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..auth import (
    get_current_user, require_manager_or_above, require_admin, require_internal,
    get_user_seller_ids,
)
from ..timezone_utils import end_of_day
from ..services.kit_import import parse_kit_workbook, match_sellers, _norm as _norm_seller
from ..services.order_import import _build_seller_alias_map
from .. import models, schemas

router = APIRouter(prefix="/cadastros", tags=["Cadastros"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
MEDIA_DIR = os.path.join(BASE_DIR, "data", "media", "products")


# ============================================================
# VALIDAÇÃO DE CHAVES ESTRANGEIRAS
# ============================================================
# Sem estas checagens o INSERT chega ao banco, viola a FK e o usuário recebe
# um "Internal Server Error" sem explicação nenhuma. Mesmo padrão que
# assign_seller_unit já usava: busca explícita e 404 com mensagem clara.
#
# NENHUMA delas filtra `active` de propósito — referenciar um seller ou uma
# unidade desativada continua permitido, como era antes. Aqui só se verifica
# que a entidade EXISTE.

def _assert_seller_exists(db: Session, seller_id: Optional[int]) -> None:
    """404 quando o seller referenciado não existe. None passa (campo opcional)."""
    if seller_id is None:
        return
    if not db.query(models.Seller.id).filter(models.Seller.id == seller_id).first():
        raise HTTPException(status_code=404, detail=f"Seller {seller_id} não encontrado")


def _assert_unit_exists(db: Session, unit_id: Optional[int]) -> None:
    """404 quando a unidade referenciada não existe. None passa (campo opcional)."""
    if unit_id is None:
        return
    if not db.query(models.Unit.id).filter(models.Unit.id == unit_id).first():
        raise HTTPException(status_code=404, detail=f"Unidade {unit_id} não encontrada")


def _assert_sellers_exist(db: Session, seller_ids: Optional[List[int]]) -> None:
    """
    404 listando quais ids não existem. Vale para o vínculo M2M do usuário:
    `_sync_sellers` resolve por `id.in_(...)` e descartaria um id inexistente
    em silêncio — o admin salvaria achando que vinculou o seller.
    """
    if not seller_ids:
        return
    encontrados = {
        row.id for row in db.query(models.Seller.id).filter(models.Seller.id.in_(seller_ids)).all()
    }
    faltando = sorted(set(seller_ids) - encontrados)
    if faltando:
        raise HTTPException(
            status_code=404,
            detail=f"Seller(s) não encontrado(s): {', '.join(str(i) for i in faltando)}",
        )


# ============================================================
# PRODUTOS
# ============================================================

@router.get("/products")
def list_products(
    seller_id: Optional[int] = None,
    search: Optional[str] = None,
    active_only: bool = True,
    page: int = 1,
    page_size: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Lista produtos com paginação server-side.
    Retorna { items, total, page, page_size, pages }.
    page_size máximo: 500. Para buscas sem paginação use page_size=0 (retorna tudo — use com filtros).
    """
    page_size = max(1, min(page_size, 500)) if page_size > 0 else 0
    page = max(1, page)

    query = db.query(models.Product).options(joinedload(models.Product.seller))
    if seller_id:
        query = query.filter(models.Product.seller_id == seller_id)
    if active_only:
        query = query.filter(models.Product.active == True)
    if search:
        term = f"%{search}%"
        query = query.filter(
            models.Product.sku.ilike(term) |
            models.Product.name.ilike(term) |
            models.Product.barcode_seller.ilike(term)
        )

    total = query.count()
    query = query.order_by(models.Product.sku)

    if page_size > 0:
        query = query.offset((page - 1) * page_size).limit(page_size)

    rows = query.all()

    items = [
        {
            "id": p.id,
            "seller_id": p.seller_id,
            "seller_name": p.seller.trade_name if p.seller else None,
            "sku": p.sku,
            "name": p.name,
            "barcode_seller": p.barcode_seller,
            "unit_value": p.unit_value or 0.0,
            "box_type": p.box_type,
            "is_input": bool(p.is_input),
            "photo_url": p.photo_url,
            "active": p.active,
        }
        for p in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, -(-total // page_size)) if page_size > 0 else 1,
    }


@router.post("/products", response_model=schemas.ProductResponse)
def create_product(
    product: schemas.ProductCreate,
    current_user: models.User = Depends(require_internal),
    db: Session = Depends(get_db),
):
    _assert_seller_exists(db, product.seller_id)

    existing = db.query(models.Product).filter(
        models.Product.seller_id == product.seller_id,
        models.Product.sku == product.sku,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"SKU '{product.sku}' já cadastrado para este seller")

    p = models.Product(**product.model_dump())
    db.add(p)
    db.flush()  # Gera o ID sem commitar

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    seller = db.query(models.Seller).filter(models.Seller.id == p.seller_id).first()
    db.add(models.AuditLog(
        entity_type="Product",
        entity_id=p.id,
        action="CREATE",
        detail=f"Produto criado: SKU={p.sku} | Nome={p.name} | Seller={seller.trade_name if seller else p.seller_id}",
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(p)
    return schemas.ProductResponse(
        id=p.id, seller_id=p.seller_id,
        seller_name=seller.trade_name if seller else None,
        sku=p.sku, name=p.name, barcode_seller=p.barcode_seller,
        unit_value=p.unit_value or 0.0, box_type=p.box_type,
        is_input=bool(p.is_input), photo_url=p.photo_url, active=p.active,
    )


@router.put("/products/{product_id}", response_model=schemas.ProductResponse)
def update_product(
    product_id: int,
    data: schemas.ProductUpdate,
    current_user: models.User = Depends(require_internal),
    db: Session = Depends(get_db),
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    changed = {f: v for f, v in data.model_dump(exclude_none=True).items()}
    for field, value in changed.items():
        setattr(product, field, value)

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    seller = db.query(models.Seller).filter(models.Seller.id == product.seller_id).first()
    db.add(models.AuditLog(
        entity_type="Product",
        entity_id=product_id,
        action="UPDATE",
        detail=(
            f"Produto atualizado: SKU={product.sku} | Seller={seller.trade_name if seller else product.seller_id} | "
            f"Campos: {', '.join(changed.keys())}"
        ),
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(product)
    return schemas.ProductResponse(
        id=product.id, seller_id=product.seller_id,
        seller_name=seller.trade_name if seller else None,
        sku=product.sku, name=product.name, barcode_seller=product.barcode_seller,
        unit_value=product.unit_value or 0.0, box_type=product.box_type,
        is_input=bool(product.is_input), photo_url=product.photo_url, active=product.active,
    )


@router.delete("/products/{product_id}", status_code=204)
def delete_product(
    product_id: int,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Remove um produto do cadastro (soft-delete: seta active=False)."""
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    seller = db.query(models.Seller).filter(models.Seller.id == product.seller_id).first()
    product.active = False

    db.add(models.AuditLog(
        entity_type="Product",
        entity_id=product_id,
        action="DELETE",
        detail=(
            f"Produto desativado: SKU={product.sku} | "
            f"Seller={seller.trade_name if seller else product.seller_id}"
        ),
        user_id=current_user.id,
    ))
    db.commit()


@router.post("/products/{product_id}/reactivate", response_model=schemas.ProductResponse)
def reactivate_product(
    product_id: int,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Reativa um produto inativado (active=False → True)."""
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    if product.active:
        raise HTTPException(status_code=400, detail="Produto já está ativo")
    product.active = True
    seller = db.query(models.Seller).filter(models.Seller.id == product.seller_id).first()
    db.add(models.AuditLog(
        entity_type="Product",
        entity_id=product_id,
        action="REACTIVATE",
        detail=f"Produto reativado: SKU={product.sku} | Seller={seller.trade_name if seller else product.seller_id}",
        user_id=current_user.id,
    ))
    db.commit()
    db.refresh(product)
    return schemas.ProductResponse(
        id=product.id, seller_id=product.seller_id,
        seller_name=seller.trade_name if seller else None,
        sku=product.sku, name=product.name, barcode_seller=product.barcode_seller,
        unit_value=product.unit_value or 0.0, box_type=product.box_type,
        is_input=bool(product.is_input), photo_url=product.photo_url, active=product.active,
    )


# Síncrono de propósito: gravação em disco é bloqueante (ver comentário em orders.py).
@router.post("/products/{product_id}/photo")
def upload_product_photo(
    product_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Upload de foto do produto para exibição durante a bipagem."""
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Apenas imagens são aceitas")

    os.makedirs(MEDIA_DIR, exist_ok=True)
    ext = file.filename.split(".")[-1]
    file_path = os.path.join(MEDIA_DIR, f"product_{product_id}.{ext}")

    with open(file_path, "wb") as f:
        f.write(file.file.read())

    product.photo_url = f"/media/products/product_{product_id}.{ext}"
    db.commit()

    return {"photo_url": product.photo_url}


# ============================================================
# KITS
# ============================================================

def _assert_kit_seller_scope(seller_id: int, user_seller_ids: Optional[List[int]]):
    """
    Manager/operator só enxergam e mexem nos kits dos sellers que atendem.
    user_seller_ids None = admin (sem filtro), conforme get_user_seller_ids.
    """
    if user_seller_ids is not None and seller_id not in (user_seller_ids or []):
        raise HTTPException(
            status_code=403,
            detail="Você não atende este seller — kit fora do seu escopo",
        )


def _resolve_kit_components(seller_id: int, items: List[schemas.KitItemCreate], db: Session):
    """
    Liga cada componente ao produto do seller (product_id) e completa o nome
    a partir do cadastro. Devolve (componentes, missing_skus, nested_kits).

    Componente cujo SKU não existe no cadastro NÃO bloqueia o salvamento:
    fica sem vínculo e é devolvido em missing_skus para a tela avisar.
    A resolução por nome (aba ACERTO SKU da planilha) não faz parte disto.
    """
    skus = [(i.component_sku or "").strip() for i in items]

    produtos = db.query(models.Product).filter(
        models.Product.seller_id == seller_id,
        models.Product.sku.in_(skus or [""]),
        models.Product.active == True,
    ).all()
    por_sku = {p.sku: p for p in produtos}

    # Componente que também é kit não é explodido (a expansão é de 1 nível só).
    kits_do_seller = {
        row[0] for row in db.query(models.Kit.kit_sku).filter(
            models.Kit.seller_id == seller_id,
            models.Kit.active == True,
        ).all()
    }

    componentes, missing, nested = [], [], []
    for item in items:
        sku = (item.component_sku or "").strip()
        prod = por_sku.get(sku)
        if not prod:
            missing.append(sku)
        if sku in kits_do_seller:
            nested.append(sku)
        componentes.append({
            "component_sku": sku,
            "component_name": (item.component_name or (prod.name if prod else None) or sku),
            "quantity": item.quantity,
            "product_id": prod.id if prod else None,
        })
    return componentes, missing, nested


def _kit_items_payload(k: models.Kit) -> List[dict]:
    return [
        {
            "id": i.id,
            "component_sku": i.component_sku,
            "component_name": i.component_name,
            "quantity": i.quantity,
            "product_id": i.product_id,
            "product_found": i.product_id is not None,
        }
        for i in k.items
    ]


def _kit_to_dict(k: models.Kit, missing_skus=None, nested_kits=None) -> dict:
    """Resposta de create/update. Dict manual — kits já deram problema com ORM mode."""
    return {
        "id": k.id,
        "seller_id": k.seller_id,
        "seller_name": k.seller.trade_name if k.seller else None,
        "kit_sku": k.kit_sku,
        "kit_name": k.kit_name,
        "active": k.active,
        "items": _kit_items_payload(k),
        "warnings": {
            "missing_skus": missing_skus or [],
            "nested_kits": nested_kits or [],
        },
    }


@router.get("/kits")
def list_kits(
    seller_id: Optional[int] = None,
    current_user: models.User = Depends(get_current_user),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    query = (
        db.query(models.Kit)
        .options(joinedload(models.Kit.items), joinedload(models.Kit.seller))
        .filter(models.Kit.active == True)
    )
    if user_seller_ids is not None:
        query = query.filter(models.Kit.seller_id.in_(user_seller_ids or [-1]))
    if seller_id:
        query = query.filter(models.Kit.seller_id == seller_id)
    kits = query.order_by(models.Kit.kit_sku).all()
    return [
        {
            "id": k.id,
            "seller_id": k.seller_id,
            "seller_name": k.seller.trade_name if k.seller else None,
            "kit_sku": k.kit_sku,
            "kit_name": k.kit_name,
            "active": k.active,
            "items": _kit_items_payload(k),
        }
        for k in kits
    ]


# Sentinelas de seller_decisions (o resto dos valores é um seller_id para vincular)
SKIP = "skip"              # não importar as linhas deste cliente
REACTIVATE = "reactivate"  # reativar o seller desativado e importar nele


def _analisa_planilha_kits(file: UploadFile, db: Session, user_seller_ids: Optional[List[int]]):
    """
    Trabalho comum de analyze e execute: lê a planilha e resolve os sellers.
    Não grava nada. Devolve (parsed, matched, unmatched, fora_de_escopo, inativos).

    `inativos` = nomes que EXISTEM no cadastro mas estão desativados. Eles caem em
    `unmatched` (o match só olha sellers ativos), mas precisam ser identificados
    para a tela não sugerir vinculá-los ao seller errado por engano.
    """
    try:
        parsed = parse_kit_workbook(file.file)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Não consegui ler a planilha: {e}")

    alias_map = _build_seller_alias_map(db, active=True)
    matched, unmatched = match_sellers(parsed["kits"], alias_map)

    alias_inativos = _build_seller_alias_map(db, active=False)
    inativos = {}
    for nome in unmatched:
        s = alias_inativos.get(nome.strip().lower())
        if s:
            inativos[_norm_seller(nome)] = s

    fora_escopo = []
    if user_seller_ids is not None:
        for chave, seller in list(matched.items()):
            if seller.id not in (user_seller_ids or []):
                fora_escopo.append(seller.trade_name)
                matched.pop(chave)
    return parsed, matched, unmatched, fora_escopo, inativos


@router.post("/kits/import-file/analyze")
def analyze_kit_import_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """
    Passo 1 do import de kits por arquivo: lê a planilha e devolve o que será feito.
    NÃO grava nada — nem kit, nem seller, nem produto.
    """
    parsed, matched, unmatched, fora_escopo, inativos = _analisa_planilha_kits(
        file, db, user_seller_ids)

    # Agrupa por cliente da planilha
    por_seller: dict[str, dict] = {}
    for k in parsed["kits"]:
        info = por_seller.setdefault(k["seller_name"], {
            "seller_name": k["seller_name"], "kits": 0, "status": "unmatched",
            "matched_seller_id": None, "matched_seller_name": None,
        })
        info["kits"] += 1

    for nome, info in por_seller.items():
        chave = _norm_seller(nome)
        seller = matched.get(chave)
        if seller:
            info["status"] = "matched"
            info["matched_seller_id"] = seller.id
            info["matched_seller_name"] = seller.trade_name
        elif chave in inativos:
            # Existe no cadastro, mas está desativado: não dá para vincular sem
            # reativar antes. A tela mostra isso e oferece pular.
            info["status"] = "inactive"
            info["matched_seller_id"] = inativos[chave].id
            info["matched_seller_name"] = inativos[chave].trade_name

    # SKUs de componente que não existem no cadastro do seller (não bloqueiam)
    missing_skus = []
    for k in parsed["kits"]:
        seller = matched.get(_norm_seller(k["seller_name"]))
        if not seller:
            continue
        skus = [c["sku"] for c in k["components"]]
        existentes = {
            p.sku for p in db.query(models.Product).filter(
                models.Product.seller_id == seller.id,
                models.Product.sku.in_(skus or [""]),
                models.Product.active == True,
            ).all()
        }
        for sku in skus:
            if sku not in existentes:
                missing_skus.append({
                    "seller_name": k["seller_name"], "kit_sku": k["kit_sku"],
                    "component_sku": sku,
                })

    return {
        "sheet": parsed["sheet"],
        "header_row": parsed["header_row"],
        "component_columns": parsed["component_columns"],
        "total_kits": len(parsed["kits"]),
        "by_seller": sorted(por_seller.values(), key=lambda x: -x["kits"]),
        "unmatched_sellers": unmatched,
        "out_of_scope_sellers": fora_escopo,
        "blocked": parsed["blocked"],
        "missing_skus": missing_skus,
        "warnings": parsed["warnings"],
        "requires_confirmation": bool(unmatched),
    }


@router.post("/kits/import-file/execute")
def execute_kit_import_file(
    file: UploadFile = File(...),
    seller_decisions: Optional[str] = Form(
        None,
        description=(
            'JSON {"nome na planilha": seller_id | "skip"}. Só vincula a seller '
            'existente — nunca cria. "skip" deixa as linhas daquele cliente de fora.'
        ),
    ),
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """
    Passo 2: grava os kits. Linhas bloqueadas na análise NUNCA entram.
    Se sobrar algum nome de seller sem decisão, aborta sem gravar nada.
    """
    decisoes = {}
    if seller_decisions:
        try:
            decisoes = json.loads(seller_decisions)
        except Exception:
            raise HTTPException(status_code=400, detail="seller_decisions inválido — esperado JSON")

    parsed, matched, unmatched, fora_escopo, inativos = _analisa_planilha_kits(
        file, db, user_seller_ids)

    def sentinela(v, alvo) -> bool:
        return isinstance(v, str) and v.strip().lower() == alvo

    # ── "Pular": vale para QUALQUER cliente da planilha, reconhecido ou não ────
    # Remove do mapa de resolvidos para que as linhas dele não sejam gravadas.
    pulados = set()
    for nome, escolha in decisoes.items():
        if sentinela(escolha, SKIP):
            pulados.add(_norm_seller(nome))
            matched.pop(_norm_seller(nome), None)

    # ── "Reativar": só para cliente que existe no cadastro mas está desativado ─
    # Religa o seller (active=True) e importa nele. Mexe no cadastro do seller a
    # partir da tela de kits, então fica registrado na trilha de auditoria.
    reativados = []
    for nome, escolha in decisoes.items():
        chave = _norm_seller(nome)
        if not sentinela(escolha, REACTIVATE) or chave in pulados:
            continue
        seller = inativos.get(chave)
        if not seller:
            raise HTTPException(
                status_code=400,
                detail=f"'{nome}' não corresponde a nenhum seller desativado — nada a reativar.",
            )
        if user_seller_ids is not None and seller.id not in (user_seller_ids or []):
            raise HTTPException(status_code=403, detail=f"Seller '{nome}' está fora do seu escopo")
        seller.active = True
        db.add(models.AuditLog(
            entity_type="Seller",
            entity_id=seller.id,
            action="REACTIVATE",
            detail=(
                f"Seller reativado durante importação de kits: {seller.trade_name} "
                f"(nome na planilha: '{nome}')"
            ),
            user_id=current_user.id,
        ))
        matched[chave] = seller
        reativados.append(seller.trade_name)

    # ── Vínculo a seller já cadastrado (nunca cria) ───────────────────────────
    pendentes = []
    for nome in unmatched:
        chave = _norm_seller(nome)
        if chave in pulados or chave in matched:
            continue
        escolha = decisoes.get(nome)
        if escolha is None or escolha == "":
            pendentes.append(nome)
            continue
        try:
            seller_id_escolhido = int(escolha)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Decisão inválida para '{nome}' — esperado um seller_id, "
                    '"skip" ou "reactivate"'
                ),
            )
        seller = db.query(models.Seller).filter(models.Seller.id == seller_id_escolhido).first()
        if not seller:
            raise HTTPException(status_code=400, detail=f"Seller escolhido para '{nome}' não existe")
        if not seller.active:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"O seller escolhido para '{nome}' está inativo. Escolha "
                    '"reativar e importar" ou "não importar" para deixar essas linhas de fora.'
                ),
            )
        if user_seller_ids is not None and seller.id not in (user_seller_ids or []):
            raise HTTPException(status_code=403, detail=f"Seller escolhido para '{nome}' está fora do seu escopo")
        matched[chave] = seller

    if pendentes:
        # Mesmo padrão do import de pedidos: pausa e devolve o que falta decidir.
        return {
            "requires_confirmation": True,
            "unmatched_sellers": pendentes,
            "message": (
                f"{len(pendentes)} nome(s) de cliente da planilha não batem com nenhum seller "
                "cadastrado. Aponte cada um para o seller correto antes de importar."
            ),
            "created": 0, "updated": 0, "skipped": 0,
        }

    resultado = {
        "requires_confirmation": False,
        "created": 0, "updated": 0, "skipped": 0,
        "blocked": parsed["blocked"],
        "skipped_sellers": [],
        "reactivated_sellers": reativados,
        "missing_skus": [],
        "nested_kits": [],
        "errors": [],
    }
    pulados_contagem: dict[str, int] = {}

    for k in parsed["kits"]:
        chave = _norm_seller(k["seller_name"])
        if chave in pulados:
            # Escolha explícita do usuário: não importar as linhas deste cliente.
            resultado["skipped"] += 1
            pulados_contagem[k["seller_name"]] = pulados_contagem.get(k["seller_name"], 0) + 1
            continue

        seller = matched.get(chave)
        if not seller:
            resultado["skipped"] += 1
            resultado["errors"].append(
                f"{k['kit_sku']}: seller '{k['seller_name']}' fora do escopo ou sem decisão"
            )
            continue

        itens = [
            schemas.KitItemCreate(component_sku=c["sku"], component_name=c["name"], quantity=c["quantity"])
            for c in k["components"]
        ]
        if any(c["sku"].strip() == k["kit_sku"].strip() for c in k["components"]):
            resultado["skipped"] += 1
            resultado["errors"].append(f"{k['kit_sku']}: kit tem ele mesmo como componente")
            continue

        existing = db.query(models.Kit).filter(
            models.Kit.seller_id == seller.id,
            models.Kit.kit_sku == k["kit_sku"],
        ).first()

        if existing:
            db.query(models.KitItem).filter(models.KitItem.kit_id == existing.id).delete()
            kit = existing
            kit.kit_name = k["kit_name"]
            kit.active = True          # reimportar um kit excluído volta a ativá-lo
            resultado["updated"] += 1
            acao = "UPDATE"
        else:
            kit = models.Kit(seller_id=seller.id, kit_sku=k["kit_sku"], kit_name=k["kit_name"])
            db.add(kit)
            db.flush()
            resultado["created"] += 1
            acao = "CREATE"

        componentes, missing, nested = _resolve_kit_components(seller.id, itens, db)
        for c in componentes:
            db.add(models.KitItem(kit_id=kit.id, **c))

        for sku in missing:
            resultado["missing_skus"].append({
                "seller_name": seller.trade_name, "kit_sku": k["kit_sku"], "component_sku": sku,
            })
        for sku in nested:
            resultado["nested_kits"].append({
                "seller_name": seller.trade_name, "kit_sku": k["kit_sku"], "component_sku": sku,
            })

        comps_txt = ", ".join("{}×{}".format(c["sku"], c["quantity"]) for c in k["components"])
        db.add(models.AuditLog(
            entity_type="Kit",
            entity_id=kit.id,
            action=acao,
            detail=(
                f"Kit importado da planilha (linha {k['row']}): SKU={k['kit_sku']} | "
                f"Nome={k['kit_name']} | Seller={seller.trade_name} | Componentes: {comps_txt}"
            ),
            user_id=current_user.id,
        ))

    resultado["skipped_sellers"] = [
        {"seller_name": nome, "kits": qtd} for nome, qtd in sorted(pulados_contagem.items())
    ]
    pulados_txt = ", ".join(f"{n} ({q})" for n, q in sorted(pulados_contagem.items())) or "—"

    db.add(models.AuditLog(
        entity_type="Kit",
        entity_id=None,
        action="IMPORT",
        detail=(
            f"Importação de kits por arquivo: {file.filename} | "
            f"Criados={resultado['created']} | Atualizados={resultado['updated']} | "
            f"Bloqueados={len(parsed['blocked'])} | Ignorados={resultado['skipped']} | "
            f"Clientes pulados por escolha do usuário: {pulados_txt} | "
            f"Sellers reativados: {', '.join(reativados) or '—'}"
        ),
        user_id=current_user.id,
    ))
    db.commit()
    return resultado


@router.get("/kits/unlinked-components")
def list_unlinked_kit_components(
    seller_id: Optional[int] = None,
    current_user: models.User = Depends(get_current_user),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """
    Componentes de kit sem vínculo com o cadastro de produtos (product_id NULL).
    Alimenta a tela /kits/vincular e o aviso do Dashboard.
    """
    query = (
        db.query(models.KitItem, models.Kit, models.Seller)
        .join(models.Kit, models.KitItem.kit_id == models.Kit.id)
        .join(models.Seller, models.Kit.seller_id == models.Seller.id)
        .filter(models.KitItem.product_id.is_(None), models.Kit.active == True)
    )
    if user_seller_ids is not None:
        query = query.filter(models.Kit.seller_id.in_(user_seller_ids or [-1]))
    if seller_id:
        query = query.filter(models.Kit.seller_id == seller_id)

    rows = query.order_by(models.Seller.trade_name, models.Kit.kit_sku).all()
    return [
        {
            "item_id": it.id,
            "kit_id": k.id,
            "kit_sku": k.kit_sku,
            "kit_name": k.kit_name,
            "seller_id": k.seller_id,
            "seller_name": s.trade_name,
            "component_sku": it.component_sku,
            "component_name": it.component_name,
            "quantity": it.quantity,
        }
        for (it, k, s) in rows
    ]


@router.post("/kits/items/{item_id}/link")
def link_kit_component(
    item_id: int,
    product_id: int = Body(..., embed=True),
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """
    Vincula um componente de kit a um produto já cadastrado.
    O produto precisa ser do MESMO seller do kit — vincular ao seller errado
    faria o pedido apontar para o estoque de outro cliente.
    """
    item = db.query(models.KitItem).filter(models.KitItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Componente não encontrado")

    kit = db.query(models.Kit).filter(models.Kit.id == item.kit_id).first()
    _assert_kit_seller_scope(kit.seller_id, user_seller_ids)

    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    if product.seller_id != kit.seller_id:
        raise HTTPException(
            status_code=400,
            detail="O produto escolhido é de outro seller — só é possível vincular a um produto do mesmo seller do kit.",
        )

    sku_antigo = item.component_sku
    item.product_id = product.id
    item.component_sku = product.sku
    item.component_name = product.name

    db.add(models.AuditLog(
        entity_type="Kit",
        entity_id=kit.id,
        action="LINK_COMPONENT",
        detail=(
            f"Componente vinculado ao produto: kit={kit.kit_sku} | "
            f"SKU {sku_antigo} → {product.sku} (produto #{product.id})"
        ),
        user_id=current_user.id,
    ))
    db.commit()
    return {
        "item_id": item.id,
        "product_id": product.id,
        "component_sku": item.component_sku,
        "component_name": item.component_name,
    }


@router.get("/kits/expansion-log")
def kit_expansion_log(
    seller_id: Optional[int] = None,
    kit_sku: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    limit: int = 500,
    current_user: models.User = Depends(get_current_user),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """
    Histórico de explosões de kit — alimenta a aba "Log Explosões" da tela de Kits.

    Cada linha é um OrderItem que nasceu do split de um kit na importação
    (is_kit_component=True), com o SKU do kit de origem em original_kit_sku.
    Pedido cancelado não aparece, como em toda a operação.

    ⚠️ Rota estática declarada ANTES das rotas /kits/{kit_id}: se um dia existir um
    GET /kits/{kit_id} registrado acima, ele capturaria "expansion-log" como id.
    """
    limit = max(1, min(limit, 2000))

    query = (
        db.query(models.OrderItem, models.Order, models.Seller)
        .join(models.Order, models.OrderItem.order_id == models.Order.id)
        .join(models.Seller, models.Order.seller_id == models.Seller.id)
        .filter(
            models.OrderItem.is_kit_component == True,
            models.OrderItem.original_kit_sku.isnot(None),
            models.Order.status != models.OrderStatus.CANCELLED,
        )
    )

    # None = admin (sem filtro). Lista = manager/operator restrito aos seus sellers.
    if user_seller_ids is not None:
        query = query.filter(models.Order.seller_id.in_(user_seller_ids or [-1]))
    if seller_id:
        query = query.filter(models.Order.seller_id == seller_id)
    if kit_sku:
        query = query.filter(models.OrderItem.original_kit_sku == kit_sku)
    if date_from:
        query = query.filter(models.Order.imported_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(models.Order.imported_at <= end_of_day(date_to))

    rows = (
        query.order_by(models.Order.imported_at.desc(), models.OrderItem.id.desc())
        .limit(limit)
        .all()
    )

    return [
        {
            "order_id": o.id,
            "created_at": o.imported_at.isoformat() if o.imported_at else None,
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "seller_id": o.seller_id,
            "seller_name": s.trade_name,
            "nf_number": o.nf_number,
            "customer_name": o.customer_name,
            "kit_sku": it.original_kit_sku,
            "component_sku": it.sku,
            "component_name": it.product_name,
            "quantity": it.quantity,
        }
        for (it, o, s) in rows
    ]


@router.post("/kits")
def create_kit(
    kit: schemas.KitCreate,
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    _assert_kit_seller_scope(kit.seller_id, user_seller_ids)
    _assert_seller_exists(db, kit.seller_id)

    # Componente igual ao próprio kit geraria referência circular na expansão.
    if any((i.component_sku or "").strip() == kit.kit_sku.strip() for i in kit.items):
        raise HTTPException(
            status_code=400,
            detail=f"O kit '{kit.kit_sku}' não pode ter ele mesmo como componente.",
        )

    # A busca NÃO filtra active de propósito: a UniqueConstraint (seller_id, kit_sku)
    # do banco ignora o soft-delete. Filtrando por active, um kit excluído homônimo
    # passava na validação e o INSERT estourava IntegrityError (500).
    # Kit ativo → erro claro. Kit excluído → reativa com a composição nova.
    existing = db.query(models.Kit).filter(
        models.Kit.seller_id == kit.seller_id,
        models.Kit.kit_sku == kit.kit_sku,
    ).first()
    if existing and existing.active:
        raise HTTPException(status_code=400, detail=f"Kit '{kit.kit_sku}' já cadastrado")

    seller = db.query(models.Seller).filter(models.Seller.id == kit.seller_id).first()
    comps = ", ".join(f"{i.component_sku}×{i.quantity}" for i in kit.items)
    seller_label = seller.trade_name if seller else kit.seller_id

    if existing:
        k = existing
        k.active = True
        k.kit_name = kit.kit_name
        db.query(models.KitItem).filter(models.KitItem.kit_id == k.id).delete()
        action = "REACTIVATE"
        detail = (
            f"Kit reativado (recriado com o mesmo SKU): SKU={kit.kit_sku} | "
            f"Nome={kit.kit_name} | Seller={seller_label} | Componentes: {comps}"
        )
    else:
        k = models.Kit(
            seller_id=kit.seller_id,
            kit_sku=kit.kit_sku,
            kit_name=kit.kit_name,
        )
        db.add(k)
        db.flush()
        action = "CREATE"
        detail = (
            f"Kit criado: SKU={kit.kit_sku} | Nome={kit.kit_name} | "
            f"Seller={seller_label} | Componentes: {comps}"
        )

    componentes, missing, nested = _resolve_kit_components(kit.seller_id, kit.items, db)
    for c in componentes:
        db.add(models.KitItem(kit_id=k.id, **c))

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    db.add(models.AuditLog(
        entity_type="Kit",
        entity_id=k.id,
        action=action,
        detail=detail,
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(k)
    return _kit_to_dict(k, missing, nested)


@router.put("/kits/{kit_id}")
def update_kit(
    kit_id: int,
    kit: schemas.KitCreate,
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    existing = db.query(models.Kit).filter(models.Kit.id == kit_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Kit não encontrado")
    _assert_kit_seller_scope(existing.seller_id, user_seller_ids)

    if any((i.component_sku or "").strip() == existing.kit_sku.strip() for i in kit.items):
        raise HTTPException(
            status_code=400,
            detail=f"O kit '{existing.kit_sku}' não pode ter ele mesmo como componente.",
        )

    # kit_sku e seller_id são imutáveis. Antes vinham no payload e eram descartados
    # em silêncio: a tela mostrava "Kit atualizado!" sem ter alterado coisa alguma.
    if kit.kit_sku and kit.kit_sku.strip() != existing.kit_sku:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Não é possível alterar o SKU de um kit ('{existing.kit_sku}' → "
                f"'{kit.kit_sku.strip()}'). Crie um novo kit com o SKU desejado."
            ),
        )
    if kit.seller_id and kit.seller_id != existing.seller_id:
        raise HTTPException(
            status_code=400,
            detail="Não é possível trocar o seller de um kit. Crie um novo kit no seller desejado.",
        )

    existing.kit_name = kit.kit_name
    db.query(models.KitItem).filter(models.KitItem.kit_id == kit_id).delete()

    componentes, missing, nested = _resolve_kit_components(existing.seller_id, kit.items, db)
    for c in componentes:
        db.add(models.KitItem(kit_id=kit_id, **c))

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    comps = ", ".join(f"{i.component_sku}×{i.quantity}" for i in kit.items)
    db.add(models.AuditLog(
        entity_type="Kit",
        entity_id=kit_id,
        action="UPDATE",
        detail=f"Kit atualizado: SKU={existing.kit_sku} | Nome={kit.kit_name} | Novos componentes: {comps}",
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(existing)
    return _kit_to_dict(existing, missing, nested)


@router.delete("/kits/{kit_id}")
def delete_kit(
    kit_id: int,
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    kit = db.query(models.Kit).filter(models.Kit.id == kit_id).first()
    if not kit:
        raise HTTPException(status_code=404, detail="Kit não encontrado")
    _assert_kit_seller_scope(kit.seller_id, user_seller_ids)
    if not kit.active:
        raise HTTPException(status_code=400, detail="Kit já está inativo")
    kit.active = False

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    db.add(models.AuditLog(
        entity_type="Kit",
        entity_id=kit_id,
        action="DELETE",
        detail=f"Kit desativado: SKU={kit.kit_sku} | Nome={kit.kit_name}",
        user_id=current_user.id,
    ))

    db.commit()
    return {"message": "Kit desativado"}


@router.post("/products/bulk-paste")
def bulk_paste_products(
    items: List[schemas.ProductBulkItem],
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Upsert em massa de produtos via colagem de tabela."""
    results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    # Cache sellers
    sellers_cache = {}
    all_sellers = db.query(models.Seller).all()
    for s in all_sellers:
        sellers_cache[s.trade_name.lower()] = s
        sellers_cache[s.name.lower()] = s

    for item in items:
        if not item.sku or not item.name or not item.seller_name:
            results["skipped"] += 1
            continue

        seller = sellers_cache.get(item.seller_name.lower())
        if not seller:
            results["errors"].append(f"Seller '{item.seller_name}' não encontrado")
            results["skipped"] += 1
            continue
        if not seller.active:
            results["errors"].append(f"Seller '{item.seller_name}' está inativo — reative-o em Sellers antes de importar")
            results["skipped"] += 1
            continue

        existing = db.query(models.Product).filter(
            models.Product.seller_id == seller.id,
            models.Product.sku == item.sku,
        ).first()

        if existing:
            existing.name = item.name
            if item.box_type:
                existing.box_type = item.box_type
            if item.barcode_seller:
                existing.barcode_seller = item.barcode_seller
            if item.unit_value is not None:
                existing.unit_value = item.unit_value
            results["updated"] += 1
        else:
            p = models.Product(
                seller_id=seller.id,
                sku=item.sku,
                name=item.name,
                box_type=item.box_type,
                barcode_seller=item.barcode_seller,
                unit_value=item.unit_value or 0.0,
            )
            db.add(p)
            results["created"] += 1

    db.commit()
    return results


# Síncrono de propósito: planilha de 20k+ linhas com commit em lotes.
@router.post("/products/bulk-upload")
def bulk_upload_products(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Upload de planilha Excel de cadastro de produtos.
    Colunas esperadas (linha 1): SKU, SELLER (ou CLIENTE), NOME,
    Valor Unitário, Caixa usada, cod barras cliente.
    Faz upsert: atualiza produtos existentes, cria novos.
    Commita em lotes de 500 para suportar planilhas grandes (20k+ linhas).
    """
    import openpyxl, io as _io

    if not file.filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(status_code=400, detail="Apenas arquivos Excel (.xlsx) são aceitos")

    content = file.file.read()

    try:
        # read_only=True: streaming — 3-5× mais rápido para arquivos grandes
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Não foi possível abrir o arquivo: {e}")

    ws = wb.active

    # ── Detecta linha de cabeçalho (procura 'SKU' nas primeiras 5 linhas) ──────
    header_row_idx = None
    header_map: dict = {}
    preview_rows = []
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        preview_rows.append(row)

    for ri, row in enumerate(preview_rows, 1):
        for ci, val in enumerate(row):
            if val and str(val).upper().strip() == 'SKU':
                header_row_idx = ri
                for ci2, cv in enumerate(row):
                    if cv:
                        header_map[str(cv).upper().strip()] = ci2
                break
        if header_row_idx:
            break

    wb.close()

    if not header_row_idx:
        raise HTTPException(status_code=400, detail="Cabeçalho não encontrado. Esperava coluna 'SKU' nas primeiras 5 linhas.")

    # ── Resolve índices de coluna (aceita variações de nome) ─────────────────
    def _col(*keys):
        for k in keys:
            v = header_map.get(k.upper().strip())
            if v is not None:
                return v
        return None

    sku_col      = _col('SKU')
    seller_col   = _col('SELLER', 'CLIENTE', 'CLIENT')
    name_col     = _col('NOME', 'NAME')
    value_col    = _col('VALOR UNITÁRIO', 'VALOR', 'UNIT VALUE', 'UNITÁRIO')
    box_col      = _col('CAIXA USADA', 'CAIXA', 'BOX', 'CAIXA USADA')
    barcode_col  = _col(
        'COD BARRAS CLIENTE', 'COD DE BARRAS', 'COD BARRAS',
        'CÓDIGO DE BARRAS', 'BARCODE', 'EAN',
    )

    if sku_col is None or seller_col is None or name_col is None:
        missing = [n for n, v in [('SKU', sku_col), ('SELLER/CLIENTE', seller_col), ('NOME', name_col)] if v is None]
        raise HTTPException(status_code=400, detail=f"Colunas obrigatórias não encontradas: {', '.join(missing)}")

    # ── Cache de sellers (trade_name e name, case-insensitive) ───────────────
    all_sellers = db.query(models.Seller).all()
    sellers_cache: dict = {}
    for s in all_sellers:
        sellers_cache[s.trade_name.strip().lower()] = s
        sellers_cache[s.name.strip().lower()] = s
        # também guarda sem espaços extras
        if s.code:
            sellers_cache[s.code.strip().lower()] = s

    # ── Cache de produtos existentes: (seller_id, sku) → Product ─────────────
    existing_products: dict = {}
    for p in db.query(models.Product).all():
        existing_products[(p.seller_id, p.sku)] = p

    results = {"created": 0, "updated": 0, "skipped": 0, "errors": [], "sellers_not_found": set(), "inactive_sellers": set()}
    BATCH = 500
    pending = 0

    # Re-abre em read_only para iterar as linhas de dados
    wb2 = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
    ws2 = wb2.active

    for row_num, row in enumerate(ws2.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        def _v(idx):
            return row[idx] if idx is not None and len(row) > idx else None

        raw_sku    = _v(sku_col)
        raw_seller = _v(seller_col)
        raw_name   = _v(name_col)

        if not raw_sku or not raw_seller or not raw_name:
            results["skipped"] += 1
            continue

        sku        = str(raw_sku).strip()
        seller_key = str(raw_seller).strip()
        name       = str(raw_name).strip()

        if not sku or not seller_key or not name:
            results["skipped"] += 1
            continue

        # Valor unitário
        raw_val = _v(value_col)
        try:
            unit_value = float(raw_val) if raw_val is not None else 0.0
        except (ValueError, TypeError):
            unit_value = 0.0

        # Caixa e código de barras
        raw_box     = _v(box_col)
        raw_barcode = _v(barcode_col)
        box_type    = str(raw_box).strip()    if raw_box    else None
        barcode     = str(raw_barcode).strip() if raw_barcode else None
        # Normaliza variações de "Própria"
        if box_type and box_type.lower() in ('propria', 'própria', 'prop.', 'próp.', 'popria', 'propria '):
            box_type = 'Própria'

        # Resolve seller
        seller = sellers_cache.get(seller_key.lower())
        if not seller:
            results["sellers_not_found"].add(seller_key)
            results["skipped"] += 1
            continue
        if not seller.active:
            results["inactive_sellers"].add(seller_key)
            results["skipped"] += 1
            continue

        key = (seller.id, sku)
        existing = existing_products.get(key)

        if existing:
            existing.name       = name
            existing.unit_value = unit_value
            if box_type:
                existing.box_type = box_type
            if barcode:
                existing.barcode_seller = barcode
            results["updated"] += 1
        else:
            prod = models.Product(
                seller_id=seller.id,
                sku=sku,
                name=name,
                unit_value=unit_value,
                box_type=box_type,
                barcode_seller=barcode,
            )
            db.add(prod)
            existing_products[key] = prod   # evita duplicata na mesma planilha
            results["created"] += 1

        pending += 1
        if pending >= BATCH:
            db.commit()
            pending = 0

    if pending:
        db.commit()

    wb2.close()

    # Converte set para lista para serialização JSON
    results["sellers_not_found"] = sorted(results["sellers_not_found"])
    results["inactive_sellers"] = sorted(results["inactive_sellers"])

    # ── Trilha de auditoria: registra import em lote ─────────────────────────
    db.add(models.AuditLog(
        entity_type="Product",
        entity_id=None,
        action="BULK_UPLOAD",
        detail=(
            f"Upload em lote de produtos: arquivo={file.filename} | "
            f"Criados={results['created']} | Atualizados={results['updated']} | "
            f"Ignorados={results['skipped']}"
        ),
        user_id=current_user.id,
    ))
    db.commit()

    return results


@router.post("/kits/bulk-import")
def bulk_import_kits(
    payload: schemas.KitBulkImport,
    current_user: models.User = Depends(require_manager_or_above),
    user_seller_ids: Optional[List[int]] = Depends(get_user_seller_ids),
    db: Session = Depends(get_db),
):
    """Import em massa de kits."""
    all_sellers = db.query(models.Seller).all()
    sellers_cache = {}
    for s in all_sellers:
        sellers_cache[s.trade_name.lower()] = s
        sellers_cache[s.name.lower()] = s

    # Cache products for name lookup
    all_products = db.query(models.Product).filter(models.Product.active == True).all()
    prod_cache = {}
    for p in all_products:
        prod_cache[(p.seller_id, p.sku)] = p

    results = {"created": 0, "updated": 0, "skipped": 0, "errors": [], "unresolved": []}

    for item in payload.items:
        seller = sellers_cache.get(item.seller_name.lower())
        if not seller:
            results["errors"].append(f"Seller '{item.seller_name}' não encontrado")
            results["skipped"] += 1
            continue
        if not seller.active:
            results["errors"].append(f"Seller '{item.seller_name}' está inativo — reative-o em Sellers antes de importar")
            results["skipped"] += 1
            continue
        if user_seller_ids is not None and seller.id not in (user_seller_ids or []):
            results["errors"].append(f"Seller '{item.seller_name}' está fora do seu escopo de atendimento")
            results["skipped"] += 1
            continue

        if not item.kit_sku or not item.components:
            results["skipped"] += 1
            continue

        # Get kit name from product if exists
        prod = prod_cache.get((seller.id, item.kit_sku))
        kit_name = prod.name if prod else item.kit_sku

        # Check if kit exists
        existing = db.query(models.Kit).filter(
            models.Kit.seller_id == seller.id,
            models.Kit.kit_sku == item.kit_sku,
        ).first()

        if existing:
            # Update: remove old items and recreate
            db.query(models.KitItem).filter(models.KitItem.kit_id == existing.id).delete()
            kit = existing
            kit.kit_name = kit_name
            # Reimportar um kit que havia sido excluído volta a ativá-lo. Sem isso
            # o kit era atualizado, contava como "updated", mas seguia active=False:
            # não aparecia na listagem nem era expandido na importação de pedidos.
            kit.active = True
            results["updated"] += 1
        else:
            kit = models.Kit(
                seller_id=seller.id,
                kit_sku=item.kit_sku,
                kit_name=kit_name,
            )
            db.add(kit)
            db.flush()
            results["created"] += 1

        for comp in item.components:
            comp_prod = prod_cache.get((seller.id, comp.sku))
            if not comp_prod:
                # Não bloqueia: o kit é salvo e o componente fica sem vínculo,
                # aparecendo depois na tela /kits/vincular.
                results["unresolved"].append({
                    "seller_name": seller.trade_name,
                    "kit_sku": item.kit_sku,
                    "component_sku": comp.sku,
                })
            ki = models.KitItem(
                kit_id=kit.id,
                component_sku=comp.sku,
                component_name=comp_prod.name if comp_prod else comp.sku,
                quantity=comp.quantity,
                product_id=comp_prod.id if comp_prod else None,
            )
            db.add(ki)

    db.commit()
    return results


# ============================================================
# ALGORITMO DE CAIXA
# ============================================================

@router.get("/box-algorithm/{seller_id}", response_model=List[schemas.BoxRuleResponse])
def get_box_algorithm(
    seller_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return db.query(models.BoxAlgorithm).filter(
        models.BoxAlgorithm.seller_id == seller_id
    ).order_by(models.BoxAlgorithm.num_products, models.BoxAlgorithm.score).all()


@router.post("/box-algorithm", response_model=schemas.BoxRuleResponse)
def create_box_rule(
    rule: schemas.BoxRuleCreate,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    _assert_seller_exists(db, rule.seller_id)

    existing = db.query(models.BoxAlgorithm).filter(
        models.BoxAlgorithm.seller_id == rule.seller_id,
        models.BoxAlgorithm.num_products == rule.num_products,
        models.BoxAlgorithm.score == rule.score,
    ).first()

    if existing:
        existing.box_type = rule.box_type
        db.commit()
        db.refresh(existing)
        return existing

    r = models.BoxAlgorithm(**rule.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.get("/box-algorithm/{seller_id}/calculate")
def calculate_box(
    seller_id: int,
    num_products: int,
    score: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Calcula a caixa para um pedido dado nº de produtos e score."""
    # Busca a regra mais próxima
    rule = db.query(models.BoxAlgorithm).filter(
        models.BoxAlgorithm.seller_id == seller_id,
        models.BoxAlgorithm.num_products <= num_products,
        models.BoxAlgorithm.score <= score,
    ).order_by(
        models.BoxAlgorithm.num_products.desc(),
        models.BoxAlgorithm.score.desc(),
    ).first()

    if not rule:
        return {"box_type": "Propria", "message": "Regra não encontrada, usar caixa própria"}

    return {"box_type": rule.box_type, "num_products": num_products, "score": score}


# ============================================================
# SELLERS
# ============================================================

@router.get("/sellers", response_model=List[schemas.SellerResponse])
def list_sellers(
    active_only: bool = False,   # False = retorna todos; True = só ativos (para dropdowns)
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from sqlalchemy import func as _func

    sellers = db.query(models.Seller)
    if active_only:
        sellers = sellers.filter(models.Seller.active == True)
    sellers = sellers.order_by(models.Seller.trade_name).all()

    # Enriquece com contagem de SKUs e SKUs com estoque (queries otimizadas com GROUP BY)
    sku_totals = dict(
        db.query(models.Product.seller_id, _func.count(models.Product.id))
        .filter(models.Product.active == True)
        .group_by(models.Product.seller_id)
        .all()
    )
    sku_with_stock = dict(
        db.query(models.StockPosition.seller_id, _func.count(models.StockPosition.id))
        .filter(models.StockPosition.current_stock > 0)
        .group_by(models.StockPosition.seller_id)
        .all()
    )

    result = []
    for s in sellers:
        d = {c.key: getattr(s, c.key) for c in s.__table__.columns}
        d["unit_display_name"] = s.unit_display_name
        d["total_skus"]      = sku_totals.get(s.id, 0)
        d["skus_with_stock"] = sku_with_stock.get(s.id, 0)
        result.append(d)
    return result


@router.get("/sellers/{seller_id}/stats")
def seller_stats(
    seller_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retorna estatísticas resumidas do seller: total SKUs, SKUs com estoque, valor total."""
    from sqlalchemy import func as _func
    total_skus = db.query(_func.count(models.Product.id)).filter(
        models.Product.seller_id == seller_id,
        models.Product.active == True,
    ).scalar() or 0

    skus_with_stock = db.query(_func.count(models.StockPosition.id)).filter(
        models.StockPosition.seller_id == seller_id,
        models.StockPosition.current_stock > 0,
    ).scalar() or 0

    total_stock_value = db.query(
        _func.sum(models.StockPosition.current_stock * models.StockPosition.unit_value)
    ).filter(
        models.StockPosition.seller_id == seller_id,
        models.StockPosition.current_stock > 0,
    ).scalar() or 0.0

    return {
        "seller_id":       seller_id,
        "total_skus":      total_skus,
        "skus_with_stock": skus_with_stock,
        "skus_zero_stock": total_skus - skus_with_stock,
        "total_stock_value": round(float(total_stock_value), 2),
    }


@router.post("/sellers", response_model=schemas.SellerResponse)
def create_seller(
    seller: schemas.SellerCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    data = seller.model_dump()
    _assert_unit_exists(db, data.get("unit_id"))
    # trade_name não pode ser nulo no banco; usa name como fallback
    if not data.get('trade_name'):
        data['trade_name'] = data['name']
    s = models.Seller(**data)
    db.add(s)
    db.flush()

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    db.add(models.AuditLog(
        entity_type="Seller",
        entity_id=s.id,
        action="CREATE",
        detail=f"Seller cadastrado: {s.trade_name} (razão: {s.name})",
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(s)
    return s


@router.post("/sellers/{seller_id}/assign-unit")
def assign_seller_unit(
    seller_id: int,
    unit_id: int = Body(..., embed=True),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Associa uma unidade a um seller — ação restrita usada na página de
    correção de sellers sem unidade. Diferente de `update_seller` (admin-only,
    edição completa do cadastro), este endpoint só altera `unit_id` e é
    liberado para manager também.
    """
    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller não encontrado")
    unit = db.query(models.Unit).filter(models.Unit.id == unit_id).first()
    if not unit:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")

    seller.unit_id = unit_id
    db.add(models.AuditLog(
        entity_type="Seller",
        entity_id=seller_id,
        action="ASSIGN_UNIT",
        detail=f"Unidade '{unit.name}' associada ao seller '{seller.trade_name}' via página de correção",
        user_id=current_user.id,
    ))
    db.commit()
    return {"seller_id": seller_id, "unit_id": unit_id}


@router.put("/sellers/{seller_id}", response_model=schemas.SellerResponse)
def update_seller(
    seller_id: int,
    data: schemas.SellerUpdate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller não encontrado")
    changed = {f: v for f, v in data.model_dump(exclude_none=True).items()}
    _assert_unit_exists(db, changed.get("unit_id"))
    for field, value in changed.items():
        setattr(seller, field, value)

    # ── Trilha de auditoria ──────────────────────────────────────────────────
    db.add(models.AuditLog(
        entity_type="Seller",
        entity_id=seller_id,
        action="UPDATE",
        detail=f"Seller atualizado: {seller.trade_name} | Campos: {', '.join(changed.keys())}",
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(seller)
    return seller


@router.delete("/sellers/{seller_id}")
def delete_seller(
    seller_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller não encontrado")
    seller.active = False
    db.commit()
    return {"message": "Seller desativado"}


# ============================================================
# SELLERS SEM UNIDADE — para warning no Dashboard
# ============================================================

@router.get("/sellers/without-unit")
def sellers_without_unit(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retorna sellers ativos sem unidade associada (unit_id nulo), com a contagem de pedidos presos em cada um."""
    sellers = db.query(models.Seller).filter(
        models.Seller.active == True,
        models.Seller.unit_id == None,  # noqa: E711
    ).all()
    result = []
    for s in sellers:
        order_count = db.query(models.Order).filter(models.Order.seller_id == s.id).count()
        result.append({"id": s.id, "trade_name": s.trade_name, "order_count": order_count})
    return result


@router.post("/sellers/{from_seller_id}/merge-orders-into/{to_seller_id}")
def merge_seller_orders(
    from_seller_id: int,
    to_seller_id: int,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Reatribui todos os pedidos (Order.seller_id) de `from_seller_id` para
    `to_seller_id` — usado para corrigir sellers duplicados/sem unidade que
    acabaram recebendo pedidos por engano. Não move produtos, estoque, kits
    ou configuração de cobrança — só pedidos.
    """
    if from_seller_id == to_seller_id:
        raise HTTPException(status_code=400, detail="Seller de origem e destino não podem ser o mesmo")

    from_seller = db.query(models.Seller).filter(models.Seller.id == from_seller_id).first()
    to_seller = db.query(models.Seller).filter(models.Seller.id == to_seller_id).first()
    if not from_seller:
        raise HTTPException(status_code=404, detail="Seller de origem não encontrado")
    if not to_seller:
        raise HTTPException(status_code=404, detail="Seller de destino não encontrado")

    count = db.query(models.Order).filter(models.Order.seller_id == from_seller_id).update(
        {models.Order.seller_id: to_seller_id}, synchronize_session=False
    )
    db.add(models.AuditLog(
        entity_type="Seller",
        entity_id=to_seller_id,
        action="MERGE_ORDERS",
        detail=(
            f"{count} pedido(s) migrado(s) de '{from_seller.trade_name}' (id={from_seller_id}) "
            f"para '{to_seller.trade_name}' (id={to_seller_id})"
        ),
        user_id=current_user.id,
    ))
    db.commit()
    return {"migrated_orders": count, "from_seller_id": from_seller_id, "to_seller_id": to_seller_id}


# ============================================================
# UNIDADES
# ============================================================

def _unit_to_response(u: models.Unit) -> schemas.UnitResponse:
    """Serializa Unit → UnitResponse incluindo seller_ids."""
    return schemas.UnitResponse(
        id=u.id,
        name=u.name,
        code=u.code,
        location=u.location,
        city=u.city,
        state=u.state,
        responsible=u.responsible,
        phone=u.phone,
        active=u.active,
        seller_ids=[s.id for s in (u.sellers or [])],
    )


@router.get("/units", response_model=List[schemas.UnitResponse])
def list_units(
    include_inactive: bool = False,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(models.Unit).options(joinedload(models.Unit.sellers))
    if not include_inactive:
        query = query.filter(models.Unit.active == True)
    return [_unit_to_response(u) for u in query.order_by(models.Unit.name).all()]


@router.post("/units", response_model=schemas.UnitResponse, status_code=201)
def create_unit(
    unit: schemas.UnitCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    u = models.Unit(**unit.model_dump())
    db.add(u)
    db.commit()
    db.refresh(u)
    db.refresh(u, attribute_names=['sellers'])
    db.add(models.AuditLog(
        entity_type="Unit", entity_id=u.id, action="CREATE",
        detail=f"Unidade criada: {u.name}", user_id=current_user.id,
    ))
    db.commit()
    return _unit_to_response(u)


@router.put("/units/{unit_id}", response_model=schemas.UnitResponse)
def update_unit(
    unit_id: int,
    data: schemas.UnitUpdate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    u = db.query(models.Unit).options(joinedload(models.Unit.sellers)).filter(
        models.Unit.id == unit_id
    ).first()
    if not u:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")

    changed = {f: v for f, v in data.model_dump(exclude_none=True).items()}
    for field, value in changed.items():
        setattr(u, field, value)

    db.add(models.AuditLog(
        entity_type="Unit", entity_id=unit_id, action="UPDATE",
        detail=f"Unidade atualizada: {u.name} | Campos: {', '.join(changed.keys())}",
        user_id=current_user.id,
    ))
    db.commit()
    db.refresh(u)
    return _unit_to_response(u)


@router.delete("/units/{unit_id}", status_code=204)
def delete_unit(
    unit_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    u = db.query(models.Unit).filter(models.Unit.id == unit_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")
    # Desassocia os sellers antes de desativar
    db.query(models.Seller).filter(models.Seller.unit_id == unit_id).update(
        {"unit_id": None}, synchronize_session="fetch"
    )
    u.active = False
    db.add(models.AuditLog(
        entity_type="Unit", entity_id=unit_id, action="DELETE",
        detail=f"Unidade desativada: {u.name}", user_id=current_user.id,
    ))
    db.commit()


@router.patch("/units/{unit_id}/sellers")
def assign_sellers_to_unit(
    unit_id: int,
    payload: dict,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Associa/desassocia sellers a uma unidade.
    Body: { "seller_ids": [1, 2, 3] }
    Substitui completamente a lista de sellers da unidade.
    """
    u = db.query(models.Unit).filter(models.Unit.id == unit_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")

    seller_ids: List[int] = payload.get("seller_ids", [])

    # Remove todos os sellers desta unidade
    db.query(models.Seller).filter(models.Seller.unit_id == unit_id).update(
        {"unit_id": None}, synchronize_session="fetch"
    )
    # Atribui os novos sellers
    if seller_ids:
        db.query(models.Seller).filter(models.Seller.id.in_(seller_ids)).update(
            {"unit_id": unit_id}, synchronize_session="fetch"
        )

    db.commit()
    return {"unit_id": unit_id, "seller_ids": seller_ids}


# ============================================================
# USUÁRIOS — helpers
# ============================================================

def _user_to_response(u: models.User) -> schemas.UserResponse:
    """Converte User ORM → UserResponse enriquecendo com nomes e seller_ids."""
    role = u.role.value if hasattr(u.role, "value") else u.role
    return schemas.UserResponse(
        id=u.id,
        name=u.name,
        email=u.email,
        role=role,
        unit_id=u.unit_id,
        unit_name=u.unit.name if u.unit else None,
        seller_id=u.seller_id,
        seller_name=(u.seller.trade_name if u.seller else None),
        seller_ids=[s.id for s in (u.sellers or [])],
        seller_names=[s.trade_name for s in (u.sellers or [])],
        active=u.active,
        force_password_change=bool(u.force_password_change),
        created_at=u.created_at,
        last_login=u.last_login,
    )


def _sync_sellers(u: models.User, seller_ids: list, db: Session) -> None:
    """Sincroniza a lista de sellers associados ao usuário (many-to-many)."""
    if seller_ids is None:
        return
    sellers = db.query(models.Seller).filter(models.Seller.id.in_(seller_ids)).all() if seller_ids else []
    u.sellers = sellers


# ============================================================
# USUÁRIOS — endpoints
# ============================================================

@router.get("/users", response_model=List[schemas.UserResponse])
def list_users(
    unit_id: Optional[int] = None,
    role: Optional[str] = None,
    active_only: bool = True,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Lista todos os usuários. Somente administrador."""
    query = db.query(models.User)
    if active_only:
        query = query.filter(models.User.active == True)
    if unit_id:
        query = query.filter(models.User.unit_id == unit_id)
    if role:
        query = query.filter(models.User.role == role)
    users = query.order_by(models.User.name).all()
    return [_user_to_response(u) for u in users]


@router.post("/users", response_model=schemas.UserResponse, status_code=201)
def create_user(
    user: schemas.UserCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Cria novo usuário. Somente administrador."""
    from ..auth import hash_password
    existing = db.query(models.User).filter(models.User.email == user.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")

    _assert_unit_exists(db, user.unit_id)
    _assert_seller_exists(db, user.seller_id)
    _assert_sellers_exist(db, user.seller_ids)

    u = models.User(
        name=user.name,
        email=user.email,
        password_hash=hash_password(user.password),
        role=user.role,
        unit_id=user.unit_id,
        seller_id=user.seller_id,   # seller principal (client)
    )
    db.add(u)
    db.flush()   # gera u.id antes de vincular sellers

    # Vincula sellers (many-to-many) para manager/operator
    _sync_sellers(u, user.seller_ids or [], db)

    db.add(models.AuditLog(
        entity_type="User",
        entity_id=u.id,
        action="CREATE",
        detail=(
            f"Usuário criado: {u.name} | email={u.email} "
            f"| role={u.role.value if hasattr(u.role, 'value') else u.role} "
            f"| sellers={user.seller_ids}"
        ),
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(u)
    return _user_to_response(u)


@router.put("/users/{user_id}", response_model=schemas.UserResponse)
def update_user(
    user_id: int,
    data: schemas.UserUpdate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Edita usuário. Somente administrador."""
    from ..auth import hash_password
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    # Campos simples
    update_data = data.model_dump(exclude_none=True)
    seller_ids_new = update_data.pop("seller_ids", None)   # trata separado

    _assert_unit_exists(db, update_data.get("unit_id"))
    _assert_seller_exists(db, update_data.get("seller_id"))
    _assert_sellers_exist(db, seller_ids_new)

    if "password" in update_data:
        update_data["password_hash"] = hash_password(update_data.pop("password"))

    for field, value in update_data.items():
        if hasattr(u, field):
            setattr(u, field, value)

    # Sincroniza sellers (many-to-many) se enviado
    if seller_ids_new is not None:
        _sync_sellers(u, seller_ids_new, db)

    db.add(models.AuditLog(
        entity_type="User",
        entity_id=user_id,
        action="UPDATE",
        detail=(
            f"Usuário atualizado: {u.name} "
            f"| Campos: {', '.join(k for k in update_data if k != 'password_hash')}"
            + (f" | sellers={seller_ids_new}" if seller_ids_new is not None else "")
        ),
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(u)
    return _user_to_response(u)


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Inativa o usuário (nunca deleta fisicamente, preserva auditoria)."""
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Você não pode inativar a si mesmo")

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    user.active = False

    db.add(models.AuditLog(
        entity_type="User",
        entity_id=user_id,
        action="DEACTIVATE",
        detail=f"Usuário inativado: {user.name} | email={user.email}",
        user_id=current_user.id,
    ))

    db.commit()
    return {"message": f"Usuário {user.name} inativado com sucesso"}


@router.post("/users/{user_id}/reactivate", response_model=schemas.UserResponse)
def reactivate_user(
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Reativa um usuário inativo."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if user.active:
        raise HTTPException(status_code=400, detail="Usuário já está ativo")

    user.active = True

    db.add(models.AuditLog(
        entity_type="User",
        entity_id=user_id,
        action="REACTIVATE",
        detail=f"Usuário reativado: {user.name} | email={user.email}",
        user_id=current_user.id,
    ))

    db.commit()
    db.refresh(user)
    return _user_to_response(user)


# ============================================================
# EXPERIENCE FILE — upload e servir roteiro de unboxing do seller
# ============================================================

EXPERIENCE_DIR = os.path.join(BASE_DIR, "data", "media", "experience")


@router.post("/sellers/{seller_id}/experience-file")
def upload_experience_file(
    seller_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """Salva o arquivo PPT/PDF de roteiro de experiencia do seller."""
    seller = db.query(models.Seller).filter(models.Seller.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller nao encontrado")

    allowed = {".pdf", ".ppt", ".pptx"}
    _, ext = os.path.splitext(file.filename or "")
    if ext.lower() not in allowed:
        raise HTTPException(status_code=400, detail=f"Extensao nao permitida. Use: {', '.join(allowed)}")

    os.makedirs(EXPERIENCE_DIR, exist_ok=True)

    dest_name = f"seller_{seller_id}{ext.lower()}"
    dest_path = os.path.join(EXPERIENCE_DIR, dest_name)

    with open(dest_path, "wb") as out:
        shutil.copyfileobj(file.file, out)

    try:
        from sqlalchemy import text
        db.execute(
            text("UPDATE sellers SET experience_file_path = :p WHERE id = :id"),
            {"p": dest_name, "id": seller_id},
        )
        db.commit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar arquivo: {e}")

    return {"filename": dest_name, "path": f"/media/experience/{dest_name}"}
