"""
WMS Kiwkiw - Router de Devoluções

Lançamento de devoluções (mercadoria que voltou do cliente final) por dois
caminhos que compartilham a MESMA validação e a MESMA gravação:

  1. planilha  -> GET /devolucoes/modelo, POST /devolucoes/analyze
  2. tela      -> a tabelinha do front monta as linhas direto

Os dois terminam em POST /devolucoes/lancar, que REVALIDA tudo no servidor
(a tela pode ser burlada e um erro aqui vira estoque errado).

REGRAS (decididas com o dono do sistema em 02/09/2026):

  * TUDO-OU-NADA: qualquer linha com problema bloqueia o lote inteiro.
  * Linha que RETORNA ao estoque  -> StockMovement de Entrada, data = hoje
    (data do lançamento, não a da NF nem a da chegada física).
  * Linha que NÃO retorna         -> nenhum movimento; só um AuditLog
    (entity_type="Devolucao"), visível na Trilha de Auditoria.
  * ⚠️ O movimento NASCE SEM `order_id`, de propósito. Amarrá-lo à NF de venda
    faria `reverse_stock_for_order()` (que trabalha por saldo líquido do
    order_id) varrer a devolução junto ao cancelar/inativar aquela NF, e o
    estoque sumiria em silêncio. O número da NF fica no `nf_number` e na
    observação — aparece na tela e no Portal do Seller do mesmo jeito.
  * NÃO existe trava contra reenvio do mesmo arquivo (decisão do dono):
    subir duas vezes lança duas vezes.
"""

import io
import json
import unicodedata
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..auth import require_manager_or_above
from ..timezone_utils import now_brasilia, today_brasilia
from .. import models
from ..services.stock_manager import update_stock_position
from ..services.order_import import _build_seller_alias_map

router = APIRouter(prefix="/devolucoes", tags=["Devoluções"])

COLUNAS_MODELO = [
    "Seller",
    "NF",
    "SKU",
    "Quantidade",
    "Retorna ao estoque",
    "Motivo",
]

# Aceita as grafias que o usuário pediu, sem diferenciar maiúscula nem acento.
_SIM = {"s", "sim", "1", "x", "true", "v", "verdadeiro"}
_NAO = {"n", "nao", "0", "false", "f", "falso"}


def _strip_accents(txt: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", txt)
        if unicodedata.category(c) != "Mn"
    )


def _parse_retorna(raw) -> Optional[bool]:
    """
    Converte a coluna 'Retorna ao estoque' em bool.
    Devolve None quando o valor está vazio ou fora do combinado — o chamador
    transforma isso em erro (nunca assume um lado).
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        if float(raw) == 1:
            return True
        if float(raw) == 0:
            return False
        return None
    key = _strip_accents(str(raw).strip().lower())
    if not key:
        return None
    if key in _SIM:
        return True
    if key in _NAO:
        return False
    return None


def _parse_qtd(raw) -> Optional[int]:
    """Inteiro > 0. Aceita o float que o Excel grava (3.0) mas recusa 3.5."""
    if raw is None or str(raw).strip() == "":
        return None
    try:
        val = float(str(raw).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    if val <= 0 or val != int(val):
        return None
    return int(val)


def _cell_str(raw) -> str:
    """
    Texto da célula. O Excel grava número como float: NF 123456 vira 123456.0
    e o SKU '678' vira 678.0 — o '.0' é removido, como no import de pedidos.
    """
    if raw is None:
        return ""
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    return str(raw).strip()


# ─────────────────────────────────────────────────────────
# MODELO DA PLANILHA
# ─────────────────────────────────────────────────────────

@router.get("/modelo")
def download_modelo(
    current_user: models.User = Depends(require_manager_or_above),
):
    """
    Gera o Excel modelo EM MEMÓRIA (nunca em disco — em produção o sistema de
    arquivos é efêmero) e devolve como download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "DEVOLUCOES"

    ws.append(COLUNAS_MODELO)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7B63E8")
        cell.alignment = Alignment(horizontal="center")

    ws.append(["NOME DO SELLER", "123456", "SKU-EXEMPLO-1", 2, "S", ""])
    ws.append(["NOME DO SELLER", "123456", "SKU-EXEMPLO-2", 1, "N", "Produto avariado"])

    larguras = [32, 16, 26, 13, 22, 40]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws2 = wb.create_sheet("INSTRUCOES")
    for linha in [
        ["Como preencher"],
        [""],
        ["Seller", "Nome do seller como esta cadastrado no WMS (apelido, razao social ou apelido alternativo)."],
        ["NF", "Numero da NF de venda que a Kiwkiw enviou e que voltou. Obrigatorio."],
        ["SKU", "SKU do seller. Precisa ter produto ativo cadastrado, senao a subida e bloqueada."],
        ["Quantidade", "Numero inteiro maior que zero."],
        ["Retorna ao estoque", "S / SIM / 1 / X para voltar ao estoque; N / NAO / 0 para nao voltar."],
        ["Motivo", "Opcional. So faz sentido quando a linha NAO retorna ao estoque."],
        [""],
        ["Observacoes"],
        ["", "O mesmo SKU pode aparecer em duas linhas (um volta, outro nao)."],
        ["", "Linhas identicas (mesmo seller, NF, SKU, quantidade e S/N) bloqueiam a subida."],
        ["", "Se qualquer linha tiver problema, NADA e lancado."],
    ]:
        ws2.append(linha)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="MODELO_DEVOLUCOES.xlsx"'},
    )


# ─────────────────────────────────────────────────────────
# VALIDAÇÃO (compartilhada pela planilha e pela tela)
# ─────────────────────────────────────────────────────────

def _validate_rows(rows: List[dict], db: Session) -> dict:
    """
    Valida as linhas e resolve seller/produto.

    `rows` vem com: seller_name OU seller_id, nf_number, sku, quantity,
    returns_stock, reason.

    Devolve {"rows": [...normalizadas...], "errors": [...]}. Não grava nada e
    não decide: quem chama é que bloqueia. Erro sempre carrega o número da
    linha, senão a pessoa não sabe onde arrumar.
    """
    errors: List[str] = []
    out: List[dict] = []

    alias_map = _build_seller_alias_map(db, active=True)
    sellers_by_id: dict = {}

    # Resolve o seller de cada linha primeiro — o produto só pode ser conferido
    # depois de saber de quem é o SKU.
    resolvidos: List[dict] = []
    for i, row in enumerate(rows):
        n = i + 1
        seller = None
        seller_id = row.get("seller_id")
        seller_name = (row.get("seller_name") or "").strip()

        if seller_id:
            if seller_id not in sellers_by_id:
                sellers_by_id[seller_id] = db.query(models.Seller).filter(
                    models.Seller.id == seller_id,
                    models.Seller.active == True,  # noqa: E712
                ).first()
            seller = sellers_by_id[seller_id]
            if not seller:
                errors.append(f"Linha {n}: seller {seller_id} não encontrado ou inativo")
        elif seller_name:
            seller = alias_map.get(seller_name.lower())
            if not seller:
                errors.append(
                    f'Linha {n}: seller "{seller_name}" não foi reconhecido. '
                    f"Confira o nome cadastrado em Sellers."
                )
        else:
            errors.append(f"Linha {n}: seller não informado")

        nf = _cell_str(row.get("nf_number"))
        if not nf:
            errors.append(f"Linha {n}: NF é obrigatória")

        sku = _cell_str(row.get("sku"))
        if not sku:
            errors.append(f"Linha {n}: SKU não informado")

        qtd = _parse_qtd(row.get("quantity"))
        if qtd is None:
            errors.append(
                f'Linha {n}: quantidade inválida ("{_cell_str(row.get("quantity"))}") — '
                f"precisa ser um número inteiro maior que zero"
            )

        retorna = _parse_retorna(row.get("returns_stock"))
        if retorna is None:
            errors.append(
                f'Linha {n}: "Retorna ao estoque" inválido '
                f'("{_cell_str(row.get("returns_stock"))}") — use S/SIM/1/X ou N/NAO/0'
            )

        resolvidos.append({
            "line": n,
            "seller_id": seller.id if seller else None,
            "seller_name": (seller.trade_name or seller.name) if seller else seller_name,
            "nf_number": nf,
            "sku": sku,
            "quantity": qtd,
            "returns_stock": retorna,
            "reason": (row.get("reason") or "").strip(),
        })

    # Produto ativo por (seller_id, sku) — uma consulta por seller envolvido,
    # nunca uma por linha.
    por_seller: dict = {}
    for r in resolvidos:
        if r["seller_id"] and r["sku"]:
            por_seller.setdefault(r["seller_id"], set()).add(r["sku"])

    produtos: dict = {}
    for sid, skus in por_seller.items():
        for p in db.query(models.Product).filter(
            models.Product.seller_id == sid,
            models.Product.sku.in_(list(skus)),
            models.Product.active == True,  # noqa: E712
        ).all():
            produtos[(sid, p.sku)] = p

    vistos: dict = {}
    for r in resolvidos:
        n = r["line"]
        if r["seller_id"] and r["sku"]:
            prod = produtos.get((r["seller_id"], r["sku"]))
            if not prod:
                errors.append(
                    f'Linha {n}: SKU "{r["sku"]}" não tem produto ativo cadastrado '
                    f'no seller {r["seller_name"]}'
                )
            else:
                r["product_id"] = prod.id
                r["product_name"] = prod.name

        # Duplicidade: só é duplicata quando TUDO bate, NF inclusive. Mesmo SKU
        # e mesma quantidade em NFs diferentes é devolução legítima.
        if all(v is not None for v in (r["seller_id"], r["quantity"], r["returns_stock"])):
            chave = (
                r["seller_id"], r["nf_number"].lower(), r["sku"].lower(),
                r["quantity"], r["returns_stock"], r["reason"].lower(),
            )
            if chave in vistos:
                errors.append(
                    f"Linha {n}: idêntica à linha {vistos[chave]} "
                    f"(mesmo seller, NF, SKU, quantidade e retorno)"
                )
            else:
                vistos[chave] = n

        out.append(r)

    return {"rows": out, "errors": errors}


def _row_to_payload(r: dict) -> dict:
    """Formato devolvido ao front."""
    return {
        "line": r["line"],
        "seller_id": r["seller_id"],
        "seller_name": r["seller_name"],
        "nf_number": r["nf_number"],
        "sku": r["sku"],
        "product_name": r.get("product_name"),
        "quantity": r["quantity"],
        "returns_stock": r["returns_stock"],
        "reason": r["reason"],
    }


# ─────────────────────────────────────────────────────────
# ANÁLISE DA PLANILHA (não grava nada)
# ─────────────────────────────────────────────────────────

@router.post("/analyze")
def analyze_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Lê a planilha e devolve as linhas normalizadas + os erros.
    NÃO grava nada. Endpoint síncrono de propósito (`def`, não `async def`):
    ler Excel é trabalho de CPU e travaria o event loop.
    """
    nome = (file.filename or "").lower()
    if not nome.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Envie um arquivo .xlsx ou .xls")

    from openpyxl import load_workbook

    try:
        conteudo = file.file.read()
        wb = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Não consegui ler o arquivo: {exc}")

    ws = wb["DEVOLUCOES"] if "DEVOLUCOES" in wb.sheetnames else wb.worksheets[0]

    linhas = list(ws.iter_rows(values_only=True))
    if len(linhas) < 2:
        raise HTTPException(status_code=422, detail="A planilha está vazia")

    # A primeira linha é o cabeçalho; linhas totalmente vazias são ignoradas.
    brutas: List[dict] = []
    for raw in linhas[1:]:
        celulas = list(raw) + [None] * 6
        if all(c is None or str(c).strip() == "" for c in celulas[:6]):
            continue
        brutas.append({
            "seller_name": _cell_str(celulas[0]),
            "nf_number": _cell_str(celulas[1]),
            "sku": _cell_str(celulas[2]),
            "quantity": celulas[3],
            "returns_stock": celulas[4],
            "reason": _cell_str(celulas[5]),
        })

    if not brutas:
        raise HTTPException(status_code=422, detail="Nenhuma linha preenchida na planilha")

    resultado = _validate_rows(brutas, db)
    linhas_ok = [_row_to_payload(r) for r in resultado["rows"]]

    return {
        "total": len(linhas_ok),
        "rows": linhas_ok,
        "errors": resultado["errors"],
        "can_submit": len(resultado["errors"]) == 0,
        "returning": sum(1 for r in linhas_ok if r["returns_stock"] is True),
        "not_returning": sum(1 for r in linhas_ok if r["returns_stock"] is False),
    }


# ─────────────────────────────────────────────────────────
# LANÇAMENTO (planilha conferida OU tabelinha da tela)
# ─────────────────────────────────────────────────────────

@router.post("/lancar", status_code=201)
def lancar_devolucoes(
    body: dict,
    current_user: models.User = Depends(require_manager_or_above),
    db: Session = Depends(get_db),
):
    """
    Grava as devoluções. TUDO-OU-NADA: revalida e, havendo qualquer erro,
    devolve 422 com a lista e não grava nada.
    """
    rows = body.get("rows") or []
    if not rows:
        raise HTTPException(status_code=422, detail="Nenhuma linha enviada")

    resultado = _validate_rows(rows, db)
    if resultado["errors"]:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Nada foi lançado — corrija os problemas abaixo.",
                "errors": resultado["errors"],
            },
        )

    agora = now_brasilia()
    hoje = today_brasilia()
    quem = current_user.name or current_user.email
    quando = agora.strftime("%d/%m/%Y")

    lancados = 0
    ignorados = 0

    try:
        for r in resultado["rows"]:
            if r["returns_stock"]:
                obs = (
                    f"DEVOLUÇÃO — NF {r['nf_number']}, {r['quantity']} un retornaram "
                    f"ao estoque. Lançado por {quem} em {quando}."
                )
                db.add(models.StockMovement(
                    seller_id=r["seller_id"],
                    product_id=r.get("product_id"),
                    sku=r["sku"],
                    product_name=r.get("product_name") or r["sku"],
                    movement_date=hoje,
                    movement_type=models.MovementType.IN,
                    quantity=r["quantity"],
                    nf_number=r["nf_number"][:20],
                    nature="Devolução",
                    # order_id fica VAZIO de propósito — ver o cabeçalho do módulo.
                    observation=obs,
                    operator_id=current_user.id,
                    created_at=agora,
                ))
                update_stock_position(
                    seller_id=r["seller_id"],
                    sku=r["sku"],
                    product_name=r.get("product_name") or r["sku"],
                    movement_type=models.MovementType.IN,
                    quantity=r["quantity"],
                    db=db,
                )
                lancados += 1
            else:
                db.add(models.AuditLog(
                    entity_type="Devolucao",
                    entity_id=r["seller_id"],
                    action="CREATE",
                    detail=json.dumps({
                        "retornou_ao_estoque": False,
                        "seller_id": r["seller_id"],
                        "seller": r["seller_name"],
                        "nf": r["nf_number"],
                        "sku": r["sku"],
                        "produto": r.get("product_name"),
                        "quantidade": r["quantity"],
                        "motivo": r["reason"] or None,
                    }, ensure_ascii=False),
                    user_id=current_user.id,
                    timestamp=agora,
                ))
                ignorados += 1

        db.add(models.AuditLog(
            entity_type="Devolucao",
            action="IMPORT",
            detail=json.dumps({
                "linhas": len(resultado["rows"]),
                "retornaram_ao_estoque": lancados,
                "nao_retornaram": ignorados,
            }, ensure_ascii=False),
            user_id=current_user.id,
            timestamp=agora,
        ))
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Falha ao lançar devoluções: {exc}")

    return {
        "total": len(resultado["rows"]),
        "returned_to_stock": lancados,
        "not_returned": ignorados,
    }
