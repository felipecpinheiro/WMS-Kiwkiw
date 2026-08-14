r"""
Reconciliação de estoque a partir da planilha "ESTOQUE <seller>.xlsx" do cliente.

Contexto: o cliente confia mais na planilha dele do que no WMS e quer que o
estoque do WMS vire uma cópia fiel dela (aba DETALHADO = histórico linha a
linha, aba ESTOQUE = resumo por SKU com Estoque Inicial/Entrada/Saída/Final).

Isto NÃO é uma reconciliação incremental (comparar e só completar o que
falta) — é um CORTE por SKU: apaga tudo que existe hoje em stock_movements
para aquele seller+SKU e recria do zero, exatamente como está na planilha,
preservando as datas originais (mesmo que sejam de anos atrás) para efeito
de auditoria futura.

Regras acertadas com o dono do sistema (13/08/2026), NÃO alterar sem
conversar antes:

1. SKU que só aparece no DETALHADO mas não na aba resumo ESTOQUE
   ("descontinuado") -> processa normalmente, sem meta pra validar contra.
2. SKU cuja soma do DETALHADO não bate com (Estoque Final - Estoque Inicial)
   da aba resumo -> NÃO É TOCADO. Fica exatamente como está hoje no banco,
   e entra no relatório para decisão manual, caso a caso.
3. SKU sem produto ativo cadastrado no seller -> NÃO É TOCADO (o dono do
   sistema já cadastrou todos os produtos ativos; se não existe, o SKU
   provavelmente não existe mais). Só entra no relatório final como aviso.
4. Linha com data implausível (ano < SUSPICIOUS_YEAR_MIN, tipicamente erro
   de digitação no Excel) -> a LINHA (não o SKU inteiro) é excluída, e
   reportada. Nunca resolver isso sozinho quando houver dúvida real — o
   dono do sistema pediu para ser chamado caso a caso.
5. Nenhuma proteção para movimentos amarrados a pedidos reais (order_id) —
   decisão consciente e explícita do dono do sistema, aceitando o risco
   documentado (ver reverse_stock_for_order / correção de 13/08/2026 para
   o que foi feito para mitigar o efeito colateral em cancelamentos futuros).
6. O alvo final é: WMS bater exatamente com a coluna "Estoque Final" da aba
   ESTOQUE para TODO SKU dela. O relatório final mostra isso SKU a SKU,
   inclusive os que ficaram de fora (para decisão manual).

Uso:
    # Só analisa e imprime o relatório — não escreve nada no banco
    python -m backend.scripts.reconcile_seller_stock --seller-id 12 --file "C:\...\ESTOQUE Feel.xlsx"

    # Executa de verdade (apaga+recria só os SKUs "OK") — precisa dos dois flags
    python -m backend.scripts.reconcile_seller_stock --seller-id 12 --file "C:\...\ESTOQUE Feel.xlsx" --execute --confirm

Por padrão roda contra o banco apontado por DATABASE_URL (o mesmo do
main.py) — local (SQLite) se a variável não estiver setada. Para produção,
setar DATABASE_URL para a URL do Postgres do Railway antes de rodar.
"""
import argparse
import json
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from backend.database import SessionLocal
from backend import models
from backend.routers.inventory import _parse_history_excel
from backend.services.stock_manager import calculate_stock_level
from backend.timezone_utils import now_brasilia

SUSPICIOUS_YEAR_MIN = 2018
RECONCILE_TAG_TEMPLATE = "[RECONCILIAÇÃO PLANILHA CLIENTE — {ts}] "


def _parse_resumo_sheet(file_path: str) -> dict:
    """
    Lê a aba ESTOQUE (resumo por SKU). Layout observado:
    linha 6 = cabeçalho, dados a partir da linha 7.
    Colunas: C=SKU, D=Nome, E=Estoque Inicial, F=Entrada, G=Saída, H=Estoque Final.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().upper() == "ESTOQUE":
            sheet = wb[name]
            break
    if sheet is None:
        raise ValueError("Aba 'ESTOQUE' (resumo) não encontrada no arquivo.")

    resumo = {}
    for row in sheet.iter_rows(min_row=7, values_only=True):
        sku = row[2] if len(row) > 2 else None
        if not sku:
            continue
        sku = str(sku).strip()
        ini = row[4] if len(row) > 4 else None
        ent = row[5] if len(row) > 5 else None
        sai = row[6] if len(row) > 6 else None
        fim = row[7] if len(row) > 7 else None
        if ini is None or ent is None or sai is None or fim is None:
            continue
        resumo[sku] = {
            "estoque_inicial": int(ini),
            "entrada": int(ent),
            "saida": int(sai),
            "estoque_final": int(fim),
        }
    wb.close()
    return resumo


def analyze(seller_id: int, file_path: str, db) -> dict:
    with open(file_path, "rb") as f:
        file_bytes = f.read()

    detail_rows = _parse_history_excel(file_bytes)
    resumo = _parse_resumo_sheet(file_path)

    # ── Separa linhas com data implausível (erro de digitação) ────────────
    suspicious_rows = []
    clean_rows = []
    for r in detail_rows:
        try:
            year = date.fromisoformat(r["movement_date"]).year
        except Exception:
            year = None
        if year is None or year < SUSPICIOUS_YEAR_MIN or year > now_brasilia().year:
            suspicious_rows.append(r)
        else:
            clean_rows.append(r)

    # ── Agrupa por SKU ──────────────────────────────────────────────────────
    by_sku = defaultdict(list)
    for r in clean_rows:
        by_sku[r["sku"]].append(r)

    # ── SKUs sem produto ativo cadastrado ──────────────────────────────────
    all_skus = set(by_sku.keys())
    registered = {
        p.sku for p in db.query(models.Product.sku).filter(
            models.Product.seller_id == seller_id,
            models.Product.sku.in_(list(all_skus)),
            models.Product.active == True,  # noqa: E712
        ).all()
    }
    missing_product_skus = sorted(all_skus - registered)

    # ── Validação SKU a SKU contra o resumo ────────────────────────────────
    ok_skus = []
    divergent_skus = []
    discontinued_skus = []  # só existem no DETALHADO, não no resumo

    current_positions = {
        p.sku: p for p in db.query(models.StockPosition).filter(
            models.StockPosition.seller_id == seller_id,
        ).all()
    }

    for sku, rows in by_sku.items():
        if sku in missing_product_skus:
            continue  # relatado à parte, nunca processado

        net_in = sum(r["quantity"] for r in rows if r["movement_type"] == "Entrada")
        net_out = sum(r["quantity"] for r in rows if r["movement_type"] == "Saida")
        net = net_in - net_out

        target = resumo.get(sku)
        current_wms = current_positions.get(sku)
        current_stock_now = current_wms.current_stock if current_wms else 0

        if target is None:
            discontinued_skus.append({
                "sku": sku,
                "rows": len(rows),
                "net_detalhado": net,
                "current_wms_stock": current_stock_now,
                "would_be_stock": net,  # initial_stock=0 pra quem não tem resumo
            })
            continue

        net_resumo = target["estoque_final"] - target["estoque_inicial"]
        if net != net_resumo:
            divergent_skus.append({
                "sku": sku,
                "net_detalhado": net,
                "net_resumo_esperado": net_resumo,
                "diferenca": net - net_resumo,
                "current_wms_stock": current_stock_now,
                "planilha_estoque_final": target["estoque_final"],
            })
            continue

        ok_skus.append({
            "sku": sku,
            "rows": len(rows),
            "estoque_inicial": target["estoque_inicial"],
            "net_detalhado": net,
            "would_be_stock": target["estoque_inicial"] + net,
            "planilha_estoque_final": target["estoque_final"],
            "current_wms_stock": current_stock_now,
        })

    # ── Relatório-mestre: TODO SKU do resumo, bate ou não ──────────────────
    master = []
    ok_by_sku = {x["sku"]: x for x in ok_skus}
    divergent_by_sku = {x["sku"]: x for x in divergent_skus}
    for sku, target in sorted(resumo.items()):
        current_wms = current_positions.get(sku)
        current_stock_now = current_wms.current_stock if current_wms else 0
        if sku in missing_product_skus:
            master.append({
                "sku": sku, "status": "SEM_PRODUTO_CADASTRADO",
                "wms_atual": current_stock_now,
                "wms_apos_script": current_stock_now,
                "planilha_estoque_final": target["estoque_final"],
                "bate": current_stock_now == target["estoque_final"],
            })
        elif sku in ok_by_sku:
            would_be = ok_by_sku[sku]["would_be_stock"]
            master.append({
                "sku": sku, "status": "OK",
                "wms_atual": current_stock_now,
                "wms_apos_script": would_be,
                "planilha_estoque_final": target["estoque_final"],
                "bate": would_be == target["estoque_final"],
            })
        elif sku in divergent_by_sku:
            master.append({
                "sku": sku, "status": "DIVERGENTE_NAO_TOCADO",
                "wms_atual": current_stock_now,
                "wms_apos_script": current_stock_now,
                "planilha_estoque_final": target["estoque_final"],
                "bate": current_stock_now == target["estoque_final"],
            })
        else:
            # SKU no resumo mas sem nenhuma linha no DETALHADO (nunca visto)
            master.append({
                "sku": sku, "status": "SEM_LINHAS_NO_DETALHADO",
                "wms_atual": current_stock_now,
                "wms_apos_script": current_stock_now,
                "planilha_estoque_final": target["estoque_final"],
                "bate": current_stock_now == target["estoque_final"],
            })

    return {
        "total_linhas_arquivo": len(detail_rows),
        "linhas_suspeitas": suspicious_rows,
        "skus_ok": ok_skus,
        "skus_divergentes": divergent_skus,
        "skus_descontinuados": discontinued_skus,
        "skus_sem_produto": missing_product_skus,
        "relatorio_mestre": master,
        "_by_sku_clean_rows": by_sku,  # uso interno do execute()
    }


def execute(seller_id: int, analysis: dict, db) -> dict:
    """Aplica o corte: apaga+recria stock_movements só dos SKUs 'OK' e 'descontinuados'."""
    ts = now_brasilia().strftime("%d/%m/%Y %H:%M")
    tag = RECONCILE_TAG_TEMPLATE.format(ts=ts)

    processable = {x["sku"]: x["estoque_inicial"] for x in analysis["skus_ok"]}
    processable.update({x["sku"]: 0 for x in analysis["skus_descontinuados"]})
    by_sku = analysis["_by_sku_clean_rows"]

    nomes_por_sku = {
        p.sku: p.name for p in db.query(models.Product.sku, models.Product.name).filter(
            models.Product.seller_id == seller_id,
        ).all()
    }

    now_ts = now_brasilia()
    movements_deleted = 0
    movements_inserted = 0

    try:
        for sku, initial_stock in processable.items():
            rows = by_sku.get(sku, [])

            deleted = db.query(models.StockMovement).filter(
                models.StockMovement.seller_id == seller_id,
                models.StockMovement.sku == sku,
            ).delete(synchronize_session=False)
            movements_deleted += deleted

            total_in = 0
            total_out = 0
            product_name = nomes_por_sku.get(sku, sku)
            for r in rows:
                is_in = r["movement_type"] == "Entrada"
                mt = models.MovementType.IN if is_in else models.MovementType.OUT
                mov_date = date.fromisoformat(r["movement_date"])
                nf_date = date.fromisoformat(r["nf_date"]) if r.get("nf_date") else None
                obs = tag + (r["observation"] or "").strip()
                name_from_sheet = r["product_name_from_sheet"] or product_name

                db.add(models.StockMovement(
                    seller_id=seller_id,
                    sku=sku,
                    product_name=nomes_por_sku.get(sku) or name_from_sheet,
                    movement_date=mov_date,
                    movement_type=mt,
                    quantity=r["quantity"],
                    adjusted_quantity=r["quantity"],
                    nf_number=r["nf_number"],
                    nf_date=nf_date,
                    observation=obs,
                    created_at=now_ts,
                ))
                movements_inserted += 1
                if is_in:
                    total_in += r["quantity"]
                else:
                    total_out += r["quantity"]

            position = db.query(models.StockPosition).filter(
                models.StockPosition.seller_id == seller_id,
                models.StockPosition.sku == sku,
            ).first()
            current_stock = initial_stock + total_in - total_out
            if position is None:
                position = models.StockPosition(seller_id=seller_id, sku=sku)
                db.add(position)
            position.product_name = product_name
            position.initial_stock = initial_stock
            position.total_in = total_in
            position.total_out = total_out
            position.current_stock = current_stock
            position.level = calculate_stock_level(current_stock)
            position.updated_at = now_ts

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "skus_processados": len(processable),
        "movimentos_apagados": movements_deleted,
        "movimentos_inseridos": movements_inserted,
    }


def print_report(analysis: dict):
    print(f"\nTotal de linhas lidas do DETALHADO: {analysis['total_linhas_arquivo']}")
    print(f"Linhas com data suspeita (ano < {SUSPICIOUS_YEAR_MIN} ou no futuro): {len(analysis['linhas_suspeitas'])}")
    for r in analysis["linhas_suspeitas"][:20]:
        print(f"    SKU={r['sku']!r} data={r['movement_date']!r} qtd={r['quantity']} obs={r['observation']!r}")
    if len(analysis["linhas_suspeitas"]) > 20:
        print(f"    ... e mais {len(analysis['linhas_suspeitas']) - 20}")

    print(f"\nSKUs sem produto cadastrado (NAO processados): {len(analysis['skus_sem_produto'])}")
    for sku in analysis["skus_sem_produto"][:30]:
        print(f"    {sku}")
    if len(analysis["skus_sem_produto"]) > 30:
        print(f"    ... e mais {len(analysis['skus_sem_produto']) - 30}")

    print(f"\nSKUs divergentes (NAO processados, decisao manual): {len(analysis['skus_divergentes'])}")
    for x in analysis["skus_divergentes"]:
        print(f"    {x['sku']}: soma DETALHADO={x['net_detalhado']} esperado(resumo)={x['net_resumo_esperado']} "
              f"diff={x['diferenca']} | WMS hoje={x['current_wms_stock']} planilha Estoque Final={x['planilha_estoque_final']}")

    print(f"\nSKUs descontinuados (sem meta no resumo, SERAO processados): {len(analysis['skus_descontinuados'])}")
    print(f"SKUs OK (batem com o resumo, SERAO processados): {len(analysis['skus_ok'])}")

    print("\n=== RELATÓRIO MESTRE (compara com a coluna 'Estoque Final' da planilha) ===")
    mismatches = [m for m in analysis["relatorio_mestre"] if not m["bate"]]
    print(f"Total de SKUs na planilha (aba ESTOQUE): {len(analysis['relatorio_mestre'])}")
    print(f"Vão bater 100% após o script: {len(analysis['relatorio_mestre']) - len(mismatches)}")
    print(f"Vão continuar diferentes (precisam de decisão sua): {len(mismatches)}")
    for m in mismatches:
        print(f"    [{m['status']}] {m['sku']}: WMS atual={m['wms_atual']} -> apos script={m['wms_apos_script']} "
              f"| meta planilha={m['planilha_estoque_final']}")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--seller-id", type=int, required=True)
    parser.add_argument("--file", type=str, required=True)
    parser.add_argument("--execute", action="store_true", help="Aplica de verdade (padrão: só analisa)")
    parser.add_argument("--confirm", action="store_true", help="Confirmação extra obrigatória junto com --execute")
    parser.add_argument("--json-out", type=str, default=None, help="Salva o relatório completo em JSON nesse caminho")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        seller = db.query(models.Seller).filter(models.Seller.id == args.seller_id).first()
        if not seller:
            print(f"ERRO: seller_id={args.seller_id} não encontrado.")
            sys.exit(1)
        print(f"Seller: {seller.trade_name or seller.name} (id={seller.id})")

        analysis = analyze(args.seller_id, args.file, db)
        print_report(analysis)

        if args.json_out:
            out = {k: v for k, v in analysis.items() if k != "_by_sku_clean_rows"}
            with open(args.json_out, "w", encoding="utf-8") as f:
                json.dump(out, f, ensure_ascii=False, indent=2, default=str)
            print(f"\nRelatório salvo em {args.json_out}")

        if args.execute:
            if not args.confirm:
                print("\n--execute exige também --confirm. Nada foi alterado.")
                sys.exit(1)
            print("\nExecutando corte (apagar + recriar SKUs OK/descontinuados)...")
            result = execute(args.seller_id, analysis, db)
            print(f"SKUs processados: {result['skus_processados']}")
            print(f"Movimentos apagados: {result['movimentos_apagados']}")
            print(f"Movimentos inseridos: {result['movimentos_inseridos']}")
        else:
            print("\n(Modo análise — nada foi alterado no banco. Use --execute --confirm para aplicar.)")
    finally:
        db.close()


if __name__ == "__main__":
    main()
