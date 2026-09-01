"""
WMS Kiwkiw - Documentos de faturamento (PDF, Excel, zip)
=======================================================
Consomem o dict devolvido por `routers/billing._build_payload`
(mesma forma de `billing_calc.compute_live`).

Sem campos fiscais (CNPJ / nº / razão social) — decisão do dono do sistema.
Identidade visual Kiwkiw: roxo #7B63E8.
"""

import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

PURPLE = colors.HexColor("#7B63E8")
DARK = colors.HexColor("#14122A")
LIGHT = colors.HexColor("#F0EEFF")
AMBER = colors.HexColor("#FBEFDD")
BORDER = colors.HexColor("#D0CCEE")


def brl(n) -> str:
    try:
        v = float(n or 0.0)
    except (TypeError, ValueError):
        v = 0.0
    s = f"{v:,.2f}"
    return "R$ " + s.replace(",", "_").replace(".", ",").replace("_", ".")


def _month_label(ref_month: str) -> str:
    y, m = ref_month.split("-")
    return f"{m}/{y}"


# ── PDF ─────────────────────────────────────────────────────────────────────

def invoice_pdf_bytes(p: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        topMargin=1.4 * cm, bottomMargin=1.4 * cm,
        title=f"Fatura {p['seller_name']} {p['ref_month']}",
    )
    title = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16,
                           textColor=PURPLE, leading=20)
    sub = ParagraphStyle("s", fontName="Helvetica", fontSize=9,
                         textColor=colors.HexColor("#6B6490"), leading=12)
    h2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=10.5,
                        textColor=DARK, leading=14, spaceBefore=10, spaceAfter=4)
    cell = ParagraphStyle("c", fontName="Helvetica", fontSize=8, leading=10)
    cellr = ParagraphStyle("cr", parent=cell, alignment=TA_RIGHT)
    hdr = ParagraphStyle("hd", fontName="Helvetica-Bold", fontSize=8,
                         textColor=colors.white, leading=10, alignment=TA_CENTER)

    el = []
    el.append(Paragraph("Fatura de serviços — Kiwkiw Fulfillment", title))
    st = {"open": "Em aberto", "closed": "Fechado"}.get(p["status"], p["status"])
    el.append(Paragraph(
        f"Seller <b>{p['seller_name']}</b> &nbsp;·&nbsp; referência "
        f"<b>{_month_label(p['ref_month'])}</b> &nbsp;·&nbsp; {st}", sub))
    el.append(Spacer(1, 8))

    pr = p["params"]
    par_rows = [
        ["Preço unitário / manuseio B2C", brl(pr["preco_unitario"]),
         "Nº mínimo de pedidos", str(pr["min_pedidos"])],
        ["Manuseio B2B", brl(pr["manuseio_b2b"]),
         "Valor caixa B2B", brl(pr["valor_caixa_b2b"])],
        ["Adicional por produto B2B", brl(pr["adic_produto_b2b"]),
         "Franquia de produtos B2B", str(pr["franquia_produtos_b2b"])],
        ["É B2B a partir de (itens)", str(pr["limite_itens_b2b"]),
         "Tipos de caixa inclusos", pr["tipos_caixa_inclusos"] or "—"],
        ["Cota de caixas / mês", str(pr["cota_caixas_mes"]),
         "Cubagem medida (m³)", str(p["cubagem_m3"])],
        ["Franquia grátis (m³)", str(pr["franquia_m3"]),
         "Preço por m³ adicional", brl(pr["preco_m3"])],
        ["Cobrar seguro", "Sim" if pr["seguro_incluso"] else "Não",
         "Alíquota do seguro (%)", str(pr["aliquota_seguro"])],
        ["Valor segurado", brl(p["valor_segurado"]),
         "Armazenagem inclusa", "Sim" if pr["armazenagem_inclusa"] else "Não"],
    ]
    el.append(Paragraph("Parâmetros do mês", h2))
    t = Table([[Paragraph(c, cell) for c in r] for r in par_rows],
              colWidths=[5.2 * cm, 3.4 * cm, 5.2 * cm, 3.4 * cm])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("BACKGROUND", (0, 0), (0, -1), LIGHT),
        ("BACKGROUND", (2, 0), (2, -1), LIGHT),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    el.append(t)

    # fatura final
    f = p["fatura"]
    el.append(Paragraph("Fatura final", h2))
    fin_rows = [
        [Paragraph("Componente", hdr), Paragraph("B2C", hdr), Paragraph("B2B", hdr)],
        ["Mínimo mensal", brl(f["b2c_min"]), brl(f["b2b_min"])],
        ["Seguro", brl(f["seguro"]), "—"],
        ["Armazenagem", brl(f["armazenagem"]), "—"],
        ["Linhas avulsas", brl(f["avulsos"]), "—"],
        ["Subtotal", brl(f["subtotal_b2c"]), brl(f["subtotal_b2b"])],
        ["TOTAL GERAL", brl(f["total_geral"]), ""],
    ]
    ft = Table(fin_rows, colWidths=[8.0 * cm, 4.6 * cm, 4.6 * cm])
    ft.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 5), (-1, 6), "Helvetica-Bold"),
        ("BACKGROUND", (0, 5), (-1, 5), LIGHT),
        ("BACKGROUND", (0, 6), (-1, 6), colors.HexColor("#ECE7FC")),
        ("SPAN", (1, 6), (2, 6)),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
    ]))
    el.append(ft)
    if f.get("min_atingiu_piso"):
        el.append(Spacer(1, 4))
        el.append(Paragraph(
            f"Mínimo mensal B2C: soma real {brl(f['soma_real_b2c'])} &lt; "
            f"piso {brl(f['floor_b2c'])} → cobra-se o maior.", sub))

    # avulsos detalhados
    if p["adjustments"]:
        el.append(Paragraph("Linhas avulsas", h2))
        av = [[Paragraph("Descrição", hdr), Paragraph("Obs.", hdr), Paragraph("Valor", hdr)]]
        for a in p["adjustments"]:
            sign = "−" if (a["sign"] or 1) < 0 else "+"
            av.append([Paragraph(a["descricao"], cell), Paragraph(a["obs"], cell),
                       Paragraph(f"{sign} {brl(a['valor'])}", cellr)])
        at = Table(av, colWidths=[8.0 * cm, 5.2 * cm, 4.0 * cm])
        at.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ]))
        el.append(at)

    # listas
    _append_list(el, "Notas fiscais B2C", h2, hdr, cell, cellr,
                 ["Data", "NF", "Cx", "Itens", "Adic. caixa", "Adic. man.", "Manus.", "Total"],
                 [[_d(l["order_date"]), l["nf_number"], l["box"] or "—", l["itens"],
                   brl(l["adic_caixa"]), brl(l.get("adic_manual") or 0.0),
                   brl(l["manuseio"]), brl(l["total"])]
                  for l in p["b2c_lines"]],
                 p["soma_b2c"])
    _append_list(el, "Notas fiscais B2B", h2, hdr, cell, cellr,
                 ["Data", "NF", "Itens", "Cx B2B", "Manus.", "Ad.prod", "Adic.", "Total"],
                 [[_d(l["order_date"]), l["nf_number"], l["itens"],
                   brl(l["valor_caixa_b2b"]), brl(l["manuseio_b2b"]),
                   brl(l.get("adic_produto") or 0.0),
                   brl(l["b2b_adicional"]), brl(l["total"])]
                  for l in p["b2b_lines"]],
                 p["soma_b2b"])

    doc.build(el)
    return buf.getvalue()


def _d(iso):
    if not iso:
        return "—"
    return iso[8:10] + "/" + iso[5:7]


def _append_list(el, title, h2, hdr, cell, cellr, headers, rows, total):
    el.append(Paragraph(f"{title} ({len(rows)})", h2))
    data = [[Paragraph(h, hdr) for h in headers]]
    for r in rows:
        data.append([Paragraph(str(c), cellr if i > 1 else cell) for i, c in enumerate(r)])
    data.append(["", "", "", "", "", "Soma", brl(total)] if len(headers) == 7 else [])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
        ("GRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), LIGHT),
    ]))
    el.append(t)
    el.append(Spacer(1, 6))


# ── Excel ──────────────────────────────────────────────────────────────────

def invoice_xlsx_bytes(p: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fatura"
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="7B63E8")
    hdr_font = Font(bold=True, color="FFFFFF")
    amber = PatternFill("solid", fgColor="FDF0D8")

    r = 1
    ws.cell(r, 1, f"Fatura — {p['seller_name']} — {_month_label(p['ref_month'])}").font = Font(bold=True, size=13)
    r += 1
    ws.cell(r, 1, "Situação").font = bold
    ws.cell(r, 2, "Fechado" if p["status"] == "closed" else "Em aberto")
    r += 2

    ws.cell(r, 1, "PARÂMETROS").font = bold
    r += 1
    pr = p["params"]
    for k, v in [
        ("Preço unitário / manuseio B2C", pr["preco_unitario"]),
        ("Nº mínimo de pedidos", pr["min_pedidos"]),
        ("Manuseio B2B", pr["manuseio_b2b"]),
        ("Valor caixa B2B", pr["valor_caixa_b2b"]),
        ("Adicional por produto B2B", pr["adic_produto_b2b"]),
        ("Franquia de produtos B2B", pr["franquia_produtos_b2b"]),
        ("É B2B a partir de (itens)", pr["limite_itens_b2b"]),
        ("Tipos de caixa inclusos", pr["tipos_caixa_inclusos"] or "—"),
        ("Cota de caixas / mês", pr["cota_caixas_mes"]),
        ("Franquia grátis (m³)", pr["franquia_m3"]),
        ("Preço por m³ adicional", pr["preco_m3"]),
        ("Cubagem medida (m³)", p["cubagem_m3"]),
        ("Cobrar seguro", "Sim" if pr["seguro_incluso"] else "Não"),
        ("Alíquota do seguro (%)", pr["aliquota_seguro"]),
        ("Valor segurado", p["valor_segurado"]),
        ("Armazenagem inclusa", "Sim" if pr["armazenagem_inclusa"] else "Não"),
    ]:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    r += 1

    ws.cell(r, 1, "FATURA FINAL").font = bold
    r += 1
    f = p["fatura"]
    for c, txt in ((1, "Componente"), (2, "B2C"), (3, "B2B")):
        cell = ws.cell(r, c, txt)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r += 1
    for lbl, a, b in [
        ("Mínimo mensal", f["b2c_min"], f["b2b_min"]),
        ("Seguro", f["seguro"], None),
        ("Armazenagem", f["armazenagem"], None),
        ("Linhas avulsas", f["avulsos"], None),
        ("Subtotal", f["subtotal_b2c"], f["subtotal_b2b"]),
        ("TOTAL GERAL", f["total_geral"], None),
    ]:
        ws.cell(r, 1, lbl)
        ws.cell(r, 2, a)
        if b is not None:
            ws.cell(r, 3, b)
        r += 1
    r += 1

    if p["adjustments"]:
        ws.cell(r, 1, "LINHAS AVULSAS").font = bold
        r += 1
        for a in p["adjustments"]:
            ws.cell(r, 1, a["descricao"])
            ws.cell(r, 2, a["obs"])
            ws.cell(r, 3, (a["sign"] or 1) * (a["valor"] or 0.0))
            r += 1
        r += 1

    r = _xlsx_list(ws, r, "NOTAS FISCAIS B2C",
                   ["Data", "NF", "Caixa", "Itens", "Adic. caixa", "Adic. manual", "Manuseio", "Total", "Sem caixa"],
                   [[l["order_date"], l["nf_number"], l["box"], l["itens"],
                     l["adic_caixa"], l.get("adic_manual") or 0.0, l["manuseio"], l["total"],
                     "SIM" if l["sem_caixa"] else ""]
                    for l in p["b2c_lines"]], hdr_fill, hdr_font, amber, sem_caixa_col=9)
    r = _xlsx_list(ws, r + 1, "NOTAS FISCAIS B2B",
                   ["Data", "NF", "Itens", "Caixa B2B", "Manuseio B2B",
                    "Adic. produto", "Adicional", "Total"],
                   [[l["order_date"], l["nf_number"], l["itens"], l["valor_caixa_b2b"],
                     l["manuseio_b2b"], l.get("adic_produto") or 0.0,
                     l["b2b_adicional"], l["total"]]
                    for l in p["b2b_lines"]], hdr_fill, hdr_font, amber)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _xlsx_list(ws, r, title, headers, rows, hdr_fill, hdr_font, amber, sem_caixa_col=None):
    from openpyxl.styles import Font
    ws.cell(r, 1, title).font = Font(bold=True)
    r += 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r += 1
    for row in rows:
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
        if sem_caixa_col and row[sem_caixa_col - 1] == "SIM":
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = amber
        r += 1
    return r


# ── consolidado ────────────────────────────────────────────────────────────

def consolidated_xlsx_bytes(ref_month: str, rows: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    hdr_fill = PatternFill("solid", fgColor="7B63E8")
    hdr_font = Font(bold=True, color="FFFFFF")

    ws.cell(1, 1, f"Consolidado de faturamento — {_month_label(ref_month)}").font = Font(bold=True, size=13)
    headers = ["Seller", "Ativo", "NFs", "B2C", "B2B", "Seguro", "Armazenagem",
               "Avulsos", "Total", "Situação"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r = 4
    tot = 0.0
    for row in rows:
        ws.cell(r, 1, row["seller_name"])
        ws.cell(r, 2, "Sim" if row["active"] else "Não")
        ws.cell(r, 3, row["nf_count"])
        ws.cell(r, 4, row["b2c"])
        ws.cell(r, 5, row["b2b"])
        ws.cell(r, 6, row["seguro"])
        ws.cell(r, 7, row["armazenagem"])
        ws.cell(r, 8, row["avulsos"])
        ws.cell(r, 9, row["total"])
        ws.cell(r, 10, row["status"])
        tot += row["total"]
        r += 1
    ws.cell(r, 1, f"{len(rows)} sellers").font = Font(bold=True)
    ws.cell(r, 9, round(tot, 2)).font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
