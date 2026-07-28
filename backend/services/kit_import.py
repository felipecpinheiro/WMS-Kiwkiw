"""
WMS Kiwkiw - Importação de kits a partir da planilha de tratamento (.xlsx/.xlsm)

Layout de referência — aba "CADASTRO KITS" da planilha TRATAMENTO KITS:

    linha 1   A=CLIENTE | B..C mesclado "DE" | D.. mesclado "PARA"
    linha 2   D="PRODUTO 1" | G="PRODUTO 2" | ... (um rótulo a cada 3 colunas)
    linha 3   B="SKU Kit" | C="Nome" | depois trincas SKU | NOME | QUANTIDADE
    linha 4+  dados

As colunas fixas são A (cliente), B (SKU do kit) e C (nome do kit). A partir de D
vêm trincas de 3 colunas por componente. A quantidade de trincas é detectada pela
largura da planilha — NÃO é fixa: a planilha de referência comporta 11 componentes
e um kit real usa todos os 11.
"""

from __future__ import annotations

import unicodedata
from typing import Optional

from openpyxl import load_workbook

SHEET_NAME = "CADASTRO KITS"
COL_CLIENTE = 1          # A
COL_KIT_SKU = 2          # B
COL_KIT_NOME = 3         # C
FIRST_COMPONENT_COL = 4  # D
COLS_PER_COMPONENT = 3   # SKU | NOME | QUANTIDADE


def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _norm(v) -> str:
    """Normaliza texto para comparação: sem acento, minúsculo, sem espaços nas pontas."""
    return _strip_accents(str(v or "").strip().lower())


def _clean_sku(v) -> str:
    """
    Converte a célula em SKU textual.

    O Excel devolve SKU só de dígitos como número: 3 (int) ou 3.0 (float).
    A planilha de referência tem 3 casos assim (HAIRCEUTICO: 1, 3 e 6).
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v).strip()
    return str(v).strip()


def _clean_qty(v):
    """
    Devolve (quantidade, erro). Quantidade vazia NÃO vira 1: bloqueia a linha.
    Quantidade de kit multiplica direto no estoque — chutar aqui vira erro de
    inventário silencioso lá na frente.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None, "quantidade em branco"
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None, f"quantidade inválida ({v!r})"
    if not f.is_integer():
        return None, f"quantidade fracionada ({v!r})"
    q = int(f)
    if q <= 0:
        return None, f"quantidade menor ou igual a zero ({q})"
    return q, None


def _find_sheet(wb):
    """Procura a aba CADASTRO KITS; ignora acento/caixa. Sem ela, erro explícito."""
    alvo = _norm(SHEET_NAME)
    for ws in wb.worksheets:
        if _norm(ws.title) == alvo:
            return ws
    disponiveis = ", ".join(f"'{ws.title}'" for ws in wb.worksheets)
    raise ValueError(
        f"A planilha não tem a aba '{SHEET_NAME}'. Abas encontradas: {disponiveis}"
    )


def _find_header_row(ws) -> int:
    """Linha de cabeçalho = a que tem 'SKU Kit' na coluna B (procura nas 15 primeiras)."""
    limite = min(ws.max_row, 15)
    for r in range(1, limite + 1):
        if _norm(ws.cell(row=r, column=COL_KIT_SKU).value) == _norm("SKU Kit"):
            return r
    raise ValueError(
        "Não encontrei a linha de cabeçalho — esperava 'SKU Kit' na coluna B "
        "nas primeiras linhas da aba."
    )


def _component_columns(ws, header_row: int) -> list[int]:
    """
    Colunas iniciais de cada trinca de componente, detectadas pela largura da aba.
    Considera a trinca válida quando o cabeçalho da 1ª coluna do grupo é 'SKU'.
    """
    grupos = []
    col = FIRST_COMPONENT_COL
    while col + 2 <= ws.max_column:
        if _norm(ws.cell(row=header_row, column=col).value) == "sku":
            grupos.append(col)
        col += COLS_PER_COMPONENT
    return grupos


def parse_kit_workbook(file_path_or_stream) -> dict:
    """
    Lê a planilha e devolve a estrutura já normalizada, SEM tocar no banco.

    {
      "sheet", "header_row", "component_columns",
      "kits":    [ {row, seller_name, kit_sku, kit_name, components:[{sku,name,quantity}]} ],
      "blocked": [ {row, seller_name, kit_sku, reason} ],
      "warnings":[ str ],
    }
    """
    wb = load_workbook(file_path_or_stream, data_only=True, read_only=True)
    try:
        ws = _find_sheet(wb)
        header_row = _find_header_row(ws)
        grupos = _component_columns(ws, header_row)
        if not grupos:
            raise ValueError(
                "Não encontrei nenhuma coluna de componente — esperava trincas "
                "SKU | NOME | QUANTIDADE a partir da coluna D."
            )

        kits, blocked, warnings = [], [], []
        vistos: dict[tuple, int] = {}   # (cliente, kit_sku) → linha, p/ detectar repetido

        for r in range(header_row + 1, ws.max_row + 1):
            cliente = str(ws.cell(row=r, column=COL_CLIENTE).value or "").strip()
            kit_sku = _clean_sku(ws.cell(row=r, column=COL_KIT_SKU).value)
            kit_nome = str(ws.cell(row=r, column=COL_KIT_NOME).value or "").strip()

            componentes, erros = [], []
            for gi, col in enumerate(grupos, start=1):
                sku = _clean_sku(ws.cell(row=r, column=col).value)
                nome = str(ws.cell(row=r, column=col + 1).value or "").strip()
                qtd_raw = ws.cell(row=r, column=col + 2).value

                if not sku and not nome and qtd_raw is None:
                    continue          # grupo vazio: componente não usado
                if not sku:
                    erros.append(f"PRODUTO {gi}: componente com nome mas sem SKU")
                    continue
                qtd, err = _clean_qty(qtd_raw)
                if err:
                    erros.append(f"PRODUTO {gi} ({sku}): {err}")
                    continue
                componentes.append({"sku": sku, "name": nome or sku, "quantity": qtd})

            # Linha totalmente vazia: ignora em silêncio
            if not cliente and not kit_sku and not kit_nome and not componentes and not erros:
                continue

            if not cliente:
                erros.append("linha sem CLIENTE")
            if not kit_sku:
                erros.append("linha sem SKU do kit")
            if not componentes and not erros:
                erros.append("kit sem nenhum componente")

            chave = (_norm(cliente), kit_sku)
            if chave in vistos:
                erros.append(
                    f"kit repetido na planilha (já aparece na linha {vistos[chave]}) — "
                    "deixe apenas uma linha por kit"
                )
            elif kit_sku:
                vistos[chave] = r

            if erros:
                blocked.append({
                    "row": r, "seller_name": cliente, "kit_sku": kit_sku,
                    "kit_name": kit_nome, "reason": "; ".join(erros),
                })
                continue

            # Componente repetido dentro do mesmo kit: soma e avisa
            somados: dict[str, dict] = {}
            for c in componentes:
                if c["sku"] in somados:
                    somados[c["sku"]]["quantity"] += c["quantity"]
                    warnings.append(
                        f"Linha {r} ({kit_sku}): componente {c['sku']} aparece mais de uma "
                        f"vez — quantidades somadas"
                    )
                else:
                    somados[c["sku"]] = dict(c)

            kits.append({
                "row": r,
                "seller_name": cliente,
                "kit_sku": kit_sku,
                "kit_name": kit_nome or kit_sku,
                "components": list(somados.values()),
            })

        return {
            "sheet": ws.title,
            "header_row": header_row,
            "component_columns": len(grupos),
            "kits": kits,
            "blocked": blocked,
            "warnings": warnings,
        }
    finally:
        wb.close()


def match_sellers(kits: list[dict], alias_map: dict) -> tuple[dict, list[str]]:
    """
    Casa o nome do cliente da planilha com os sellers cadastrados.

    Devolve (mapa_nome_normalizado → Seller, nomes_não_reconhecidos).
    Nunca cria seller: a decisão é sempre do usuário, na tela.
    """
    encontrados, faltando = {}, []
    vistos: set[str] = set()          # dedupe pela chave normalizada, não pelo nome cru
    for k in kits:
        chave = _norm(k["seller_name"])
        if chave in vistos:
            continue
        vistos.add(chave)
        seller = alias_map.get(k["seller_name"].strip().lower())
        if seller:
            encontrados[chave] = seller
        else:
            faltando.append(k["seller_name"].strip())
    return encontrados, faltando
