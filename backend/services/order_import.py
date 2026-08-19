"""
WMS Kiwkiw - Serviço de Importação de Pedidos
Processa o arquivo Excel gerado pelo robô de captura dos ERPs.
Reproduz a lógica das macros 'colar_info()' e validações do MENU.
"""

import openpyxl
import pandas as pd
from datetime import date, datetime
from typing import Optional, List
import os
import re

from sqlalchemy.orm import Session, joinedload
from .. import models, schemas
from .kit_handler import process_order_items
from .stock_manager import apply_stock_for_orders
from .excel_utils import ensure_xlsx_path
from ..timezone_utils import now_brasilia, today_brasilia


# Mapeamento de naturezas de operação para tipo (Entrada/Saída)
NATURE_TYPE_MAP = {
    "Venda de mercadorias": "Saída",
    "Venda de Mercadoria para Cliente Final": "Saída",
    "Venda de mercadoria - Revenda": "Saída",
    "Venda de mercadoria a nao contribuinte": "Saída",
    "Operacao de demonstracao - Brinde": "Saída",
    "Remessa e retorno em garantia troca de mercadorias - Saida d": "Saída",
    "Remessa de mercadoria em consignacao - Revenda": "Saída",
    "Remessa em bonificacao": "Saída",
    "Simples Remessa": "Saída",
    "Remessa em demonstracao": "Saída",
    "Remessa em consignacao": "Saída",
    "Remessa para armazem-geral": "Entrada",  # Retorno = Entrada
    "Remessa para deposito fechado ou armazem do RS p/ fora do RS": "Entrada",
    "Remessa em Bonificacao, Doacao ou Brinde": "Saída",
    "Venda de Producao do Estabelecimento": "Saída",
    "Devolucao simbolica pelo consignatario": "Saída",
    "Remessa de mercadoria em consignacao": "Saída",
}


def parse_nf_safe(value) -> Optional[str]:
    """
    Converte valor da coluna 'Nota'/'NF' para string canônica.
    Aceita int, float (Excel grava números como float), string.
    Remove '.0' que o Excel costuma adicionar.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # Excel grava 2433 como 2433.0 — mantém apenas parte inteira
        if value.is_integer():
            return str(int(value))
        return str(value)
    s = str(value).strip()
    # Tira '.0' final se for algo como "2433.0"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


def parse_danfe_key_safe(value) -> Optional[str]:
    """
    Converte valor da coluna 'Chave DANFE' para string de até 44 dígitos.
    Trata:
      - string já correta ('53260443823086...')
      - int grande (raro — Python aceita, Excel às vezes)
      - float em notação científica (Bling/Excel fazem isso com chaves de 44 díg.
        — nesse caso há PERDA DE PRECISÃO nos últimos ~15 dígitos. Fazemos o
        melhor esforço convertendo para string decimal inteira sem 'e+XX').
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # 3.5260455718181e+43 → "35260455718180999976320753831245549000949760"
        # A precisão é perdida, mas pelo menos não vira notação científica.
        try:
            return f"{int(value):d}"
        except (OverflowError, ValueError):
            return None
    s = str(value).strip()
    # Se veio como string em notação científica ("3.52e+43"), converte
    if "e" in s.lower():
        try:
            return f"{int(float(s)):d}"
        except (OverflowError, ValueError):
            pass
    # Remove '.0' se Excel enviou
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


def parse_date_safe(value) -> Optional[date]:
    """Converte diversos formatos de data para objeto date."""
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        # Remove timestamp se presente
        value = value.split()[0].strip()
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    # Excel serial date
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta
            excel_origin = date(1899, 12, 30)
            return excel_origin + timedelta(days=int(value))
        except Exception:
            pass
    return None


def _build_seller_alias_map(db: Session, active: bool = True) -> dict:
    """
    Carrega os sellers (ativos ou inativos, conforme `active`) uma única vez e
    constrói um dict de lookup: alias_normalizado (lower+strip) → objeto models.Seller

    Inclui três fontes de chave por seller:
      1. trade_name  — apelido oficial (ex: "YUGEN")
      2. name        — razão social (ex: "YUGEN COSMETICOS LTDA")
      3. other_aliases — strings extras separadas por ";" cadastradas pelo usuário
                        (ex: "MERU SERVICOS EMPRESARIAIS E VAREJO LTDA;MERU SERVICOS")
    """
    sellers = db.query(models.Seller).filter(models.Seller.active == active).all()
    alias_map: dict[str, models.Seller] = {}
    for s in sellers:
        for key in (s.trade_name, s.name):
            if key:
                alias_map[key.strip().lower()] = s
        # Expande outros apelidos cadastrados
        if getattr(s, "other_aliases", None):
            for alias in s.other_aliases.split(";"):
                alias = alias.strip()
                if alias:
                    alias_map[alias.lower()] = s
    return alias_map


def _resolve_seller_name(raw_name: str, alias_map: dict) -> str:
    """
    Dado um nome vindo do Excel (pode ser razão social longa, apelido, etc.),
    tenta encontrar o seller correspondente no mapa de aliases e retorna o
    trade_name canônico. Se não encontrar, retorna o nome original sem alteração.

    A resolução é case-insensitive e ignora espaços extras nas extremidades.
    """
    if not raw_name:
        return raw_name
    key = raw_name.strip().lower()
    seller = alias_map.get(key)
    if seller:
        return seller.trade_name   # sempre retorna o apelido oficial
    return raw_name.strip()


def _is_valid_seller_link_decision(decision) -> bool:
    """
    Valida o formato de uma decisão de seller_link_decisions:
      {"action": "create", "unit_id": int} ou {"action": "link", "seller_id": int}
    """
    if not isinstance(decision, dict):
        return False
    action = decision.get("action")
    if action == "create":
        return isinstance(decision.get("unit_id"), int)
    if action == "link":
        return isinstance(decision.get("seller_id"), int)
    return False


def get_or_create_seller(
    db: Session,
    seller_name: str,
    alias_map: dict | None = None,
) -> models.Seller:
    """
    Busca seller pelo apelido, razão social, outros aliases ou cria um novo.

    Se alias_map for fornecido (recomendado no import em lote), o lookup é O(1).
    Sem o mapa, faz queries individuais no banco (compatível com chamadas avulsas).
    """
    # 1. Tenta pelo mapa em memória (sem query extra)
    if alias_map is not None:
        key = seller_name.strip().lower()
        seller = alias_map.get(key)
        if seller:
            return seller

    # 2. Fallback: queries individuais (trade_name, name, other_aliases)
    seller = db.query(models.Seller).filter(
        models.Seller.trade_name == seller_name
    ).first()

    if not seller:
        seller = db.query(models.Seller).filter(
            models.Seller.name == seller_name
        ).first()

    if not seller:
        # Busca em other_aliases — matching parcial por substring do campo
        # (SQLite: LIKE '%alias%'; funciona tb em PostgreSQL)
        from sqlalchemy import or_
        key_lower = seller_name.strip().lower()
        sellers_with_aliases = db.query(models.Seller).filter(
            models.Seller.other_aliases.isnot(None)
        ).all()
        for s in sellers_with_aliases:
            if s.other_aliases:
                for alias in s.other_aliases.split(";"):
                    if alias.strip().lower() == key_lower:
                        seller = s
                        break
            if seller:
                break

    if not seller:
        # Cria seller automaticamente se não existir
        seller = models.Seller(
            name=seller_name,
            trade_name=seller_name,
        )
        db.add(seller)
        db.flush()

    return seller


def validate_session_checks(session: models.PickingSession, db: Session) -> dict:
    """
    Realiza as 4 checagens do MENU (P6/P8/P10/P12).
    Retorna dict com status de cada checagem.
    """
    orders = db.query(models.Order).filter(
        models.Order.session_id == session.id
    ).all()

    if not orders:
        return {
            "check_transport": False,
            "check_separation": False,
            "check_planning": False,
            "check_stock": False,
            "issues": ["Nenhum pedido na sessão"],
        }

    issues = []
    warnings = []

    # Checagem 1: Validar expedição e meio de transporte
    orders_without_carrier = [o for o in orders if not o.carrier]
    check_transport = len(orders_without_carrier) == 0
    if not check_transport:
        issues.append(f"{len(orders_without_carrier)} pedidos sem transportadora definida")

    # Checagem 2: Relatório de separação gerado
    check_separation = bool(session.separation_pdf)

    # Checagem 3: Relatório de planejamento gerado
    check_planning = bool(session.expedition_pdf)

    # Checagem 4: Verificar duplicatas
    nf_keys = [o.danfe_key for o in orders if o.danfe_key]
    duplicates = len(nf_keys) - len(set(nf_keys))
    check_stock = duplicates == 0
    if not check_stock:
        issues.append(f"{duplicates} notas fiscais duplicadas encontradas")

    return {
        "check_transport": check_transport,
        "check_separation": check_separation,
        "check_planning": check_planning,
        "check_stock": check_stock,
        "issues": issues,
        "warnings": warnings,
    }


def _normalize_header(h) -> str:
    """Normaliza cabeçalho para comparação: remove acentos, espaços e caixa."""
    if h is None:
        return ""
    s = str(h).strip().lower()
    # Remove acentos simples
    for a, b in (("á", "a"), ("ã", "a"), ("â", "a"), ("é", "e"), ("ê", "e"),
                 ("í", "i"), ("ó", "o"), ("ô", "o"), ("õ", "o"), ("ú", "u"), ("ç", "c")):
        s = s.replace(a, b)
    return s


def _build_column_map(header_row) -> dict:
    """
    Constrói mapa {campo_lógico: índice_da_coluna} a partir do cabeçalho real do arquivo.
    Aceita os dois layouts conhecidos:
      - Importador original (16 colunas, sem índice)
      - Exportação Tiny "Notas Pendentes" (15 colunas, com índice na col 0)
    """
    # ATENÇÃO: a ordem dentro de cada lista importa — aliases mais específicos
    # devem vir ANTES dos curtos (startswith), pra evitar match errado.
    # Ex.: 'nota' casa com header 'nota'/'Nota' (ERP Bling/contingência);
    # 'hora nota' precisa ser mapeado para 'hour', não para 'nf_number'.
    aliases = {
        "erp_code":        ["cod erp", "cod tiny", "codigo erp", "codigo tiny", "cod"],
        # 'nota' é o cabeçalho real nos arquivos Bling/contingência
        "nf_number":       ["# nf", "numero nf", "num nf", "n nf", "nota", "nf"],
        "customer":        ["cliente", "destinatario", "nome cliente"],
        "order_date":      ["data nota", "dt nota", "emissao", "data"],
        "sku":             ["cod sku", "codigo sku", "codigo produto", "sku"],
        "product_name":    ["nome do produto", "nome produto", "produto", "descricao"],
        "quantity":        ["quantidade", "qtd", "qtde", "qnt"],
        # 'transportador' (singular, arquivo contingência) e 'transportadora'
        "carrier":         ["transportadora", "transportador", "transp"],
        "status":          ["status", "situacao"],
        # 'hora nota' precisa vir ANTES de 'hora' (mesmo prefixo com nf_number)
        "hour":            ["hora nota", "hora"],
        "danfe_key":       ["chave danfe", "chave nfe", "chave de acesso", "chave"],
        "seller_name":     ["seller", "cliente seller", "loja"],
        "expedition_date": ["data expedicao", "dt expedicao", "expedicao"],
        "nature":          ["natureza da operacao", "natureza operacao", "natureza"],
        "nf_key":          ["chave (nf+seller)", "chave nf+seller", "chave pedido"],
        "unit_name":       ["unidade", "uf unidade", "filial"],
    }
    cmap = {}
    # Passo 1: match EXATO (mais confiável — evita colisões tipo
    # 'hora nota' casando com alias 'nota' de nf_number).
    for idx, cell in enumerate(header_row):
        norm = _normalize_header(cell)
        if not norm:
            continue
        for field, opts in aliases.items():
            if field in cmap:
                continue
            if any(norm == o for o in opts):
                cmap[field] = idx
                break

    # Passo 2: match por prefixo para os campos ainda não encontrados.
    for idx, cell in enumerate(header_row):
        norm = _normalize_header(cell)
        if not norm:
            continue
        # Evita sobrescrever colunas já mapeadas a outros campos
        if idx in cmap.values():
            continue
        for field, opts in aliases.items():
            if field in cmap:
                continue
            if any(norm.startswith(o) for o in opts):
                cmap[field] = idx
                break
    return cmap


class WatcherImportError(Exception):
    """
    Exceção estruturada para o folder_watcher.
    Separa o título curto (reason) dos detalhes adicionais (details)
    para que o painel do robô exiba informações úteis sem jargão técnico.
    """
    def __init__(self, reason: str, details: str = ""):
        super().__init__(reason)
        self.reason  = reason    # título de 1 linha
        self.details = details   # contexto específico (seller, NFs, colunas…)


def import_orders_from_excel_sync(
    file_path: str,
    db: Session,
    unit_id: Optional[int] = None,
) -> int:
    """
    Wrapper de import_excel_orders para uso no folder_watcher.

    Mantido como função separada por causa do tratamento de erro rico
    (WatcherImportError com reason + details) que o painel do robô consome.
    O import em si já é síncrono — não há event loop envolvido.

    Retorna a quantidade de pedidos importados ou lança WatcherImportError
    com mensagens ricas (reason + details) para exibição no painel do robô.
    """
    result = import_excel_orders(file_path=file_path, unit_id=unit_id, db=db)

    # ── Falha definitiva (não é confirmação de duplicata) ──────────────────────
    if not result.success and not result.requires_confirmation:
        msg = result.message or "Erro ao importar arquivo"
        msg_l = msg.lower()

        # Colunas faltando — extrai quais são
        if "faltam colunas" in msg_l or "cabeçalho não reconhecido" in msg_l:
            # Mensagem do order_import: "Faltam colunas essenciais: nf_number, sku. Cabeçalho lido: [...]"
            # Extrai apenas a parte útil
            details = msg
            if ". Cabeçalho lido:" in msg:
                parts = msg.split(". Cabeçalho lido:")
                details = parts[0]          # só "Faltam colunas essenciais: nf_number, sku"
                header_sample = parts[1].strip()[:120]
                details += f"\nCabeçalho encontrado: {header_sample}"
            raise WatcherImportError(
                reason="Colunas obrigatórias não encontradas no arquivo",
                details=details,
            )

        # Linha(s) com SKU vazio — checar antes do "vazio" genérico abaixo,
        # senão cai no ramo errado ("Arquivo vazio ou sem dados").
        if result.missing_sku_lines:
            lines = result.missing_sku_lines
            detail_lines = [
                f"NF {l.nf_number} ({l.seller_name}): {l.product_name or 'produto sem nome'}"
                for l in lines[:8]
            ]
            if len(lines) > 8:
                detail_lines.append(f"...e mais {len(lines) - 8} linha(s)")
            raise WatcherImportError(
                reason=f"{len(lines)} linha(s) com SKU vazio — arquivo não importado",
                details="\n".join(detail_lines),
            )

        if "vazio" in msg_l or "sem cabeçalho" in msg_l:
            raise WatcherImportError(
                reason="Arquivo vazio ou sem dados",
                details="O arquivo não contém linhas ou o cabeçalho está ausente.",
            )

        # Erros linha-a-linha acumulados
        if result.errors:
            # Agrupa erros por tipo para não exibir 100 linhas
            seller_errors = [e for e in result.errors if "seller" in e.lower()]
            sku_errors    = [e for e in result.errors if "sku" in e.lower() or "produto" in e.lower()]
            other_errors  = [e for e in result.errors if e not in seller_errors and e not in sku_errors]

            detail_lines = []
            if seller_errors:
                # Extrai nomes de sellers mencionados
                sellers = list({e.split("'")[1] for e in seller_errors if "'" in e}
                                | {e.split('"')[1] for e in seller_errors if '"' in e})
                if sellers:
                    detail_lines.append(f"Sellers com problema: {', '.join(sellers[:5])}")
                else:
                    detail_lines.append(f"{len(seller_errors)} erro(s) relacionados a sellers")
            if sku_errors:
                detail_lines.append(f"{len(sku_errors)} erro(s) de SKU/produto")
            if other_errors:
                detail_lines.append(f"{len(other_errors)} outro(s) erro(s)")
            detail_lines.append(f"Exemplo: {result.errors[0][:120]}")

            raise WatcherImportError(
                reason=f"Importação falhou com {len(result.errors)} erro(s)",
                details="\n".join(detail_lines),
            )

        # Mensagem genérica do import
        raise WatcherImportError(
            reason=msg[:120] if len(msg) > 120 else msg,
            details="",
        )

    # ── Confirmação necessária (sellers inativos) ───────────────────────────────
    # O watcher roda sozinho, sem alguém pra confirmar — trata como falha e
    # deixa o arquivo para reimportação manual via tela, onde há o modal de decisão.
    if result.requires_confirmation and result.inactive_sellers:
        inactives = result.inactive_sellers
        detail_lines = [f"{i.seller_name}: {len(i.nf_numbers)} NF(s)" for i in inactives[:5]]
        raise WatcherImportError(
            reason=f"{len(inactives)} seller(s) inativo(s) no arquivo — requer confirmação manual",
            details="\n".join(detail_lines),
        )

    # ── Confirmação necessária (NFs duplicadas) ────────────────────────────────
    if result.requires_confirmation:
        dups = result.duplicates or []
        # Agrupa NFs duplicadas por seller para facilitar a identificação
        by_seller: dict[str, list[str]] = {}
        for d in dups:
            by_seller.setdefault(d.seller_name, []).append(d.nf_number)

        detail_lines = []
        for seller, nfs in list(by_seller.items())[:5]:   # mostra até 5 sellers
            nf_list = ", ".join(nfs[:8])                   # mostra até 8 NFs por seller
            if len(nfs) > 8:
                nf_list += f" … (+{len(nfs)-8})"
            detail_lines.append(f"{seller}: NF(s) {nf_list}")

        raise WatcherImportError(
            reason=f"{len(dups)} NF(s) duplicadas — reimportação requer confirmação manual",
            details="\n".join(detail_lines) if detail_lines else "",
        )

    # Sucesso — retorna contagem
    return result.orders_imported


def import_excel_orders(
    file_path: str,
    unit_id: Optional[int] = None,
    db: Session = None,
    created_by_id: Optional[int] = None,
    file_type: Optional["models.FileType"] = None,
    for_billing: bool = True,
    force_duplicates: bool = False,
    inactive_seller_decisions: Optional[dict] = None,
    seller_link_decisions: Optional[dict] = None,
) -> schemas.ImportResult:
    """
    Importa arquivo Excel de pedidos (gerado pelo robô de captura dos ERPs).

    Faz auto-detecção do cabeçalho para tolerar variações de layout:
      - Importador original (16 colunas)
      - Exportação Tiny/Bling "Notas Pendentes" (15 colunas, com coluna índice)
    Campos obrigatórios detectados no cabeçalho: nf_number, sku, seller_name.

    Parâmetros:
      - force_duplicates: se False (default), aborta quando detecta NFs que já
        existem no banco e retorna a lista de duplicatas para confirmação
        pelo usuário. Se True, reimporta os duplicados como novos pedidos.
      - inactive_seller_decisions: dict {seller_id: "reactivate" | "ignore"}.
        Se o arquivo referenciar sellers inativos e este dict não cobrir todos
        eles, a importação aborta e retorna `requires_confirmation=True` com a
        lista em `inactive_sellers` para o usuário decidir. "reactivate"
        reativa o seller (active=True) antes de importar seus pedidos;
        "ignore" pula apenas os pedidos daquele seller, sem tocar nos demais.
      - seller_link_decisions: dict {seller_name: {"action": "create", "unit_id": int}
        | {"action": "link", "seller_id": int}}. Se o arquivo referenciar um nome de
        seller que não bate com nenhum cadastro (ativo ou inativo) e este dict não
        cobrir todos eles, a importação aborta e retorna `requires_confirmation=True`
        com a lista em `unmatched_sellers` para o usuário decidir. Nunca cria um
        seller novo silenciosamente — "create" exige unit_id explícito; "link"
        associa os pedidos a um seller já cadastrado.
    """
    errors = []
    warnings = []
    duplicates_info: List[schemas.DuplicateOrderInfo] = []
    inactive_seller_decisions = inactive_seller_decisions or {}
    seller_link_decisions = seller_link_decisions or {}
    orders_imported = 0
    kits_expanded = 0

    try:
        # Lê o arquivo Excel — .xls (formato antigo) é convertido pra .xlsx antes,
        # porque openpyxl não sabe ler .xls
        file_path = ensure_xlsx_path(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Tenta encontrar a aba correta
        sheet = None
        for sheet_name in ["1. excel", "excel", "pedidos", "Sheet1"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                break

        if sheet is None:
            sheet = wb.active

        # Lê cabeçalho e constrói mapa de colunas
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Arquivo vazio ou sem cabeçalho")

        cmap = _build_column_map(header_row)

        # ── Constrói mapas de aliases uma única vez ─────────────────────────────
        # Permite resolução O(1) de "MERU SERVICOS EMPRESARIAIS" → "YUGEN" para
        # todas as linhas do arquivo, sem queries adicionais por row.
        # Mapa de inativos é separado para nunca ser usado silenciosamente na
        # criação/atualização de pedidos sem confirmação explícita do usuário.
        seller_alias_map = _build_seller_alias_map(db, active=True)
        inactive_seller_alias_map = _build_seller_alias_map(db, active=False)
        # Usado só para resolver o nome canônico exibido (não decide ativo/inativo)
        seller_name_resolve_map = {**inactive_seller_alias_map, **seller_alias_map}

        # Valida campos essenciais
        required = ["nf_number", "sku", "seller_name"]
        missing = [f for f in required if f not in cmap]
        if missing:
            raise ValueError(
                "Cabeçalho não reconhecido. Faltam colunas essenciais: "
                + ", ".join(missing)
                + f". Cabeçalho lido: {list(header_row)}"
            )

        def get(row, field):
            """Extrai valor da coluna pelo nome lógico (ou None se coluna ausente)."""
            idx = cmap.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        # Lê os dados em um dict por NF
        orders_dict = {}  # nf_key → {order_data, items: []}
        total_rows = 0
        missing_sku_rows = []  # linhas com NF/seller identificados mas sem SKU

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Pula linhas vazias
            if not any(row):
                continue

            # Extrai campos via mapa (tolerante à ordem)
            erp_code     = str(get(row, "erp_code")).strip() if get(row, "erp_code") is not None else None
            # parse_nf_safe tolera int, float (Excel), string
            nf_number    = parse_nf_safe(get(row, "nf_number"))
            customer     = str(get(row, "customer")).strip() if get(row, "customer") else None
            order_date   = parse_date_safe(get(row, "order_date"))
            sku_raw      = get(row, "sku")
            sku          = str(sku_raw).strip() if sku_raw is not None else None
            # Remove '.0' se o SKU veio como float do Excel (ex.: 123.0)
            if sku and sku.endswith(".0") and sku[:-2].isdigit():
                sku = sku[:-2]
            product_name = str(get(row, "product_name")).strip() if get(row, "product_name") else None

            # Quantidade: tolera strings ("1"), floats (1.0), vazio → 1
            qty_raw = get(row, "quantity")
            try:
                quantity = int(float(qty_raw)) if qty_raw not in (None, "") else 1
            except (ValueError, TypeError):
                warnings.append(f"Quantidade inválida ({qty_raw!r}) na NF {nf_number} — assumindo 1")
                quantity = 1

            carrier         = str(get(row, "carrier")).strip() if get(row, "carrier") else None
            # parse_danfe_key_safe: trata floats em notação científica (Bling)
            danfe_key       = parse_danfe_key_safe(get(row, "danfe_key"))
            seller_name_raw = str(get(row, "seller_name")).strip() if get(row, "seller_name") else None
            # ── Resolução de aliases: substitui razão social / apelido alternativo
            # pelo trade_name canônico do seller (ex: "MERU SERVICOS..." → "YUGEN")
            seller_name     = _resolve_seller_name(seller_name_raw, seller_name_resolve_map) if seller_name_raw else None
            expedition_date = parse_date_safe(get(row, "expedition_date"))
            nature          = str(get(row, "nature")).strip() if get(row, "nature") else None
            nf_key_raw      = get(row, "nf_key")
            nf_key_is_auto  = not bool(nf_key_raw)
            nf_key          = str(nf_key_raw).strip() if nf_key_raw else (
                f"{nf_number}_{seller_name}" if nf_number and seller_name else None
            )

            # Validações básicas
            if not nf_number or not seller_name:
                warnings.append(f"Linha ignorada: dados incompletos (NF={nf_number})")
                continue

            # SKU vazio: diferente do caso acima, aqui a NF e o seller são
            # conhecidos — só falta o código do produto. Sem SKU não existe
            # código de barras pra bipar, então a linha nunca pode virar item
            # de pedido. Não é só um warning: registra pra bloquear o arquivo
            # inteiro logo abaixo (decisão do dono do sistema, 19/08/2026).
            if not sku:
                missing_sku_rows.append({
                    "nf_number": nf_number,
                    "seller_name": seller_name,
                    "customer_name": customer,
                    "product_name": product_name,
                })
                continue

            total_rows += 1

            # Chave única do pedido: NF + seller
            order_key = f"{nf_number}_{seller_name}"

            if order_key not in orders_dict:
                orders_dict[order_key] = {
                    "erp_code": erp_code,
                    "nf_number": nf_number,
                    "customer_name": customer or "N/A",
                    "order_date": order_date or today_brasilia(),
                    "seller_name": seller_name,
                    "carrier": carrier,
                    "expedition_date": expedition_date,
                    "nature": nature,
                    "danfe_key": danfe_key,
                    "nf_key": nf_key,
                    "nf_key_auto": nf_key_is_auto,
                    "items": [],
                }

            # ── Deduplicação de SKU dentro da mesma NF ──────────────
            # Se o mesmo SKU aparece em mais de uma linha da NF (ex.: ERP
            # que gera uma linha por unidade), somamos as quantidades em vez
            # de criar dois itens separados — evita bipagem dupla no scanner.
            existing_item = next(
                (it for it in orders_dict[order_key]["items"] if it["sku"] == sku),
                None,
            )
            if existing_item:
                existing_item["quantity"] += quantity
            else:
                orders_dict[order_key]["items"].append({
                    "sku": sku,
                    "product_name": product_name or sku,
                    "quantity": quantity,
                })

        # ── Linha com SKU vazio: bloqueia o arquivo inteiro ──────────────
        # Sem SKU não há como casar barcode nenhum na bipagem, então a linha
        # não pode simplesmente ser descartada em silêncio (era o que
        # acontecia antes: a NF entrava faltando esse item e ninguém
        # percebia). Aborta antes de qualquer resolução de seller ou
        # gravação — nada deste arquivo é importado até a planilha de
        # origem ser corrigida e reenviada.
        if missing_sku_rows:
            return schemas.ImportResult(
                success=False,
                message=f"Não subimos o arquivo porque tem {len(missing_sku_rows)} linha(s) com SKU vazio",
                total_rows=total_rows,
                orders_imported=0,
                orders_with_kits=0,
                errors=[],
                warnings=warnings,
                missing_sku_lines=[
                    schemas.MissingSkuLineInfo(
                        nf_number=r["nf_number"],
                        seller_name=r["seller_name"],
                        customer_name=r["customer_name"],
                        product_name=r["product_name"],
                    )
                    for r in missing_sku_rows
                ],
            )

        # ── RESOLUÇÃO ÚNICA DE SELLER POR PEDIDO ────────────────────
        # Resolve o seller de cada order_key uma única vez (reaproveitado nas
        # checagens abaixo e na persistência), já distinguindo sellers ativos
        # de inativos — o mapa de inativos nunca é usado silenciosamente.
        order_key_seller_map: dict[str, models.Seller] = {}
        for order_key, order_data in orders_dict.items():
            resolved_key = order_data["seller_name"].strip().lower()
            seller_obj = seller_alias_map.get(resolved_key) or inactive_seller_alias_map.get(resolved_key)
            if not seller_obj:
                # Fallback para query direta (seller novo, ainda não em nenhum mapa)
                seller_obj = db.query(models.Seller).filter(
                    (models.Seller.trade_name == order_data["seller_name"]) |
                    (models.Seller.name == order_data["seller_name"])
                ).first()
            if seller_obj:
                order_key_seller_map[order_key] = seller_obj

        # ── PRÉ-CHECAGEM DE SELLERS INATIVOS ────────────────────────
        # Um seller inativo referenciado no arquivo nunca deve receber pedido
        # novo silenciosamente. Exige decisão explícita: reativar ou ignorar
        # apenas os pedidos daquele seller neste arquivo.
        inactive_sellers_found: dict[int, schemas.InactiveSellerInfo] = {}
        for order_key, order_data in orders_dict.items():
            seller_obj = order_key_seller_map.get(order_key)
            if seller_obj and not seller_obj.active:
                info = inactive_sellers_found.setdefault(
                    seller_obj.id,
                    schemas.InactiveSellerInfo(seller_id=seller_obj.id, seller_name=seller_obj.trade_name, nf_numbers=[]),
                )
                info.nf_numbers.append(order_data["nf_number"])

        missing_decisions = [
            sid for sid in inactive_sellers_found
            if inactive_seller_decisions.get(sid) not in ("reactivate", "ignore")
        ]
        if inactive_sellers_found and missing_decisions:
            # Aborta sem criar nada — frontend vai perguntar ao usuário
            return schemas.ImportResult(
                success=False,
                message=(
                    f"{len(inactive_sellers_found)} seller(s) deste arquivo estão inativos. "
                    f"Confirme se deseja reativá-los ou ignorar os pedidos deles neste arquivo."
                ),
                total_rows=total_rows,
                orders_imported=0,
                orders_with_kits=0,
                errors=[],
                warnings=warnings,
                requires_confirmation=True,
                inactive_sellers=list(inactive_sellers_found.values()),
            )

        # Aplica as decisões: reativa (com auditoria) ou marca pedidos p/ ignorar
        ignored_order_keys: set[str] = set()
        for seller_id, decision in inactive_seller_decisions.items():
            if seller_id not in inactive_sellers_found:
                continue
            seller_obj = next(
                (s for s in order_key_seller_map.values() if s.id == seller_id), None
            )
            if not seller_obj:
                continue
            if decision == "reactivate":
                seller_obj.active = True
                db.add(models.AuditLog(
                    entity_type="Seller",
                    entity_id=seller_obj.id,
                    action="REACTIVATE",
                    detail=f"Seller '{seller_obj.trade_name}' reativado automaticamente durante import de pedidos (arquivo {os.path.basename(file_path)})",
                    user_id=created_by_id,
                ))
            elif decision == "ignore":
                for order_key, order_data in orders_dict.items():
                    if order_key_seller_map.get(order_key) and order_key_seller_map[order_key].id == seller_id:
                        ignored_order_keys.add(order_key)
                        warnings.append(
                            f"NF {order_data['nf_number']} do seller '{order_data['seller_name']}' ignorada — seller inativo (confirmado pelo usuário)"
                        )

        # ── PRÉ-CHECAGEM DE SELLERS NÃO RECONHECIDOS ────────────────
        # Nome de seller no arquivo que não bateu com nenhum cadastro (ativo
        # nem inativo). Nunca cria um seller novo silenciosamente — exige
        # decisão explícita do usuário: vincular a um seller já cadastrado ou
        # criar um novo (com unidade obrigatória).
        unmatched_sellers_found: dict[str, schemas.UnmatchedSellerInfo] = {}
        for order_key, order_data in orders_dict.items():
            if order_key in ignored_order_keys or order_key_seller_map.get(order_key):
                continue
            name = order_data["seller_name"]
            norm_key = name.strip().lower()
            info = unmatched_sellers_found.setdefault(
                norm_key,
                schemas.UnmatchedSellerInfo(seller_name=name, nf_numbers=[]),
            )
            info.nf_numbers.append(order_data["nf_number"])

        missing_seller_link_decisions = [
            norm_key for norm_key, info in unmatched_sellers_found.items()
            if not _is_valid_seller_link_decision(seller_link_decisions.get(info.seller_name))
        ]
        if unmatched_sellers_found and missing_seller_link_decisions:
            # Aborta sem criar nada — frontend vai perguntar ao usuário
            return schemas.ImportResult(
                success=False,
                message=(
                    f"{len(unmatched_sellers_found)} nome(s) de seller neste arquivo não batem com "
                    f"nenhum cadastro. Confirme se deve vincular a um seller existente ou criar um novo."
                ),
                total_rows=total_rows,
                orders_imported=0,
                orders_with_kits=0,
                errors=[],
                warnings=warnings,
                requires_confirmation=True,
                unmatched_sellers=list(unmatched_sellers_found.values()),
            )

        # Aplica as decisões: cria (unidade obrigatória) ou vincula a seller existente
        for norm_key, info in unmatched_sellers_found.items():
            decision = seller_link_decisions.get(info.seller_name)
            if decision.get("action") == "create":
                seller_obj = models.Seller(
                    name=info.seller_name,
                    trade_name=info.seller_name,
                    unit_id=decision["unit_id"],
                )
                db.add(seller_obj)
                db.flush()
                db.add(models.AuditLog(
                    entity_type="Seller",
                    entity_id=seller_obj.id,
                    action="CREATE",
                    detail=(
                        f"Seller '{info.seller_name}' criado durante import de pedidos "
                        f"(arquivo {os.path.basename(file_path)}), unidade confirmada pelo usuário"
                    ),
                    user_id=created_by_id,
                ))
            else:  # "link"
                seller_obj = db.query(models.Seller).filter(
                    models.Seller.id == decision["seller_id"]
                ).first()
                if not seller_obj:
                    errors.append(f"Seller id {decision['seller_id']} não encontrado para vincular '{info.seller_name}'")
                    continue
            for order_key, order_data in orders_dict.items():
                if order_key not in order_key_seller_map and order_data["seller_name"].strip().lower() == norm_key:
                    order_key_seller_map[order_key] = seller_obj
                    # nf_key foi montado com o nome bruto do arquivo (fallback, sem
                    # coluna própria) — agora que sabemos o seller de verdade, troca
                    # pelo trade_name curto para não estourar o VARCHAR(50)
                    if order_data.get("nf_key_auto"):
                        order_data["nf_key"] = f"{order_data['nf_number']}_{seller_obj.trade_name}"

        # ── PRÉ-CHECAGEM DE DUPLICATAS ──────────────────────────────
        # Antes de criar qualquer coisa, verifica se alguma NF deste arquivo
        # já existe no banco. Se sim e o usuário não confirmou (force_duplicates=False),
        # abortamos e devolvemos a lista pro frontend exibir e perguntar.
        for order_key, order_data in orders_dict.items():
            if order_key in ignored_order_keys:
                continue
            seller_obj = order_key_seller_map.get(order_key)
            if not seller_obj:
                # seller ainda não cadastrado → impossível ser duplicata
                continue
            existing = db.query(models.Order).filter(
                models.Order.nf_number == order_data["nf_number"],
                models.Order.seller_id == seller_obj.id,
            ).first()
            if existing:
                duplicates_info.append(schemas.DuplicateOrderInfo(
                    nf_number=order_data["nf_number"],
                    seller_name=order_data["seller_name"],
                    existing_order_id=existing.id,
                    existing_session_id=existing.session_id,
                    existing_imported_at=existing.imported_at,
                ))

        if duplicates_info and not force_duplicates:
            # Aborta sem criar sessão — frontend vai perguntar ao usuário
            return schemas.ImportResult(
                success=False,
                message=(
                    f"{len(duplicates_info)} NF(s) deste arquivo já foram importadas anteriormente. "
                    f"Confirme se deseja reimportá-las."
                ),
                total_rows=total_rows,
                orders_imported=0,
                orders_with_kits=0,
                errors=[],
                warnings=warnings,
                requires_confirmation=True,
                duplicates=duplicates_info,
            )

        # Cria a sessão de picking para este lote de importação.
        # file_type e for_billing são configurações de NÍVEL DE ARQUIVO —
        # herdadas por todos os pedidos desta sessão.
        now_local = now_brasilia()
        today = now_local.date()
        session_file_type = file_type if file_type is not None else models.FileType.EXPORT
        session = models.PickingSession(
            session_date=today,
            unit_id=unit_id,
            status=models.OrderStatus.PENDING,
            total_orders=len(orders_dict) - len(ignored_order_keys),
            completed_orders=0,
            created_by_id=created_by_id,
            source_file=os.path.basename(file_path),
            file_type=session_file_type,
            for_billing=for_billing,
            created_at=now_local,
        )
        db.add(session)
        db.flush()

        # Conjunto de chaves de duplicatas para log
        duplicate_keys = {(d.nf_number, d.seller_name) for d in duplicates_info}

        # ── Caches por seller (evitam o N+1 do laço de persistência) ──────────
        # Antes: 1 query de kit + 1 query de produto POR ITEM do arquivo.
        # Agora: 2 queries por seller, na primeira vez que ele aparece.
        # Preenchidos sob demanda de propósito — um seller pode ser criado
        # dentro do próprio laço, e aí nasce sem kit e sem produto.
        kits_cache: dict = {}
        products_cache: dict = {}

        def _kits_of(sid: int) -> dict:
            """{kit_sku: Kit} dos kits ATIVOS do seller, com os itens já carregados."""
            cached = kits_cache.get(sid)
            if cached is None:
                kits = (
                    db.query(models.Kit)
                    .options(joinedload(models.Kit.items))
                    .filter(models.Kit.seller_id == sid, models.Kit.active == True)
                    .all()
                )
                # UniqueConstraint (seller_id, kit_sku) garante 1 kit por SKU
                cached = {k.kit_sku: k for k in kits}
                kits_cache[sid] = cached
            return cached

        def _products_of(sid: int) -> dict:
            """
            {sku: product_id} do seller. SEM filtro de active — reproduz
            exatamente o lookup item a item que existia aqui.
            """
            cached = products_cache.get(sid)
            if cached is None:
                cached = {}
                rows = (
                    db.query(models.Product.sku, models.Product.id)
                    .filter(models.Product.seller_id == sid)
                    .order_by(models.Product.id)
                    .all()
                )
                for _sku, _pid in rows:
                    cached.setdefault(_sku, _pid)
                products_cache[sid] = cached
            return cached

        # Persiste os pedidos
        for order_key, order_data in orders_dict.items():
            if order_key in ignored_order_keys:
                continue
            try:
                # Reaproveita o seller já resolvido (inclui reativados); se for
                # seller novo (não visto na resolução acima), cria agora.
                seller = order_key_seller_map.get(order_key) or get_or_create_seller(
                    db, order_data["seller_name"], alias_map=seller_alias_map
                )

                is_duplicate = (order_data["nf_number"], order_data["seller_name"]) in duplicate_keys
                if is_duplicate and force_duplicates:
                    # Usuário confirmou reimportação — cria novo pedido mesmo assim.
                    # Adiciona sufixo à nf_number pra não quebrar unique constraints
                    # em cima desse par (NF+seller) se houver no futuro.
                    warnings.append(
                        f"NF {order_data['nf_number']} ({order_data['seller_name']}) reimportada por confirmação do usuário"
                    )
                elif is_duplicate and not force_duplicates:
                    # Defensivo: já deveríamos ter abortado acima
                    warnings.append(
                        f"NF {order_data['nf_number']} do seller {order_data['seller_name']} já importada (ignorando)"
                    )
                    continue

                # Processa kits (expande SKUs)
                processed_items = process_order_items(
                    seller_id=seller.id,
                    raw_items=order_data["items"],
                    db=db,
                    kits_by_sku=_kits_of(seller.id),
                )

                # Conta kits expandidos
                kits_in_order = sum(1 for i in processed_items if i["is_kit_component"])
                if kits_in_order > 0:
                    kits_expanded += 1

                # file_type e for_billing são HERDADOS da sessão (nível do arquivo).
                # A natureza da NF continua sendo armazenada para auditoria/relatórios.

                # Cria o pedido — imported_at explícito com hora local
                # unit_id: usa a unidade do seller se configurada;
                # só cai no unit_id da sessão se o seller não tiver unidade associada.
                order_unit_id = seller.unit_id if seller.unit_id else unit_id
                order = models.Order(
                    erp_code=order_data["erp_code"],
                    nf_number=order_data["nf_number"],
                    customer_name=order_data["customer_name"],
                    order_date=order_data["order_date"],
                    seller_id=seller.id,
                    unit_id=order_unit_id,
                    carrier=order_data["carrier"],
                    expedition_date=order_data["expedition_date"],
                    nature=order_data["nature"],
                    danfe_key=order_data["danfe_key"],
                    nf_key=order_data["nf_key"],
                    file_type=session.file_type,
                    for_billing=session.for_billing,
                    status=models.OrderStatus.PENDING,
                    session_id=session.id,
                    imported_at=now_local,
                )
                db.add(order)
                db.flush()

                # Cria os itens (com kits já expandidos)
                seller_products = _products_of(seller.id)
                for item in processed_items:
                    # Componente de kit já vem com o product_id resolvido do cadastro.
                    # Para os demais (e para kit sem vínculo), busca pelo SKU.
                    product_id = item.get("product_id")
                    if product_id is None:
                        product_id = seller_products.get(item["sku"])

                    order_item = models.OrderItem(
                        order_id=order.id,
                        product_id=product_id,
                        sku=item["sku"],
                        product_name=item["product_name"],
                        quantity=item["quantity"],
                        is_kit_component=item["is_kit_component"],
                        original_kit_sku=item["original_kit_sku"],
                    )
                    db.add(order_item)

                orders_imported += 1

            except Exception as e:
                errors.append(f"Erro ao importar pedido {order_key}: {str(e)}")
                continue

        # Atualiza totais da sessão
        session.total_orders = orders_imported

        # ── Baixa de estoque (06/08/2026) ──────────────────────────────────
        # Acontece AQUI, dentro da mesma transação do import e ANTES do commit
        # único lá embaixo: se a baixa estourar, o `except` faz rollback e nada
        # é criado — nem sessão, nem pedido, nem movimento. Era exatamente o
        # requisito ("deve falhar, desfazer e deixar o erro explícito").
        #
        # O flush + recarga com joinedload é necessário: os OrderItem foram
        # criados com order_id (não via relationship), e a sessão usa
        # autoflush=False — sem isso, `order.items` viria VAZIO e nenhuma NF
        # baixaria estoque, em silêncio.
        db.flush()
        session_orders = db.query(models.Order).options(
            joinedload(models.Order.items),
            joinedload(models.Order.seller),
        ).filter(models.Order.session_id == session.id).all()
        stock_raw = apply_stock_for_orders(session_orders, db, operator_id=created_by_id)

        stock_report = schemas.StockApplyReport(
            applied_orders=len(stock_raw["applied"]),
            pending_orders=[schemas.PendingStockOrderInfo(**p) for p in stock_raw["pending"]],
            missing_products=[
                schemas.MissingProductInfo(**m) for m in stock_raw["missing_skus"].values()
            ],
            negatives=[schemas.NegativeStockInfo(**n) for n in stock_raw["negatives"]],
        )

        if stock_report.pending_orders:
            warnings.append(
                f"{len(stock_report.pending_orders)} NF(s) NÃO baixaram estoque — "
                f"falta transportadora ou produto cadastrado"
            )

        # Log de auditoria
        audit = models.AuditLog(
            entity_type="PickingSession",
            entity_id=session.id,
            action="IMPORT",
            detail=(
                f"Importação de {orders_imported} pedidos do arquivo {os.path.basename(file_path)}. "
                f"Kits expandidos: {kits_expanded}. "
                f"Estoque baixado em {stock_report.applied_orders} NF(s); "
                f"{len(stock_report.pending_orders)} pendente(s)."
            ),
            user_id=created_by_id,
        )
        db.add(audit)
        db.commit()

        return schemas.ImportResult(
            success=True,
            message=f"Importação concluída: {orders_imported} pedidos importados",
            session_id=session.id,
            total_rows=total_rows,
            orders_imported=orders_imported,
            orders_with_kits=kits_expanded,
            errors=errors,
            warnings=warnings,
            stock=stock_report,
        )

    except Exception as e:
        db.rollback()
        return schemas.ImportResult(
            success=False,
            message=f"Erro na importação: {str(e)}",
            total_rows=0,
            orders_imported=0,
            errors=[str(e)] + errors,
            warnings=warnings,
        )
